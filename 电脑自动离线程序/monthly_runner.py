# -*- coding: utf-8 -*-
"""Wizard-style local monthly Excel automation launcher."""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path
from typing import Any

import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox


LAUNCHER_DIR = Path(__file__).resolve().parent
PROJECT_DIR = LAUNCHER_DIR.parent
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))

from dingbian_classifier.exceptions import ClassifierError
from monthly_automation import (
    DEFAULT_INPUT_DIR,
    DEFAULT_LOG_DIR,
    create_run_log_path,
    ensure_input_dir,
    load_monthly_flow,
    run_monthly_flow,
    run_stage,
    validate_input_file,
    write_exception_log,
    write_log,
)


CONFIG_PATH = LAUNCHER_DIR / "monthly_flow.json"
LOG_DIR = DEFAULT_LOG_DIR
STATE_PATH = LAUNCHER_DIR / "wizard_state.json"
COEFFICIENT_PENDING_SHEETS = ["系数仍缺失", "系数补充"]
SHEET_METAL_PENDING_SHEETS = ["钣金型号补充"]
MATERIAL_DESCRIPTION_PENDING_SHEETS = ["物料描述仍缺失", "物料描述补充"]


def choose_excel_file(title: str, initial_dir: Path, *, include_text: bool = False) -> Path | None:
    filetypes = [("Excel 文件", "*.xlsx *.xlsm *.xls")]
    if include_text:
        filetypes.append(("查询表文本导出", "*.txt *.csv *.tsv *.xls"))
    filetypes.append(("所有文件", "*.*"))
    selected = filedialog.askopenfilename(
        title=title,
        initialdir=str(initial_dir),
        filetypes=filetypes,
    )
    if not selected:
        return None
    return Path(selected)


def choose_input_file() -> Path | None:
    return choose_excel_file("请选择本次月度定编 Excel 文件", DEFAULT_INPUT_DIR)


def choose_coefficient_lookup() -> Path | None:
    return choose_excel_file("请选择系数查询表", DEFAULT_INPUT_DIR, include_text=True)


def choose_sheet_metal_lookup() -> Path | None:
    selected = filedialog.askopenfilename(
        title="请选择钣金型号查询表",
        initialdir=str(DEFAULT_INPUT_DIR),
        filetypes=[
            ("Excel 文件", "*.xlsx *.xlsm"),
            ("所有文件", "*.*"),
        ],
    )
    if not selected:
        return None
    return Path(selected)


def choose_sheet_metal_bom_lookup() -> Path | None:
    selected = filedialog.askopenfilename(
        title="请选择钣金型号 BOM 表",
        initialdir=str(DEFAULT_INPUT_DIR),
        filetypes=[
            ("Excel 文件", "*.xlsx *.xlsm *.xls"),
            ("所有文件", "*.*"),
        ],
    )
    if not selected:
        return None
    return Path(selected)


def choose_material_description_lookup() -> Path | None:
    return choose_excel_file("请选择物料描述查询表", DEFAULT_INPUT_DIR, include_text=True)


def load_state() -> dict[str, Any] | None:
    if not STATE_PATH.exists():
        return None
    with STATE_PATH.open("r", encoding="utf-8") as state_file:
        return json.load(state_file)


def save_state(state: dict[str, Any]) -> None:
    with STATE_PATH.open("w", encoding="utf-8") as state_file:
        json.dump(state, state_file, ensure_ascii=False, indent=2)


def clear_state() -> None:
    if STATE_PATH.exists():
        STATE_PATH.unlink()


def make_state(step: str, current_file: Path, log_path: Path, output_dir: Path) -> dict[str, Any]:
    return {
        "step": step,
        "current_file": str(current_file),
        "log_path": str(log_path),
        "output_dir": str(output_dir),
    }


def get_state_path(state: dict[str, Any], key: str) -> Path:
    value = state.get(key)
    if not value:
        raise ValueError(f"向导状态缺少字段：{key}")
    return Path(value)


def sheet_has_data_rows(workbook_path: Path, sheet_names: list[str]) -> bool:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        for sheet_name in sheet_names:
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            if sheet.max_row >= 2:
                return True
        return False
    finally:
        workbook.close()


def has_coefficient_pending(workbook_path: Path) -> bool:
    return sheet_has_data_rows(workbook_path, COEFFICIENT_PENDING_SHEETS)


def has_sheet_metal_pending(workbook_path: Path) -> bool:
    return sheet_has_data_rows(workbook_path, SHEET_METAL_PENDING_SHEETS)


def has_material_description_pending(workbook_path: Path) -> bool:
    return sheet_has_data_rows(workbook_path, MATERIAL_DESCRIPTION_PENDING_SHEETS)


def next_foundation_step(workbook_path: Path) -> str:
    if has_coefficient_pending(workbook_path):
        return "choose_coefficient_lookup"
    if has_sheet_metal_pending(workbook_path):
        return "choose_sheet_metal_lookup"
    if has_material_description_pending(workbook_path):
        return "choose_material_description_lookup"
    return "final_stages"


def open_output_dir(output_dir: Path, log_path: Path) -> None:
    output_dir.mkdir(exist_ok=True)
    try:
        os.startfile(str(output_dir))
        write_log(log_path, f"已打开输出目录：{output_dir}")
    except OSError as exc:
        write_log(log_path, f"打开输出目录失败：{exc}")


def open_workbook(path: Path, log_path: Path) -> None:
    try:
        os.startfile(str(path))
        write_log(log_path, f"已打开工作簿：{path}")
    except OSError as exc:
        write_log(log_path, f"打开工作簿失败：{exc}")


def show_error(title: str, exc: BaseException) -> None:
    messagebox.showerror(title, f"{type(exc).__name__}: {exc}")


def pause_for_lookup(
    *,
    state: dict[str, Any],
    message: str,
    log_path: Path,
) -> Path | None:
    save_state(state)
    write_log(log_path, message)
    messagebox.showinfo("流程暂停", message)
    return None


def pause_for_manual(
    *,
    state: dict[str, Any],
    workbook_path: Path,
    output_dir: Path,
    log_path: Path,
    sheets: str,
) -> Path | None:
    save_state(state)
    open_workbook(workbook_path, log_path)
    open_output_dir(output_dir, log_path)
    message = (
        f"自动查询后仍有数据需要手工补充。\n\n"
        f"当前结果文件：\n{workbook_path}\n\n"
        f"请打开并填写工作表：{sheets}\n"
        f"填写完成后请保存并关闭 Excel，然后再次双击“启动月度处理.bat”继续。"
    )
    write_log(log_path, message.replace("\n", " "))
    messagebox.showinfo("需要手工补充", message)
    return None


def continue_prompt(state: dict[str, Any]) -> bool:
    current_file = state.get("current_file", "")
    return messagebox.askyesno(
        "继续上次流程",
        f"检测到上次流程尚未完成。\n\n当前文件：\n{current_file}\n\n是否继续上次流程？",
    )


def run_wizard(state: dict[str, Any]) -> Path | None:
    step = state["step"]
    current_file = get_state_path(state, "current_file")
    output_dir = get_state_path(state, "output_dir")
    log_path = get_state_path(state, "log_path")
    stages, _configured_output_dir = load_monthly_flow(CONFIG_PATH)

    while True:
        write_log(log_path, f"向导步骤：{step}")

        if step == "prepare_standard_units":
            step = "final_stages"
            state = make_state(step, current_file, log_path, output_dir)
            continue

        if step in {
            "prepare_foundation_data",
            "prepare_coefficients",
            "prepare_sheet_metal",
            "prepare_material_description",
        }:
            current_file = run_stage(current_file, output_dir, "prepare-foundation-data", log_path)
            step = next_foundation_step(current_file)
            state = make_state(step, current_file, log_path, output_dir)
            continue

        if step == "choose_coefficient_lookup":
            lookup = choose_coefficient_lookup()
            if lookup is None:
                return pause_for_lookup(
                    state=make_state(step, current_file, log_path, output_dir),
                    message="已暂停：请下次继续时选择系数查询表。",
                    log_path=log_path,
                )
            current_file = run_stage(
                current_file,
                output_dir,
                "fill-coefficients",
                log_path,
                coefficient_lookup=lookup,
            )
            if has_coefficient_pending(current_file):
                return pause_for_manual(
                    state=make_state("wait_manual_coefficients", current_file, log_path, output_dir),
                    workbook_path=current_file,
                    output_dir=output_dir,
                    log_path=log_path,
                    sheets="系数补充 / 系数仍缺失",
                )
            step = next_foundation_step(current_file)
            state = make_state(step, current_file, log_path, output_dir)
            continue

        if step == "wait_manual_coefficients":
            current_file = run_stage(current_file, output_dir, "apply-manual-coefficients", log_path)
            if has_coefficient_pending(current_file):
                return pause_for_manual(
                    state=make_state("wait_manual_coefficients", current_file, log_path, output_dir),
                    workbook_path=current_file,
                    output_dir=output_dir,
                    log_path=log_path,
                    sheets="系数补充 / 系数仍缺失",
                )
            step = next_foundation_step(current_file)
            state = make_state(step, current_file, log_path, output_dir)
            continue

        if step == "choose_sheet_metal_lookup":
            lookup = choose_sheet_metal_lookup()
            if lookup is None:
                use_bom = messagebox.askyesno(
                    "钣金型号查询表未选择",
                    "没有选择钣金型号查询表。\n\n是否改为选择 BOM 表，先生成钣金型号候选供你确认？",
                )
                if use_bom:
                    step = "choose_sheet_metal_bom_lookup"
                    state = make_state(step, current_file, log_path, output_dir)
                    continue
                return pause_for_lookup(
                    state=make_state(step, current_file, log_path, output_dir),
                    message="已暂停：请下次继续时选择钣金型号查询表，或选择 BOM 表生成候选。",
                    log_path=log_path,
                )
            current_file = run_stage(
                current_file,
                output_dir,
                "fill-sheet-metal",
                log_path,
                sheet_metal_lookup=lookup,
            )
            if has_sheet_metal_pending(current_file):
                use_bom = messagebox.askyesno(
                    "钣金型号仍有缺失",
                    "钣金型号查询表仍有覆盖不到的行。\n\n是否选择 BOM 表，先把箱体组件候选写入“钣金型号补充”供你确认？",
                )
                if use_bom:
                    step = "choose_sheet_metal_bom_lookup"
                    state = make_state(step, current_file, log_path, output_dir)
                    continue
                return pause_for_manual(
                    state=make_state("wait_manual_sheet_metal", current_file, log_path, output_dir),
                    workbook_path=current_file,
                    output_dir=output_dir,
                    log_path=log_path,
                    sheets="钣金型号补充",
                )
            step = next_foundation_step(current_file)
            state = make_state(step, current_file, log_path, output_dir)
            continue

        if step == "choose_sheet_metal_bom_lookup":
            lookup = choose_sheet_metal_bom_lookup()
            if lookup is None:
                return pause_for_lookup(
                    state=make_state(step, current_file, log_path, output_dir),
                    message="已暂停：请下次继续时选择钣金型号 BOM 表，或回到钣金型号补充表手工填写。",
                    log_path=log_path,
                )
            current_file = run_stage(
                current_file,
                output_dir,
                "suggest-sheet-metal-bom",
                log_path,
                sheet_metal_bom_lookup=lookup,
            )
            return pause_for_manual(
                state=make_state("wait_manual_sheet_metal", current_file, log_path, output_dir),
                workbook_path=current_file,
                output_dir=output_dir,
                log_path=log_path,
                sheets="钣金型号补充（请确认 BOM 候选，必要时手工修改）",
            )

        if step == "wait_manual_sheet_metal":
            current_file = run_stage(current_file, output_dir, "apply-manual-sheet-metal", log_path)
            if has_sheet_metal_pending(current_file):
                return pause_for_manual(
                    state=make_state("wait_manual_sheet_metal", current_file, log_path, output_dir),
                    workbook_path=current_file,
                    output_dir=output_dir,
                    log_path=log_path,
                    sheets="钣金型号补充",
                )
            step = next_foundation_step(current_file)
            state = make_state(step, current_file, log_path, output_dir)
            continue

        if step == "choose_material_description_lookup":
            lookup = choose_material_description_lookup()
            if lookup is None:
                return pause_for_lookup(
                    state=make_state(step, current_file, log_path, output_dir),
                    message="已暂停：请下次继续时选择物料描述查询表。",
                    log_path=log_path,
                )
            current_file = run_stage(
                current_file,
                output_dir,
                "fill-material-description",
                log_path,
                material_description_lookup=lookup,
            )
            if has_material_description_pending(current_file):
                return pause_for_manual(
                    state=make_state("wait_manual_material_description", current_file, log_path, output_dir),
                    workbook_path=current_file,
                    output_dir=output_dir,
                    log_path=log_path,
                    sheets="物料描述补充 / 物料描述仍缺失",
                )
            step = next_foundation_step(current_file)
            state = make_state(step, current_file, log_path, output_dir)
            continue

        if step == "wait_manual_material_description":
            current_file = run_stage(current_file, output_dir, "apply-manual-material-description", log_path)
            if has_material_description_pending(current_file):
                return pause_for_manual(
                    state=make_state("wait_manual_material_description", current_file, log_path, output_dir),
                    workbook_path=current_file,
                    output_dir=output_dir,
                    log_path=log_path,
                    sheets="物料描述补充 / 物料描述仍缺失",
                )
            step = next_foundation_step(current_file)
            state = make_state(step, current_file, log_path, output_dir)
            continue

        if step == "final_stages":
            current_file = run_monthly_flow(current_file, stages, output_dir, log_path)
            clear_state()
            return current_file

        raise ValueError(f"未知向导步骤：{step}")


def main() -> int:
    root = tk.Tk()
    root.withdraw()

    try:
        ensure_input_dir()
        existing_state = load_state()
        if existing_state and continue_prompt(existing_state):
            state = existing_state
            log_path = get_state_path(state, "log_path")
            write_log(log_path, "继续上次月度向导流程")
        else:
            if existing_state:
                clear_state()
            log_path = create_run_log_path(LOG_DIR)
            stages, output_dir = load_monthly_flow(CONFIG_PATH)
            write_log(log_path, "月度离线向导启动")
            write_log(log_path, f"项目目录：{PROJECT_DIR}")
            write_log(log_path, f"启动目录：{LAUNCHER_DIR}")
            write_log(log_path, f"最终流程：{' -> '.join(stages)}")
            input_path = choose_input_file()
            if input_path is None:
                write_log(log_path, "用户取消选择文件，程序退出。")
                root.destroy()
                return 0
            validate_input_file(input_path)
            state = make_state("prepare_foundation_data", input_path, log_path, output_dir)

        output_path = run_wizard(state)
    except (ClassifierError, FileNotFoundError, ValueError, json.JSONDecodeError) as exc:
        log_path = Path(locals()["log_path"]) if "log_path" in locals() else create_run_log_path(LOG_DIR)
        write_log(log_path, f"处理失败：{exc}")
        write_exception_log(log_path, exc)
        show_error("月度处理失败", exc)
        root.destroy()
        return 1
    except Exception as exc:
        log_path = Path(locals()["log_path"]) if "log_path" in locals() else create_run_log_path(LOG_DIR)
        write_log(log_path, f"未知错误：{exc}")
        write_exception_log(log_path, exc)
        show_error("月度处理发生未知错误", exc)
        root.destroy()
        return 1

    if output_path is None:
        root.destroy()
        return 0

    output_dir = output_path.parent
    write_log(log_path, f"月度离线向导完成：{output_path}")
    open_output_dir(output_dir, log_path)
    messagebox.showinfo(
        "月度处理完成",
        f"全部流程已完成。\n\n输出文件：\n{output_path}\n\n日志文件：\n{log_path}",
    )
    root.destroy()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
