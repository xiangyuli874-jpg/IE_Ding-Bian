#!/usr/bin/env python
"""Read-only inspection for Dingbian production-plan workbooks."""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl


TARGET_PATTERN = re.compile(r"^W\d+\s*-\s*W\d+.*周排产明细")
SUPPORTED_SUFFIXES = {".xlsx", ".xlsm"}
PENDING_GROUPS = {
    "coefficient_pending": ("系数仍缺失", "系数补充"),
    "sheet_metal_pending": ("钣金型号补充",),
    "material_description_pending": ("物料描述仍缺失", "物料描述补充"),
}
DECOMPOSITION_SHEETS = ("排单分解表明细", "各线体分类明细表")
CLASSIFICATION_SHEETS = ("分类结果汇总表", "未分类数据")
FINAL_SHEETS = (*DECOMPOSITION_SHEETS, *CLASSIFICATION_SHEETS)
TYPE_COLUMN = "类型"
STANDARD_UNITS_COLUMN = "标台数"
DECOMPOSITION_DETAIL_SHEET = "排单分解表明细"
EXTRA_SUMMARY_METRICS = (
    "外协烘道数量",
    "滚筒喷粉数量",
    "波轮喷粉数量",
    "PCM板中需喷涂前门板的箱体数量",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="只读检查定编工作簿状态并输出 JSON")
    parser.add_argument("workbook", type=Path, help="待检查的 .xlsx 或 .xlsm 文件")
    return parser.parse_args()


def fail(message: str, exit_code: int = 2) -> int:
    print(json.dumps({"ok": False, "error": message}, ensure_ascii=False, indent=2))
    return exit_code


def count_data_rows(sheet: Any) -> int:
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(value is not None and str(value).strip() for value in row):
            count += 1
    return count


def pending_count(workbook: Any, sheet_names: tuple[str, ...]) -> int:
    return max(
        (count_data_rows(workbook[name]) for name in sheet_names if name in workbook.sheetnames),
        default=0,
    )


def header_values(sheet: Any) -> list[str]:
    headers: list[str] = []
    for cell in next(sheet.iter_rows(min_row=1, max_row=1), []):
        value = cell.value
        headers.append(str(value).strip() if value is not None else "")
    return headers


def sheet_contains_text(sheet: Any, text: str) -> bool:
    for row in sheet.iter_rows(values_only=True):
        for value in row:
            if value is not None and text in str(value):
                return True
    return False


def main() -> int:
    args = parse_args()
    raw_path = args.workbook
    if raw_path.name.startswith("~$"):
        return fail("不能检查 Excel 临时锁文件；请关闭 Excel 后使用原始文件。")
    path = raw_path.expanduser().resolve()
    warnings: list[str] = []

    if not path.exists():
        return fail(f"文件不存在：{path}")
    if path.suffix.lower() not in SUPPORTED_SUFFIXES:
        return fail("主工作簿仅支持 .xlsx 或 .xlsm。")
    lock_path = path.with_name(f"~${path.name}")
    if lock_path.exists():
        warnings.append(f"发现 Excel 临时锁文件：{lock_path}；如需继续写入阶段，请先保存并关闭 Excel。")

    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return fail(f"无法打开工作簿：{type(exc).__name__}: {exc}")

    try:
        target_sheets = [name for name in workbook.sheetnames if TARGET_PATTERN.search(name)]
        main_sheet = workbook[target_sheets[0]] if len(target_sheets) == 1 else None
        headers = header_values(main_sheet) if main_sheet is not None else []
        result: dict[str, Any] = {
            "ok": len(target_sheets) == 1,
            "path": str(path),
            "target_sheets": target_sheets,
            "target_sheet": target_sheets[0] if len(target_sheets) == 1 else None,
            "target_data_rows": (
                count_data_rows(main_sheet) if main_sheet is not None else None
            ),
            "sheet_count": len(workbook.sheetnames),
            "has_type_column": TYPE_COLUMN in headers,
            "has_standard_units_column": STANDARD_UNITS_COLUMN in headers,
        }

        for key, sheet_names in PENDING_GROUPS.items():
            result[key] = pending_count(workbook, sheet_names)

        result["unclassified_rows"] = (
            count_data_rows(workbook["未分类数据"]) if "未分类数据" in workbook.sheetnames else 0
        )
        result["foundation_ready"] = (
            result["coefficient_pending"] == 0
            and result["sheet_metal_pending"] == 0
            and result["material_description_pending"] == 0
        )
        result["decomposition_sheets"] = {
            name: name in workbook.sheetnames for name in DECOMPOSITION_SHEETS
        }
        result["classification_sheets"] = {
            name: name in workbook.sheetnames for name in CLASSIFICATION_SHEETS
        }
        result["final_sheets"] = {
            name: name in workbook.sheetnames for name in FINAL_SHEETS
        }
        result["decomposition_ready"] = all(result["decomposition_sheets"].values())
        result["classification_ready"] = all(result["classification_sheets"].values())
        if DECOMPOSITION_DETAIL_SHEET in workbook.sheetnames:
            detail_sheet = workbook[DECOMPOSITION_DETAIL_SHEET]
            result["extra_summary_metrics"] = {
                metric: sheet_contains_text(detail_sheet, metric)
                for metric in EXTRA_SUMMARY_METRICS
            }
        else:
            result["extra_summary_metrics"] = {
                metric: False for metric in EXTRA_SUMMARY_METRICS
            }
        result["extra_summary_ready"] = all(result["extra_summary_metrics"].values())
        result["warnings"] = warnings

        if not target_sheets:
            result["error"] = "未找到符合命名规则的周排产明细主工作表。"
            result["sheet_names"] = workbook.sheetnames
        elif len(target_sheets) > 1:
            result["error"] = "找到多个候选主工作表，需要用户指定。"

        if not result["ok"]:
            result["next_action"] = "fix_workbook_selection"
        elif not result["foundation_ready"]:
            result["next_action"] = "resolve_foundation_data"
            if result["decomposition_ready"] or result["classification_ready"]:
                result["warnings"].append(
                    "当前文件已有下游结果表但基础数据仍有待补；补齐后应重新执行受影响的标准台数、格式、分解、额外汇总和 classify。"
                )
        elif not result["has_standard_units_column"]:
            result["next_action"] = "prepare_standard_units"
        elif not result["has_type_column"] or not result["decomposition_ready"]:
            result["next_action"] = "run_standard_format_and_decomposition"
        elif not result["extra_summary_ready"]:
            result["next_action"] = "run_decompose_extra_summary"
            result["warnings"].append("额外订单信息汇总缺少最新产能规划指标；请运行 decompose-extra-summary 后再 classify。")
        elif not result["classification_ready"]:
            result["next_action"] = "run_classify"
        elif result["unclassified_rows"] > 0:
            result["next_action"] = "review_unclassified_data"
        else:
            result["next_action"] = "complete"
        if "分类结果" in path.stem and not result["classification_ready"]:
            result["warnings"].append("文件名包含“分类结果”，但工作簿缺少“分类结果汇总表/未分类数据”；不要仅凭文件名判断定编完成。")
        result["is_complete"] = result["next_action"] == "complete"

        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0 if result["ok"] else 3
    finally:
        workbook.close()


if __name__ == "__main__":
    sys.exit(main())
