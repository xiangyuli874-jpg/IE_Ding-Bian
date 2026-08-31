"""Standard-unit and sheet-metal-model preparation helpers."""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .data_cleaner import deduplicate_headers, normalize_header
from .excel_io import remove_sheet_if_exists
from .exceptions import MissingRequiredFieldsError
from .logger import ProcessingLogger
from .reporter import style_header

STANDARD_UNITS_FIELD = "标台数"
SHEET_METAL_SUPPLEMENT_SHEET = "钣金型号补充"
SHEET_METAL_LOOKUP_SHEET = "钣金型号查询表"
SHEET_METAL_BOM_LOOKUP_SHEET = "钣金型号BOM查询表"
WAVE_LINES = {"B线", "B线夜", "C线", "C线夜"}
LOOKUP_CODE_FIELDS = ("物料编码", "新总装编码")
LOOKUP_MODEL_FIELDS = ("钣金型号", "半成品机型")
BOM_CODE_FIELD = "BOM号"
BOM_MODEL_FIELD = "物料描述"
BOM_MODEL_KEYWORD = "箱体组件"


@dataclass
class SheetMetalPrepareResult:
    standard_units_formulas: int
    sheet_metal_missing_rows: int


@dataclass
class SheetMetalFillResult:
    filled_rows: int
    remaining_rows: int
    lookup_rows: int


@dataclass
class ManualSheetMetalApplyResult:
    applied_rows: int
    remaining_rows: int


def prepare_sheet_metal(
    workbook: Workbook,
    values_workbook: Workbook,
    target_sheet_name: str,
    logger: ProcessingLogger,
) -> SheetMetalPrepareResult:
    formula_sheet = workbook[target_sheet_name]
    values_sheet = values_workbook[target_sheet_name]

    missing_count = create_sheet_metal_supplement_sheet(
        workbook,
        formula_sheet,
        values_sheet,
        logger,
    )
    return SheetMetalPrepareResult(
        standard_units_formulas=0,
        sheet_metal_missing_rows=missing_count,
    )


def ensure_standard_units_column(sheet: Worksheet, logger: ProcessingLogger) -> int:
    headers = _header_map(sheet)
    _require_columns(headers, ["系数", "订单数"], "主数据表")

    if STANDARD_UNITS_FIELD not in headers:
        insert_col = headers["订单数"]
        sheet.insert_cols(insert_col, 1)
        sheet.cell(1, insert_col).value = STANDARD_UNITS_FIELD
        logger.info(f"已在“订单数”列前新增“{STANDARD_UNITS_FIELD}”列。")
    else:
        logger.info(f"检测到已存在“{STANDARD_UNITS_FIELD}”列，将刷新公式和格式。")

    headers = _header_map(sheet)
    _require_columns(headers, ["系数", STANDARD_UNITS_FIELD, "订单数"], "主数据表")
    coefficient_col = headers["系数"]
    standard_col = headers[STANDARD_UNITS_FIELD]
    order_col = headers["订单数"]

    coefficient_letter = get_column_letter(coefficient_col)
    order_letter = get_column_letter(order_col)
    for row_index in range(2, sheet.max_row + 1):
        cell = sheet.cell(row_index, standard_col)
        cell.value = f"={coefficient_letter}{row_index}*{order_letter}{row_index}"
        cell.number_format = "0.00"
    sheet.cell(1, standard_col).number_format = "General"
    _style_standard_units_column(sheet, standard_col)

    logger.info(f"已刷新“{STANDARD_UNITS_FIELD}”公式：{max(sheet.max_row - 1, 0)} 行。")
    return max(sheet.max_row - 1, 0)


def refresh_sheet_metal_lookup_formulas(sheet: Worksheet, logger: ProcessingLogger) -> int:
    headers = _header_map(sheet)
    if "钣金型号" not in headers or "物料编码" not in headers:
        logger.warning("未刷新钣金型号查找公式：主数据表缺少 钣金型号 或 物料编码。")
        return 0

    sheet_metal_col = headers["钣金型号"]
    code_col = headers["物料编码"]
    code_letter = get_column_letter(code_col)
    refreshed = 0
    for row_index in range(2, sheet.max_row + 1):
        cell = sheet.cell(row_index, sheet_metal_col)
        value = cell.value
        if not (isinstance(value, str) and value.strip().upper().startswith("=VLOOKUP(")):
            continue
        new_formula = f"=VLOOKUP({code_letter}{row_index},钣金型号!A:C,3,0)"
        if value != new_formula:
            cell.value = new_formula
            refreshed += 1

    logger.info(f"已刷新钣金型号列 VLOOKUP 当前行引用：{refreshed} 行。")
    return refreshed


def _style_standard_units_column(sheet: Worksheet, standard_col: int) -> None:
    blue_fill = PatternFill(fill_type="solid", fgColor="0070C0")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_cell = sheet.cell(1, standard_col)
    header_cell.fill = blue_fill
    header_cell.font = Font(color="FFFFFF", bold=True, size=14)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    header_cell.border = border

    for row_index in range(2, sheet.max_row + 1):
        cell = sheet.cell(row_index, standard_col)
        cell.fill = PatternFill(fill_type=None)
        cell.font = Font(color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        cell.number_format = "0.00"


def refresh_daily_output_formulas(sheet: Worksheet, logger: ProcessingLogger) -> None:
    headers = _header_map(sheet)
    required = ["日产", "订单数", "线体", "基本开始日期"]
    if any(name not in headers for name in required):
        logger.warning("未刷新“日产”公式：缺少 日产/订单数/线体/基本开始日期 中的一个字段。")
        return

    daily_col = headers["日产"]
    order_col = headers["订单数"]
    line_col = headers["线体"]
    date_col = headers["基本开始日期"]
    order_letter = get_column_letter(order_col)
    line_letter = get_column_letter(line_col)
    date_letter = get_column_letter(date_col)

    for row_index in range(2, sheet.max_row + 1):
        sheet.cell(row_index, daily_col).value = (
            f"=SUMIFS({order_letter}:{order_letter},"
            f"{line_letter}:{line_letter},{line_letter}{row_index},"
            f"{date_letter}:{date_letter},{date_letter}{row_index})"
        )
    logger.info("已同步刷新“日产”列公式，确保继续引用移动后的“订单数”列。")


def create_sheet_metal_supplement_sheet(
    workbook: Workbook,
    formula_sheet: Worksheet,
    values_sheet: Worksheet,
    logger: ProcessingLogger,
) -> int:
    values_headers = _header_map(values_sheet)
    _require_columns(values_headers, ["钣金型号"], "主数据表")
    sheet_metal_col = values_headers["钣金型号"]

    remove_sheet_if_exists(workbook, SHEET_METAL_SUPPLEMENT_SHEET)
    supplement = workbook.create_sheet(SHEET_METAL_SUPPLEMENT_SHEET)
    output_headers = ["原始行号"] + [
        normalize_header(cell.value) or f"空列{index}"
        for index, cell in enumerate(formula_sheet[1], start=1)
    ]
    supplement.append(output_headers)

    missing_count = 0
    for row_index in range(2, values_sheet.max_row + 1):
        value = values_sheet.cell(row_index, sheet_metal_col).value
        if not (_is_blank(value) or _is_na(value)):
            continue
        missing_count += 1
        row_values = [
            formula_sheet.cell(row_index, col_index).value
            for col_index in range(1, formula_sheet.max_column + 1)
        ]
        supplement.append([row_index] + row_values)

    supplement.freeze_panes = "A2"
    if supplement.max_row > 1:
        supplement.auto_filter.ref = supplement.dimensions
    style_header(supplement)
    _fit_columns(supplement)
    logger.info(f"钣金型号为 #N/A 的行已摘取到“{SHEET_METAL_SUPPLEMENT_SHEET}”：{missing_count} 行。")
    return missing_count


def fill_sheet_metal_models(
    workbook: Workbook,
    values_workbook: Workbook,
    target_sheet_name: str,
    lookup_path,
    logger: ProcessingLogger,
) -> SheetMetalFillResult:
    lookup_pairs = read_sheet_metal_lookup(lookup_path, logger)
    lookup_map: dict[Any, Any] = {}
    conflicts = 0
    for code, model in lookup_pairs:
        if _is_blank(code) or _is_blank(model) or _is_na(model):
            continue
        code_key = _lookup_key(code)
        if _is_blank(code_key):
            continue
        if code_key in lookup_map and lookup_map[code_key] != model:
            conflicts += 1
            continue
        lookup_map[code_key] = model

    import_sheet_metal_lookup_sheet(workbook, lookup_pairs, logger)

    formula_sheet = workbook[target_sheet_name]
    values_sheet = values_workbook[target_sheet_name]
    headers = _header_map(values_sheet)
    _require_columns(headers, ["钣金型号", "物料编码"], "主数据表")
    sheet_metal_col = headers["钣金型号"]
    code_col = headers["物料编码"]

    remaining_rows: list[int] = []
    filled_rows = 0
    for row_index in range(2, values_sheet.max_row + 1):
        current_value = values_sheet.cell(row_index, sheet_metal_col).value
        if not _is_na(current_value):
            continue
        code = values_sheet.cell(row_index, code_col).value
        matched_model = lookup_map.get(_lookup_key(code))
        if matched_model is None:
            remaining_rows.append(row_index)
            continue
        formula_sheet.cell(row_index, sheet_metal_col).value = matched_model
        filled_rows += 1

    refresh_sheet_metal_supplement_sheet(workbook, formula_sheet, remaining_rows, logger)
    if conflicts:
        logger.warning(f"钣金型号查询表存在同一物料编码对应多个钣金型号的情况，已保留首次匹配：{conflicts} 条。")
    logger.info(
        f"钣金型号已按物料编码回填：{filled_rows} 行；仍待补充 {len(remaining_rows)} 行。"
    )
    return SheetMetalFillResult(
        filled_rows=filled_rows,
        remaining_rows=len(remaining_rows),
        lookup_rows=len(lookup_pairs),
    )


def fill_sheet_metal_models_from_bom(
    workbook: Workbook,
    values_workbook: Workbook,
    target_sheet_name: str,
    lookup_path,
    logger: ProcessingLogger,
) -> SheetMetalFillResult:
    lookup_pairs = read_sheet_metal_bom_lookup(lookup_path, logger)
    lookup_map: dict[str, Any] = {}
    conflicts = 0
    for code, model in lookup_pairs:
        code_key = _lookup_key(code)
        if _is_blank(code_key) or _is_blank(model) or _is_na(model):
            continue
        if code_key in lookup_map and lookup_map[code_key] != model:
            conflicts += 1
            continue
        lookup_map[code_key] = model

    import_sheet_metal_bom_lookup_sheet(workbook, lookup_pairs, logger)

    formula_sheet = workbook[target_sheet_name]
    values_sheet = values_workbook[target_sheet_name]
    headers = _header_map(values_sheet)
    _require_columns(headers, ["钣金型号", "物料编码", "线体"], "主数据表")
    sheet_metal_col = headers["钣金型号"]
    code_col = headers["物料编码"]
    line_col = headers["线体"]

    remaining_rows: list[int] = []
    filled_rows = 0
    skipped_line_mismatch_rows = 0
    for row_index in range(2, values_sheet.max_row + 1):
        current_value = values_sheet.cell(row_index, sheet_metal_col).value
        if not _is_na(current_value):
            continue
        line = str(values_sheet.cell(row_index, line_col).value or "").strip()
        code = values_sheet.cell(row_index, code_col).value
        matched_model = lookup_map.get(_lookup_key(code))
        if matched_model is None:
            remaining_rows.append(row_index)
            continue
        if not _bom_model_matches_line(line, matched_model):
            remaining_rows.append(row_index)
            skipped_line_mismatch_rows += 1
            continue
        formula_sheet.cell(row_index, sheet_metal_col).value = matched_model
        filled_rows += 1

    refresh_sheet_metal_supplement_sheet(workbook, formula_sheet, remaining_rows, logger)
    if conflicts:
        logger.warning(f"钣金型号BOM表存在同一BOM号对应多个箱体组件的情况，已保留首次匹配：{conflicts} 条。")
    logger.info(
        f"钣金型号已按BOM箱体组件回填：{filled_rows} 行；仍待补充 {len(remaining_rows)} 行。"
    )
    if skipped_line_mismatch_rows:
        logger.warning(
            f"已跳过 {skipped_line_mismatch_rows} 行疑似线体类型不匹配的钣金型号BOM回填；"
            "BOM只用于当前钣金型号缺失/异常行，不覆盖已有正常钣金型号。"
        )
    return SheetMetalFillResult(
        filled_rows=filled_rows,
        remaining_rows=len(remaining_rows),
        lookup_rows=len(lookup_pairs),
    )


def suggest_sheet_metal_models_from_bom(
    workbook: Workbook,
    values_workbook: Workbook,
    target_sheet_name: str,
    lookup_path,
    logger: ProcessingLogger,
) -> SheetMetalFillResult:
    """Write BOM matches into the supplement sheet only, leaving the main sheet untouched."""
    lookup_pairs = read_sheet_metal_bom_lookup(lookup_path, logger)
    lookup_map: dict[str, Any] = {}
    conflicts = 0
    for code, model in lookup_pairs:
        code_key = _lookup_key(code)
        if _is_blank(code_key) or _is_blank(model) or _is_na(model):
            continue
        if code_key in lookup_map and lookup_map[code_key] != model:
            conflicts += 1
            continue
        lookup_map[code_key] = model

    import_sheet_metal_bom_lookup_sheet(workbook, lookup_pairs, logger)

    formula_sheet = workbook[target_sheet_name]
    values_sheet = values_workbook[target_sheet_name]
    if SHEET_METAL_SUPPLEMENT_SHEET not in workbook.sheetnames:
        create_sheet_metal_supplement_sheet(workbook, formula_sheet, values_sheet, logger)

    supplement = workbook[SHEET_METAL_SUPPLEMENT_SHEET]
    headers = _header_map(supplement)
    _require_columns(headers, ["物料编码", "钣金型号"], SHEET_METAL_SUPPLEMENT_SHEET)
    code_col = headers["物料编码"]
    model_col = headers["钣金型号"]
    line_col = headers.get("线体")
    _ensure_supplement_column_before(supplement, "BOM匹配说明", "钣金型号")
    _ensure_supplement_column_before(supplement, "BOM箱体组件原文", "钣金型号")
    headers = _header_map(supplement)
    code_col = headers["物料编码"]
    model_col = headers["钣金型号"]
    line_col = headers.get("线体")
    note_col = headers["BOM匹配说明"]
    source_col = headers["BOM箱体组件原文"]

    suggested_rows = 0
    skipped_existing_rows = 0
    skipped_line_mismatch_rows = 0
    unmatched_rows = 0
    candidate_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    mismatch_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")

    for row_index in range(2, supplement.max_row + 1):
        current_value = supplement.cell(row_index, model_col).value
        if _is_valid_manual_model(current_value):
            skipped_existing_rows += 1
            continue

        code = supplement.cell(row_index, code_col).value
        matched_model = lookup_map.get(_lookup_key(code))
        if matched_model is None:
            supplement.cell(row_index, note_col).value = "BOM未匹配"
            unmatched_rows += 1
            continue

        line = str(supplement.cell(row_index, line_col).value or "").strip() if line_col else ""
        if not _bom_model_matches_line(line, matched_model):
            supplement.cell(row_index, note_col).value = "疑似线体类型不匹配，未自动写入"
            supplement.cell(row_index, source_col).value = matched_model
            supplement.cell(row_index, note_col).fill = mismatch_fill
            skipped_line_mismatch_rows += 1
            continue

        supplement.cell(row_index, model_col).value = matched_model
        supplement.cell(row_index, model_col).fill = candidate_fill
        supplement.cell(row_index, note_col).value = "BOM候选，待确认"
        supplement.cell(row_index, source_col).value = matched_model
        suggested_rows += 1

    remaining_rows = 0
    for row_index in range(2, supplement.max_row + 1):
        if not _is_valid_manual_model(supplement.cell(row_index, model_col).value):
            remaining_rows += 1

    style_header(supplement)
    _fit_columns(supplement)
    if conflicts:
        logger.warning(f"钣金型号BOM表存在同一BOM号对应多个箱体组件的情况，已保留首次匹配：{conflicts} 条。")
    if skipped_line_mismatch_rows:
        logger.warning(f"BOM候选中有 {skipped_line_mismatch_rows} 行疑似线体类型不匹配，已保留在说明列等待人工确认。")
    logger.info(
        f"已将BOM候选钣金型号写入“{SHEET_METAL_SUPPLEMENT_SHEET}”："
        f"候选 {suggested_rows} 行；未匹配 {unmatched_rows} 行；"
        f"已有人工作值跳过 {skipped_existing_rows} 行；仍待确认/补充 {remaining_rows} 行。"
    )
    return SheetMetalFillResult(
        filled_rows=suggested_rows,
        remaining_rows=remaining_rows,
        lookup_rows=len(lookup_pairs),
    )


def _bom_model_matches_line(line: str, model: Any) -> bool:
    """Avoid filling obvious rolling/wave mismatches from a broad BOM export."""
    model_text = str(model or "")
    is_wave_line = line in WAVE_LINES
    if is_wave_line and "滚筒" in model_text:
        return False
    if not is_wave_line and "波轮" in model_text:
        return False
    return True


def _ensure_supplement_column_before(sheet: Worksheet, header: str, before_header: str) -> int:
    headers = _header_map(sheet)
    if before_header not in headers:
        raise MissingRequiredFieldsError(f"{sheet.title}缺少关键字段：" + before_header)
    before_col = headers[before_header]
    if header in headers:
        existing_col = headers[header]
        if existing_col < before_col:
            return existing_col
        column_values = []
        for row_index in range(1, sheet.max_row + 1):
            cell = sheet.cell(row_index, existing_col)
            column_values.append(
                {
                    "value": cell.value,
                    "fill": copy(cell.fill),
                    "font": copy(cell.font),
                    "alignment": copy(cell.alignment),
                    "border": copy(cell.border),
                    "number_format": cell.number_format,
                    "protection": copy(cell.protection),
                }
            )
        width = sheet.column_dimensions[get_column_letter(existing_col)].width
        sheet.delete_cols(existing_col)
        sheet.insert_cols(before_col)
        for row_index, saved in enumerate(column_values, start=1):
            cell = sheet.cell(row_index, before_col)
            cell.value = saved["value"]
            cell.fill = copy(saved["fill"])
            cell.font = copy(saved["font"])
            cell.alignment = copy(saved["alignment"])
            cell.border = copy(saved["border"])
            cell.number_format = saved["number_format"]
            cell.protection = copy(saved["protection"])
        if width:
            sheet.column_dimensions[get_column_letter(before_col)].width = width
        return before_col

    sheet.insert_cols(before_col)
    sheet.cell(1, before_col).value = header
    return before_col


def apply_manual_sheet_metal_models(
    workbook: Workbook,
    values_workbook: Workbook,
    target_sheet_name: str,
    logger: ProcessingLogger,
) -> ManualSheetMetalApplyResult:
    if SHEET_METAL_SUPPLEMENT_SHEET not in workbook.sheetnames:
        logger.info(f"未找到“{SHEET_METAL_SUPPLEMENT_SHEET}”，无需手工回填钣金型号。")
        return ManualSheetMetalApplyResult(applied_rows=0, remaining_rows=0)

    source_sheet = workbook[SHEET_METAL_SUPPLEMENT_SHEET]
    values_source_sheet = (
        values_workbook[SHEET_METAL_SUPPLEMENT_SHEET]
        if SHEET_METAL_SUPPLEMENT_SHEET in values_workbook.sheetnames
        else None
    )
    formula_sheet = workbook[target_sheet_name]
    source_headers = _header_map(source_sheet)
    main_headers = _header_map(formula_sheet)
    _require_columns(source_headers, ["原始行号", "钣金型号"], SHEET_METAL_SUPPLEMENT_SHEET)
    _require_columns(main_headers, ["钣金型号"], "主数据表")

    row_no_col = source_headers["原始行号"]
    source_code_col = source_headers.get("物料编码")
    source_model_col = source_headers["钣金型号"]
    values_source_model_col = None
    if values_source_sheet is not None:
        values_source_headers = _header_map(values_source_sheet)
        values_source_model_col = values_source_headers.get("钣金型号")
    main_model_col = main_headers["钣金型号"]
    main_code_col = main_headers.get("物料编码")
    remaining_rows: list[int] = []
    applied_rows = 0

    rows_by_code: dict[str, list[int]] = {}
    if source_code_col and main_code_col:
        supplement_codes = {
            _lookup_key(source_sheet.cell(row_index, source_code_col).value)
            for row_index in range(2, source_sheet.max_row + 1)
        }
        supplement_codes.discard("")
        for row_index in range(2, formula_sheet.max_row + 1):
            code = _lookup_key(formula_sheet.cell(row_index, main_code_col).value)
            if code in supplement_codes:
                rows_by_code.setdefault(code, []).append(row_index)

    for row_index in range(2, source_sheet.max_row + 1):
        original_row = source_sheet.cell(row_index, row_no_col).value
        model = _manual_sheet_metal_model(
            source_sheet.cell(row_index, source_model_col).value,
            (
                values_source_sheet.cell(row_index, values_source_model_col).value
                if values_source_sheet is not None and values_source_model_col is not None
                else None
            ),
        )
        source_code = _lookup_key(source_sheet.cell(row_index, source_code_col).value) if source_code_col else ""
        target_rows = rows_by_code.get(source_code, []) if source_code else []
        if not target_rows:
            if _is_blank(original_row):
                continue
            try:
                target_rows = [int(original_row)]
            except (TypeError, ValueError):
                logger.warning(f"{SHEET_METAL_SUPPLEMENT_SHEET} 第 {row_index} 行原始行号无效，已跳过：{original_row}")
                continue

        if _is_valid_manual_model(model):
            for target_row in target_rows:
                formula_sheet.cell(target_row, main_model_col).value = model
                applied_rows += 1
        else:
            remaining_rows.extend(target_rows)

    remaining_rows = sorted(set(remaining_rows))
    refresh_sheet_metal_supplement_sheet(workbook, formula_sheet, remaining_rows, logger)
    logger.info(f"已回填手工钣金型号：{applied_rows} 行；剩余暂不处理 {len(remaining_rows)} 行。")
    return ManualSheetMetalApplyResult(applied_rows=applied_rows, remaining_rows=len(remaining_rows))


def read_sheet_metal_lookup(lookup_path, logger: ProcessingLogger) -> list[tuple[Any, Any]]:
    import openpyxl

    if not lookup_path.exists():
        raise FileNotFoundError(f"钣金型号查询表不存在：{lookup_path}")
    lookup_wb = openpyxl.load_workbook(lookup_path, data_only=True)
    source_sheet = lookup_wb[lookup_wb.sheetnames[0]]
    headers = _header_map(source_sheet)
    code_field = _first_existing_header(headers, LOOKUP_CODE_FIELDS)
    model_field = _first_existing_header(headers, LOOKUP_MODEL_FIELDS)
    missing = []
    if code_field is None:
        missing.append("物料编码/新总装编码")
    if model_field is None:
        missing.append("钣金型号/半成品机型")
    if missing:
        raise MissingRequiredFieldsError("钣金型号查询表缺少关键字段：" + "、".join(missing))
    code_col = headers[code_field]
    model_col = headers[model_field]

    rows: list[tuple[Any, Any]] = []
    for row_index in range(2, source_sheet.max_row + 1):
        code = source_sheet.cell(row_index, code_col).value
        model = source_sheet.cell(row_index, model_col).value
        if _is_blank(code):
            continue
        rows.append((code, model))
    logger.info(f"已读取钣金型号查询表：{len(rows)} 条；编码列={code_field}，钣金型号列={model_field}。")
    return rows


def read_sheet_metal_bom_lookup(lookup_path, logger: ProcessingLogger) -> list[tuple[Any, Any]]:
    import openpyxl

    if not lookup_path.exists():
        raise FileNotFoundError(f"钣金型号BOM表不存在：{lookup_path}")
    lookup_wb = openpyxl.load_workbook(lookup_path, data_only=True)
    source_sheet = lookup_wb[lookup_wb.sheetnames[0]]
    headers = _header_map(source_sheet)
    _require_columns(headers, [BOM_CODE_FIELD, BOM_MODEL_FIELD], "钣金型号BOM表")
    code_col = headers[BOM_CODE_FIELD]
    model_col = headers[BOM_MODEL_FIELD]

    rows: list[tuple[Any, Any]] = []
    for row_index in range(2, source_sheet.max_row + 1):
        code = source_sheet.cell(row_index, code_col).value
        model = source_sheet.cell(row_index, model_col).value
        if _is_blank(code) or _is_blank(model):
            continue
        if BOM_MODEL_KEYWORD not in str(model):
            continue
        rows.append((code, model))
    lookup_wb.close()
    logger.info(
        f"已读取钣金型号BOM表：{len(rows)} 条含“{BOM_MODEL_KEYWORD}”的BOM组件记录。"
    )
    return rows


def import_sheet_metal_lookup_sheet(
    workbook: Workbook,
    lookup_pairs: list[tuple[Any, Any]],
    logger: ProcessingLogger,
) -> None:
    remove_sheet_if_exists(workbook, SHEET_METAL_LOOKUP_SHEET)
    sheet = workbook.create_sheet(SHEET_METAL_LOOKUP_SHEET)
    sheet.append(["物料编码", "钣金型号"])
    for code, model in lookup_pairs:
        sheet.append([code, model])
    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = sheet.dimensions
    style_header(sheet)
    _fit_columns(sheet)
    logger.info(f"已导入“{SHEET_METAL_LOOKUP_SHEET}”：{len(lookup_pairs)} 条。")


def import_sheet_metal_bom_lookup_sheet(
    workbook: Workbook,
    lookup_pairs: list[tuple[Any, Any]],
    logger: ProcessingLogger,
) -> None:
    remove_sheet_if_exists(workbook, SHEET_METAL_BOM_LOOKUP_SHEET)
    sheet = workbook.create_sheet(SHEET_METAL_BOM_LOOKUP_SHEET)
    sheet.append(["BOM号", "箱体组件描述"])
    for code, model in lookup_pairs:
        sheet.append([code, model])
    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = sheet.dimensions
    style_header(sheet)
    _fit_columns(sheet)
    logger.info(f"已导入“{SHEET_METAL_BOM_LOOKUP_SHEET}”：{len(lookup_pairs)} 条。")


def refresh_sheet_metal_supplement_sheet(
    workbook: Workbook,
    formula_sheet: Worksheet,
    row_numbers: list[int],
    logger: ProcessingLogger,
) -> None:
    remove_sheet_if_exists(workbook, SHEET_METAL_SUPPLEMENT_SHEET)
    supplement = workbook.create_sheet(SHEET_METAL_SUPPLEMENT_SHEET)
    output_headers = ["原始行号"] + [
        normalize_header(cell.value) or f"空列{index}"
        for index, cell in enumerate(formula_sheet[1], start=1)
    ]
    supplement.append(output_headers)
    for row_index in row_numbers:
        row_values = [
            formula_sheet.cell(row_index, col_index).value
            for col_index in range(1, formula_sheet.max_column + 1)
        ]
        supplement.append([row_index] + row_values)

    supplement.freeze_panes = "A2"
    if supplement.max_row > 1:
        supplement.auto_filter.ref = supplement.dimensions
    style_header(supplement)
    _fit_columns(supplement)
    logger.info(f"已刷新“{SHEET_METAL_SUPPLEMENT_SHEET}”：剩余待补充 {len(row_numbers)} 行。")


def _header_map(sheet: Worksheet) -> dict[str, int]:
    raw_headers = [normalize_header(cell.value) for cell in sheet[1]]
    headers = deduplicate_headers(raw_headers)
    return {header: index + 1 for index, header in enumerate(headers)}


def _require_columns(header_map: dict[str, int], columns: list[str], context: str) -> None:
    missing = [column for column in columns if column not in header_map]
    if missing:
        raise MissingRequiredFieldsError(f"{context}缺少关键字段：" + "、".join(missing))


def _first_existing_header(header_map: dict[str, int], candidates: tuple[str, ...]) -> str | None:
    for candidate in candidates:
        if candidate in header_map:
            return candidate
    return None


def _lookup_key(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_na(value: Any) -> bool:
    if value is None:
        return False
    value_text = str(value).strip()
    return value_text.upper() in {"#N/A", "#NA", "N/A"} or value_text in {"0", "0.0", "0.00", "-"}


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _is_valid_manual_model(value: Any) -> bool:
    if _is_blank(value) or _is_na(value):
        return False
    if isinstance(value, str) and value.strip().startswith("="):
        return False
    return True


def _manual_sheet_metal_model(formula_value: Any, cached_value: Any) -> Any:
    """Return a usable supplement value, including an evaluated formula result."""
    if isinstance(formula_value, str) and formula_value.strip().startswith("="):
        return cached_value if _is_valid_manual_model(cached_value) else None
    return formula_value


def _fit_columns(sheet: Worksheet) -> None:
    for column_cells in sheet.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 45)
