"""Preprocessing and coefficient backfill helpers."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
import csv

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .data_cleaner import deduplicate_headers, normalize_header
from .excel_io import remove_sheet_if_exists
from .exceptions import ClassifierError, MissingRequiredFieldsError
from .logger import ProcessingLogger
from .reporter import style_header

COEFFICIENT_SUPPLEMENT_SHEET = "系数补充"
COEFFICIENT_LOOKUP_SHEET = "系数查询表"
COEFFICIENT_STILL_MISSING_SHEET = "系数仍缺失"
EXCLUDED_MATERIAL_CODES = {
    "Z4U6010100",
    "Z4U6010108",
    "Z4U60501080",
}


@dataclass
class CoefficientPrepareResult:
    deleted_blank_order_rows: int
    deleted_blank_line_rows: int
    deleted_excluded_material_rows: int
    coefficient_missing_rows: int


@dataclass
class CoefficientFillResult:
    formulas_written: int
    still_missing_rows: int


@dataclass
class ManualCoefficientApplyResult:
    applied_rows: int
    remaining_rows: int


@dataclass
class OrderCleanupResult:
    deleted_blank_order_rows: int
    deleted_blank_line_rows: int
    deleted_excluded_material_rows: int


@dataclass
class CoefficientSupplementRowsCleanupResult:
    deleted_rows: int
    deleted_codes: int


def _header_map(sheet: Worksheet) -> dict[str, int]:
    raw_headers = [normalize_header(cell.value) for cell in sheet[1]]
    headers = deduplicate_headers(raw_headers)
    return {header: index + 1 for index, header in enumerate(headers)}


def _normalized_header_map(headers: list[Any]) -> dict[str, int]:
    return {normalize_header(header): index for index, header in enumerate(headers)}


def _require_columns(header_map: dict[str, int], columns: list[str], context: str) -> None:
    missing = [column for column in columns if column not in header_map]
    if missing:
        raise MissingRequiredFieldsError(f"{context}缺少关键字段：" + "、".join(missing))


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _is_na(value: Any) -> bool:
    return value is not None and str(value).strip().upper() in {"#N/A", "#NA", "N/A"}


def _code_key(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def delete_blank_order_rows(sheet: Worksheet, logger: ProcessingLogger) -> int:
    headers = _header_map(sheet)
    _require_columns(headers, ["订单数"], "主数据表")
    order_col = headers["订单数"]

    rows_to_delete = [
        row_index
        for row_index in range(2, sheet.max_row + 1)
        if _is_blank(sheet.cell(row_index, order_col).value)
    ]
    for row_index in reversed(rows_to_delete):
        sheet.delete_rows(row_index, 1)

    logger.info(f"订单数为空白的行已删除：{len(rows_to_delete)} 行。")
    return len(rows_to_delete)


def delete_blank_line_rows(sheet: Worksheet, logger: ProcessingLogger) -> int:
    headers = _header_map(sheet)
    _require_columns(headers, ["线体"], "主数据表")
    line_col = headers["线体"]

    rows_to_delete = [
        row_index
        for row_index in range(2, sheet.max_row + 1)
        if _is_blank(sheet.cell(row_index, line_col).value)
    ]
    for row_index in reversed(rows_to_delete):
        sheet.delete_rows(row_index, 1)

    logger.info(f"线体为空白的行已删除：{len(rows_to_delete)} 行。")
    return len(rows_to_delete)


def delete_excluded_material_code_rows(sheet: Worksheet, logger: ProcessingLogger) -> int:
    headers = _header_map(sheet)
    _require_columns(headers, ["物料编码"], "主数据表")
    code_col = headers["物料编码"]

    rows_to_delete = [
        row_index
        for row_index in range(2, sheet.max_row + 1)
        if _code_key(sheet.cell(row_index, code_col).value) in EXCLUDED_MATERIAL_CODES
    ]
    for row_index in reversed(rows_to_delete):
        sheet.delete_rows(row_index, 1)

    logger.info(
        "指定物料编码订单行已删除："
        f"{len(rows_to_delete)} 行；物料编码={', '.join(sorted(EXCLUDED_MATERIAL_CODES))}。"
    )
    return len(rows_to_delete)


def cleanup_order_rows(sheet: Worksheet, logger: ProcessingLogger) -> OrderCleanupResult:
    deleted_blank_rows = delete_blank_order_rows(sheet, logger)
    deleted_blank_line_rows = delete_blank_line_rows(sheet, logger)
    deleted_excluded_rows = delete_excluded_material_code_rows(sheet, logger)
    return OrderCleanupResult(
        deleted_blank_order_rows=deleted_blank_rows,
        deleted_blank_line_rows=deleted_blank_line_rows,
        deleted_excluded_material_rows=deleted_excluded_rows,
    )


def cleanup_order_rows_preserve_blank_orders(sheet: Worksheet, logger: ProcessingLogger) -> OrderCleanupResult:
    """Run all foundation cleanup rules except deleting blank-order rows."""
    deleted_blank_line_rows = delete_blank_line_rows(sheet, logger)
    deleted_excluded_rows = delete_excluded_material_code_rows(sheet, logger)
    return OrderCleanupResult(
        deleted_blank_order_rows=0,
        deleted_blank_line_rows=deleted_blank_line_rows,
        deleted_excluded_material_rows=deleted_excluded_rows,
    )


def delete_coefficient_supplement_rows(
    workbook: Workbook,
    target_sheet_name: str,
    logger: ProcessingLogger,
) -> CoefficientSupplementRowsCleanupResult:
    if COEFFICIENT_SUPPLEMENT_SHEET not in workbook.sheetnames:
        logger.info(f"未找到“{COEFFICIENT_SUPPLEMENT_SHEET}”，无需删除系数补充对应订单行。")
        return CoefficientSupplementRowsCleanupResult(deleted_rows=0, deleted_codes=0)

    supplement = workbook[COEFFICIENT_SUPPLEMENT_SHEET]
    supplement_headers = _header_map(supplement)
    _require_columns(supplement_headers, ["物料编码"], COEFFICIENT_SUPPLEMENT_SHEET)
    supplement_code_col = supplement_headers["物料编码"]
    codes = {
        _code_key(supplement.cell(row_index, supplement_code_col).value)
        for row_index in range(2, supplement.max_row + 1)
        if not _is_blank(supplement.cell(row_index, supplement_code_col).value)
    }
    codes.discard("")
    if not codes:
        logger.info(f"“{COEFFICIENT_SUPPLEMENT_SHEET}”没有可删除的物料编码。")
        return CoefficientSupplementRowsCleanupResult(deleted_rows=0, deleted_codes=0)

    sheet = workbook[target_sheet_name]
    headers = _header_map(sheet)
    _require_columns(headers, ["物料编码"], "主数据表")
    main_code_col = headers["物料编码"]
    rows_to_delete = [
        row_index
        for row_index in range(2, sheet.max_row + 1)
        if _code_key(sheet.cell(row_index, main_code_col).value) in codes
    ]
    for row_index in reversed(rows_to_delete):
        sheet.delete_rows(row_index, 1)

    refresh_coefficient_supplement_sheet(workbook, sheet, [], logger)
    create_still_missing_sheet(workbook, sheet, [], logger)
    logger.info(
        f"已删除系数补充对应订单行：{len(rows_to_delete)} 行；"
        f"涉及物料编码 {len(codes)} 个。"
    )
    return CoefficientSupplementRowsCleanupResult(
        deleted_rows=len(rows_to_delete),
        deleted_codes=len(codes),
    )


def refresh_coefficient_lookup_formulas(sheet: Worksheet, logger: ProcessingLogger) -> int:
    headers = _header_map(sheet)
    if "系数" not in headers or "物料编码" not in headers:
        logger.warning("未刷新系数查找公式：主数据表缺少 系数 或 物料编码。")
        return 0

    coefficient_col = headers["系数"]
    code_col = headers["物料编码"]
    code_letter = get_column_letter(code_col)
    refreshed = 0
    for row_index in range(2, sheet.max_row + 1):
        cell = sheet.cell(row_index, coefficient_col)
        value = cell.value
        if not (isinstance(value, str) and value.strip().upper().startswith("=VLOOKUP(")):
            continue
        if "系数" not in value:
            continue
        new_formula = f"=VLOOKUP({code_letter}{row_index},系数!B:G,5,0)"
        if value != new_formula:
            cell.value = new_formula
            refreshed += 1

    logger.info(f"已刷新系数列 VLOOKUP 当前行引用：{refreshed} 行。")
    return refreshed


def create_coefficient_supplement_sheet(
    workbook: Workbook,
    formula_sheet: Worksheet,
    values_sheet: Worksheet,
    logger: ProcessingLogger,
) -> int:
    headers = _header_map(values_sheet)
    _require_columns(headers, ["系数", "物料编码"], "主数据表")
    coefficient_col = headers["系数"]

    remove_sheet_if_exists(workbook, COEFFICIENT_SUPPLEMENT_SHEET)
    sheet = workbook.create_sheet(COEFFICIENT_SUPPLEMENT_SHEET)
    output_headers = ["原始行号"] + [
        normalize_header(cell.value) or f"空列{index}"
        for index, cell in enumerate(formula_sheet[1], start=1)
    ]
    sheet.append(output_headers)

    missing_count = 0
    for row_index in range(2, values_sheet.max_row + 1):
        if not _is_na(values_sheet.cell(row_index, coefficient_col).value):
            continue
        missing_count += 1
        row_values = [formula_sheet.cell(row_index, col_index).value for col_index in range(1, formula_sheet.max_column + 1)]
        sheet.append([row_index] + row_values)

    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = sheet.dimensions
    style_header(sheet)
    _fit_columns(sheet)
    logger.info(f"系数为 #N/A 的行已摘取到“{COEFFICIENT_SUPPLEMENT_SHEET}”：{missing_count} 行。")
    return missing_count


def prepare_coefficients(
    workbook: Workbook,
    values_workbook: Workbook,
    target_sheet_name: str,
    logger: ProcessingLogger,
    cleanup_rows: bool = True,
) -> CoefficientPrepareResult:
    formula_sheet = workbook[target_sheet_name]
    # A prior manual-fill pass may have left an outdated “系数仍缺失” sheet.
    # The fresh “系数补充” sheet built below is the source of truth for this
    # preparation pass, so remove the stale snapshot before recomputing it.
    remove_sheet_if_exists(workbook, COEFFICIENT_STILL_MISSING_SHEET)
    cleanup_result = (
        cleanup_order_rows(formula_sheet, logger)
        if cleanup_rows
        else OrderCleanupResult(0, 0, 0)
    )
    refresh_coefficient_lookup_formulas(formula_sheet, logger)

    # The values workbook still reflects the pre-delete copy. Delete the same rows
    # there so row numbers stay aligned when detecting #N/A rows.
    values_sheet = values_workbook[target_sheet_name]
    if cleanup_rows and (
        cleanup_result.deleted_blank_order_rows
        or cleanup_result.deleted_blank_line_rows
        or cleanup_result.deleted_excluded_material_rows
    ):
        cleanup_order_rows(values_sheet, logger)

    missing_count = create_coefficient_supplement_sheet(workbook, formula_sheet, values_sheet, logger)
    return CoefficientPrepareResult(
        deleted_blank_order_rows=cleanup_result.deleted_blank_order_rows,
        deleted_blank_line_rows=cleanup_result.deleted_blank_line_rows,
        deleted_excluded_material_rows=cleanup_result.deleted_excluded_material_rows,
        coefficient_missing_rows=missing_count,
    )


def copy_lookup_sheet_into_workbook(
    workbook: Workbook,
    lookup_path: Path,
    logger: ProcessingLogger,
) -> tuple[int, int]:
    if not lookup_path.exists():
        raise FileNotFoundError(f"系数查询表不存在：{lookup_path}")

    lookup_rows = read_lookup_rows(lookup_path)

    remove_sheet_if_exists(workbook, COEFFICIENT_LOOKUP_SHEET)
    target_sheet = workbook.create_sheet(COEFFICIENT_LOOKUP_SHEET)
    target_sheet.append(["物料编码", "系数"])
    count = 0
    for code, coefficient in lookup_rows:
        if _is_blank(code):
            continue
        target_sheet.append([code, coefficient])
        count += 1

    if count == 0:
        raise ClassifierError("系数查询表没有可用的物料编码数据。")

    target_sheet.freeze_panes = "A2"
    target_sheet.auto_filter.ref = target_sheet.dimensions
    style_header(target_sheet)
    _fit_columns(target_sheet)
    logger.info(f"已导入系数查询表：{count} 条物料编码。")
    return target_sheet.max_row, count


def read_lookup_rows(lookup_path: Path) -> list[tuple[Any, Any]]:
    """Read lookup rows from .xlsx or SAP text-style .xls exports."""
    suffix = lookup_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return read_lookup_rows_from_xlsx(lookup_path)
    return read_lookup_rows_from_text_export(lookup_path)


def read_lookup_rows_from_xlsx(lookup_path: Path) -> list[tuple[Any, Any]]:
    lookup_wb = openpyxl.load_workbook(lookup_path, data_only=True)
    source_sheet = lookup_wb[lookup_wb.sheetnames[0]]
    source_headers = _header_map(source_sheet)
    coefficient_header = find_coefficient_header(source_headers)
    _require_columns(source_headers, ["物料编码", coefficient_header], "系数查询表")

    code_col = source_headers["物料编码"]
    coefficient_col = source_headers[coefficient_header]
    rows: list[tuple[Any, Any]] = []
    for row_index in range(2, source_sheet.max_row + 1):
        rows.append((
            source_sheet.cell(row_index, code_col).value,
            source_sheet.cell(row_index, coefficient_col).value,
        ))
    return rows


def read_lookup_rows_from_text_export(lookup_path: Path) -> list[tuple[Any, Any]]:
    data = lookup_path.read_bytes()
    for encoding in ("utf-16", "utf-8-sig", "gb18030"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ClassifierError(f"无法识别系数查询表编码：{lookup_path}")

    parsed_rows = list(csv.reader(text.splitlines(), delimiter="\t"))
    if not parsed_rows:
        raise ClassifierError("系数查询表为空。")
    header_row_index = None
    source_headers: dict[str, int] | None = None
    coefficient_header = None
    for row_index, row in enumerate(parsed_rows):
        candidate_headers = _normalized_header_map(row)
        try:
            candidate_coefficient_header = find_coefficient_header(candidate_headers)
            _require_text_columns(candidate_headers, ["物料编码", candidate_coefficient_header], "系数查询表")
        except MissingRequiredFieldsError:
            continue
        header_row_index = row_index
        source_headers = candidate_headers
        coefficient_header = candidate_coefficient_header
        break

    if header_row_index is None or source_headers is None or coefficient_header is None:
        raise MissingRequiredFieldsError("系数查询表缺少关键字段：物料编码、系数 或 等级D")

    code_index = source_headers["物料编码"]
    coefficient_index = source_headers[coefficient_header]
    rows: list[tuple[Any, Any]] = []
    for row in parsed_rows[header_row_index + 1:]:
        code = row[code_index] if code_index < len(row) else None
        coefficient = row[coefficient_index] if coefficient_index < len(row) else None
        if code is None or not str(code).strip():
            continue
        rows.append((code, coefficient))
    return rows


def find_coefficient_header(header_map: dict[str, int]) -> str:
    for candidate in ("系数", "等级D"):
        if candidate in header_map:
            return candidate
    raise MissingRequiredFieldsError("系数查询表缺少关键字段：系数 或 等级D")


def _require_text_columns(header_map: dict[str, int], columns: list[str], context: str) -> None:
    missing = [column for column in columns if column not in header_map]
    if missing:
        raise MissingRequiredFieldsError(f"{context}缺少关键字段：" + "、".join(missing))


def fill_coefficients(
    workbook: Workbook,
    values_workbook: Workbook,
    target_sheet_name: str,
    lookup_path: Path,
    logger: ProcessingLogger,
) -> CoefficientFillResult:
    lookup_max_row, lookup_count = copy_lookup_sheet_into_workbook(workbook, lookup_path, logger)
    if lookup_count <= 0:
        raise ClassifierError("系数查询表为空，无法回填系数。")

    formula_sheet = workbook[target_sheet_name]
    values_sheet = values_workbook[target_sheet_name]
    headers = _header_map(values_sheet)
    _require_columns(headers, ["系数", "物料编码"], "主数据表")
    coefficient_col = headers["系数"]
    code_col = headers["物料编码"]
    coefficient_col_letter = get_column_letter(coefficient_col)
    code_col_letter = get_column_letter(code_col)

    lookup_codes = _load_lookup_codes(workbook[COEFFICIENT_LOOKUP_SHEET])
    still_missing_rows: list[int] = []
    formulas_written = 0

    for row_index in range(2, values_sheet.max_row + 1):
        if not _is_na(values_sheet.cell(row_index, coefficient_col).value):
            continue
        code = values_sheet.cell(row_index, code_col).value
        formula_sheet.cell(row_index, coefficient_col).value = (
            f'=IFERROR(VLOOKUP({code_col_letter}{row_index},'
            f"'{COEFFICIENT_LOOKUP_SHEET}'!$A$2:$B${lookup_max_row},2,0),\"#N/A\")"
        )
        formulas_written += 1
        if code not in lookup_codes:
            still_missing_rows.append(row_index)

    create_still_missing_sheet(workbook, formula_sheet, still_missing_rows, logger)
    refresh_coefficient_supplement_sheet(workbook, formula_sheet, still_missing_rows, logger)
    logger.info(
        f"系数回填公式已写入：{formulas_written} 行；查询表暂未覆盖：{len(still_missing_rows)} 行。"
    )
    return CoefficientFillResult(
        formulas_written=formulas_written,
        still_missing_rows=len(still_missing_rows),
    )


def apply_manual_coefficients(
    workbook: Workbook,
    values_workbook: Workbook,
    target_sheet_name: str,
    logger: ProcessingLogger,
) -> ManualCoefficientApplyResult:
    source_sheet = find_manual_coefficient_sheet(workbook)
    values_source_sheet = values_workbook[source_sheet.title] if source_sheet.title in values_workbook.sheetnames else None
    formula_sheet = workbook[target_sheet_name]
    source_headers = _header_map(source_sheet)
    main_headers = _header_map(formula_sheet)
    _require_columns(source_headers, ["原始行号", "系数"], source_sheet.title)
    _require_columns(main_headers, ["系数", "物料编码"], "主数据表")

    row_no_col = source_headers["原始行号"]
    source_coefficient_col = source_headers["系数"]
    values_source_coefficient_col = None
    if values_source_sheet is not None:
        values_source_headers = _header_map(values_source_sheet)
        values_source_coefficient_col = values_source_headers.get("系数")
    source_code_col = source_headers.get("物料编码")
    main_coefficient_col = main_headers["系数"]
    main_code_col = main_headers["物料编码"]

    remaining_rows: list[int] = []
    applied_rows = 0
    if source_code_col:
        coefficient_by_code: dict[str, Any] = {}
        remaining_codes: set[str] = set()
        invalid_original_rows: list[int] = []

        for row_index in range(2, source_sheet.max_row + 1):
            code = _code_key(source_sheet.cell(row_index, source_code_col).value)
            coefficient = _manual_coefficient_value(
                source_sheet.cell(row_index, source_coefficient_col).value,
                (
                    values_source_sheet.cell(row_index, values_source_coefficient_col).value
                    if values_source_sheet is not None and values_source_coefficient_col is not None
                    else None
                ),
            )
            if _is_manual_coefficient_value(coefficient):
                if code:
                    coefficient_by_code[code] = coefficient
            elif code:
                remaining_codes.add(code)
            else:
                original_row = source_sheet.cell(row_index, row_no_col).value
                if _is_blank(original_row):
                    continue
                try:
                    invalid_original_rows.append(int(original_row))
                except (TypeError, ValueError):
                    logger.warning(f"{source_sheet.title} 第 {row_index} 行原始行号无效，已跳过：{original_row}")

        if coefficient_by_code:
            for row_index in range(2, formula_sheet.max_row + 1):
                code = _code_key(formula_sheet.cell(row_index, main_code_col).value)
                if code not in coefficient_by_code:
                    continue
                current_value = formula_sheet.cell(row_index, main_coefficient_col).value
                if not _is_missing_or_lookup_formula(current_value):
                    continue
                formula_sheet.cell(row_index, main_coefficient_col).value = coefficient_by_code[code]
                applied_rows += 1

        if remaining_codes:
            for row_index in range(2, formula_sheet.max_row + 1):
                code = _code_key(formula_sheet.cell(row_index, main_code_col).value)
                if code in remaining_codes and _is_missing_or_lookup_formula(
                    formula_sheet.cell(row_index, main_coefficient_col).value
                ):
                    remaining_rows.append(row_index)

        remaining_rows.extend(
            row_index
            for row_index in invalid_original_rows
            if 2 <= row_index <= formula_sheet.max_row
            and _is_missing_or_lookup_formula(formula_sheet.cell(row_index, main_coefficient_col).value)
        )
        remaining_rows = sorted(set(remaining_rows))
        create_still_missing_sheet(workbook, formula_sheet, remaining_rows, logger)
        refresh_coefficient_supplement_sheet(workbook, formula_sheet, remaining_rows, logger)
        logger.info(
            f"已从“{source_sheet.title}”按物料编码回填手工系数：{applied_rows} 行；仍待补充 {len(remaining_rows)} 行。"
        )
        return ManualCoefficientApplyResult(applied_rows=applied_rows, remaining_rows=len(remaining_rows))

    for row_index in range(2, source_sheet.max_row + 1):
        original_row = source_sheet.cell(row_index, row_no_col).value
        coefficient = _manual_coefficient_value(
            source_sheet.cell(row_index, source_coefficient_col).value,
            (
                values_source_sheet.cell(row_index, values_source_coefficient_col).value
                if values_source_sheet is not None and values_source_coefficient_col is not None
                else None
            ),
        )
        if _is_blank(original_row):
            continue
        try:
            original_row_number = int(original_row)
        except (TypeError, ValueError):
            logger.warning(f"{source_sheet.title} 第 {row_index} 行原始行号无效，已跳过：{original_row}")
            continue

        if _is_manual_coefficient_value(coefficient):
            formula_sheet.cell(original_row_number, main_coefficient_col).value = coefficient
            applied_rows += 1
        else:
            remaining_rows.append(original_row_number)

    create_still_missing_sheet(workbook, formula_sheet, remaining_rows, logger)
    refresh_coefficient_supplement_sheet(workbook, formula_sheet, remaining_rows, logger)
    logger.info(
        f"已从“{source_sheet.title}”回填手工系数：{applied_rows} 行；仍待补充 {len(remaining_rows)} 行。"
    )
    return ManualCoefficientApplyResult(applied_rows=applied_rows, remaining_rows=len(remaining_rows))


def find_manual_coefficient_sheet(workbook: Workbook) -> Worksheet:
    for sheet_name in ("系数待补充", COEFFICIENT_STILL_MISSING_SHEET, COEFFICIENT_SUPPLEMENT_SHEET):
        if sheet_name in workbook.sheetnames and workbook[sheet_name].max_row > 1:
            return workbook[sheet_name]
    for sheet_name in ("系数待补充", COEFFICIENT_STILL_MISSING_SHEET, COEFFICIENT_SUPPLEMENT_SHEET):
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
    raise ClassifierError("未找到可用于手工系数回填的工作表：系数待补充、系数仍缺失 或 系数补充")


def _is_manual_coefficient_value(value: Any) -> bool:
    if _is_blank(value) or _is_na(value):
        return False
    if isinstance(value, str) and value.strip().startswith("="):
        return False
    return True


def _manual_coefficient_value(formula_value: Any, cached_value: Any) -> Any:
    """Return a confirmed coefficient from a supplement cell.

    A supplement may contain a VLOOKUP formula while the paired data-only
    workbook contains Excel's already-calculated numeric result.  Treat that
    numeric cache as a usable value; formula errors and empty results remain
    unresolved.
    """
    if isinstance(formula_value, str) and formula_value.strip().startswith("="):
        return cached_value if _is_manual_coefficient_value(cached_value) else None
    return formula_value


def _is_missing_or_lookup_formula(value: Any) -> bool:
    if _is_blank(value) or _is_na(value):
        return True
    return isinstance(value, str) and value.strip().startswith("=")


def _load_lookup_codes(lookup_sheet: Worksheet) -> set[Any]:
    return {
        lookup_sheet.cell(row_index, 1).value
        for row_index in range(2, lookup_sheet.max_row + 1)
        if not _is_blank(lookup_sheet.cell(row_index, 1).value)
    }


def create_still_missing_sheet(
    workbook: Workbook,
    formula_sheet: Worksheet,
    row_numbers: list[int],
    logger: ProcessingLogger,
) -> None:
    remove_sheet_if_exists(workbook, COEFFICIENT_STILL_MISSING_SHEET)
    sheet = workbook.create_sheet(COEFFICIENT_STILL_MISSING_SHEET)
    output_headers = ["原始行号"] + [
        normalize_header(cell.value) or f"空列{index}"
        for index, cell in enumerate(formula_sheet[1], start=1)
    ]
    sheet.append(output_headers)
    for row_index in row_numbers:
        row_values = [formula_sheet.cell(row_index, col_index).value for col_index in range(1, formula_sheet.max_column + 1)]
        sheet.append([row_index] + row_values)

    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = sheet.dimensions
    style_header(sheet)
    _fit_columns(sheet)
    logger.info(f"已生成“{COEFFICIENT_STILL_MISSING_SHEET}”：{len(row_numbers)} 行。")


def refresh_coefficient_supplement_sheet(
    workbook: Workbook,
    formula_sheet: Worksheet,
    row_numbers: list[int],
    logger: ProcessingLogger,
) -> None:
    remove_sheet_if_exists(workbook, COEFFICIENT_SUPPLEMENT_SHEET)
    sheet = workbook.create_sheet(COEFFICIENT_SUPPLEMENT_SHEET)
    output_headers = ["原始行号"] + [
        normalize_header(cell.value) or f"空列{index}"
        for index, cell in enumerate(formula_sheet[1], start=1)
    ]
    sheet.append(output_headers)
    for row_index in row_numbers:
        row_values = [formula_sheet.cell(row_index, col_index).value for col_index in range(1, formula_sheet.max_column + 1)]
        sheet.append([row_index] + row_values)

    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = sheet.dimensions
    style_header(sheet)
    _fit_columns(sheet)
    logger.info(f"已刷新“{COEFFICIENT_SUPPLEMENT_SHEET}”：剩余待补充 {len(row_numbers)} 行。")


def _fit_columns(sheet: Worksheet) -> None:
    for column_cells in sheet.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 45)
