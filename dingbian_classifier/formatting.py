"""Main worksheet formatting helpers."""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .data_cleaner import deduplicate_headers, normalize_header
from .exceptions import MissingRequiredFieldsError
from .logger import ProcessingLogger

TYPE_FIELD = "类型"
REMARK_FIELD = "备注"
HEADER_BLUE = "0070C0"
HIDDEN_FIELDS = ["周次", "日产", "已交货数", "产品销售单号", "创建日期"]


@dataclass
class FormatMainSheetResult:
    inserted_type_column: bool
    hidden_fields: list[str]
    formatted_rows: int
    formatted_columns: int


def format_main_sheet(sheet: Worksheet, logger: ProcessingLogger) -> FormatMainSheetResult:
    inserted_type = ensure_type_column_after_remark(sheet, logger)
    apply_main_sheet_styles(sheet, logger)
    hidden_fields = hide_unnecessary_columns(sheet, logger)
    logger.info(
        f"主表格式处理完成：类型列{'已新增' if inserted_type else '已存在'}，隐藏列 {len(hidden_fields)} 个。"
    )
    return FormatMainSheetResult(
        inserted_type_column=inserted_type,
        hidden_fields=hidden_fields,
        formatted_rows=sheet.max_row,
        formatted_columns=sheet.max_column,
    )


def ensure_type_column_after_remark(sheet: Worksheet, logger: ProcessingLogger) -> bool:
    headers = _header_map(sheet)
    if TYPE_FIELD in headers:
        logger.info("检测到已存在“类型”列，不重复新增。")
        return False
    if REMARK_FIELD not in headers:
        raise MissingRequiredFieldsError("主数据表缺少关键字段：备注")

    insert_col = headers[REMARK_FIELD] + 1
    sheet.insert_cols(insert_col, 1)
    sheet.cell(1, insert_col).value = TYPE_FIELD
    logger.info("已在“备注”列后新增“类型”列。")
    return True


def apply_main_sheet_styles(sheet: Worksheet, logger: ProcessingLogger) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_BLUE)
    header_font = Font(color="FFFFFF", bold=True, size=14)
    header_alignment = Alignment(horizontal="center", vertical="center")
    body_font_color = "000000"
    no_fill = PatternFill(fill_type=None)
    thin_side = Side(style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            existing_font = copy(cell.font)
            existing_font.color = body_font_color
            cell.font = existing_font
            cell.border = thin_border
            if cell.row > 1:
                cell.fill = no_fill

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    logger.info(f"已统一主表字体、填充和边框格式：{sheet.max_row} 行，{sheet.max_column} 列。")


def hide_unnecessary_columns(sheet: Worksheet, logger: ProcessingLogger) -> list[str]:
    headers = _header_map(sheet)
    hidden: list[str] = []
    for field in HIDDEN_FIELDS:
        col_index = headers.get(field)
        if not col_index:
            logger.warning(f"未找到需隐藏字段：{field}")
            continue
        sheet.column_dimensions[get_column_letter(col_index)].hidden = True
        hidden.append(field)
    logger.info("已隐藏不必要列：" + "、".join(hidden))
    return hidden


def _header_map(sheet: Worksheet) -> dict[str, int]:
    raw_headers = [normalize_header(cell.value) for cell in sheet[1]]
    headers = deduplicate_headers(raw_headers)
    return {header: index + 1 for index, header in enumerate(headers)}
