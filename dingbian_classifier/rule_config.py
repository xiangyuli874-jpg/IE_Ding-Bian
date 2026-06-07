"""Classification rule configuration."""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook

from .exceptions import RuleConfigError
from .logger import ProcessingLogger

CONFIG_SHEET_NAME = "分类规则配置"

CONFIG_HEADERS = [
    "启用",
    "分类名称",
    "判断字段",
    "包含关键词",
    "排除关键词",
    "输出工作表名称",
    "是否保留原始行",
    "是否去重",
    "去重字段",
    "是否按订单数汇总",
    "汇总维度字段",
    "优先级",
    "备注",
]


@dataclass
class ClassificationRule:
    enabled: bool
    category_name: str
    field_name: str
    include_keywords: list[str]
    exclude_keywords: list[str]
    output_sheet_name: str
    keep_original_rows: bool
    deduplicate: bool
    deduplicate_fields: list[str]
    summarize_order_qty: bool
    summary_fields: list[str]
    priority: int
    note: str = ""


DEFAULT_RULE_ROWS = [
    ["是", "预走货提示", "预走货", "需求;可装;走货", "", "预走货提示", "是", "否", "", "否", "", 10, "示例规则，可按业务需要修改或删除"],
    ["是", "返工相关", "物料描述", "返工;返修", "", "返工相关", "是", "否", "", "否", "", 20, "示例规则"],
    ["是", "摔机相关", "备注", "摔机", "", "摔机相关", "是", "否", "", "否", "", 30, "示例规则"],
    ["是", "项目订单提示", "产品销售单号", "DDW;项目", "", "项目订单提示", "是", "否", "", "否", "", 40, "示例规则"],
]


def truthy(value: Any, default: bool = False) -> bool:
    if value is None or value == "":
        return default
    return str(value).strip().lower() in {"是", "yes", "y", "true", "1", "启用"}


def split_keywords(value: Any) -> list[str]:
    if value is None:
        return []
    text = str(value).replace("；", ";").replace(",", ";").replace("，", ";")
    return [part.strip() for part in text.split(";") if part.strip()]


def ensure_config_sheet(workbook: Workbook, logger: ProcessingLogger) -> None:
    if CONFIG_SHEET_NAME in workbook.sheetnames:
        logger.info(f"检测到已有“{CONFIG_SHEET_NAME}”工作表，将按其中规则执行。")
        return

    sheet = workbook.create_sheet(CONFIG_SHEET_NAME)
    sheet.append(CONFIG_HEADERS)
    for row in DEFAULT_RULE_ROWS:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    logger.info(f"未找到“{CONFIG_SHEET_NAME}”，已创建默认示例规则配置表。")


def load_rules(workbook: Workbook, headers: list[str], logger: ProcessingLogger) -> list[ClassificationRule]:
    sheet = workbook[CONFIG_SHEET_NAME]
    config_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    missing = [header for header in CONFIG_HEADERS if header not in config_headers]
    if missing:
        raise RuleConfigError("分类规则配置表缺少字段：" + "、".join(missing))

    index = {header: config_headers.index(header) for header in CONFIG_HEADERS}
    rules: list[ClassificationRule] = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        enabled = truthy(row[index["启用"]])
        if not enabled:
            continue

        category_name = str(row[index["分类名称"]] or "").strip()
        field_name = str(row[index["判断字段"]] or "").strip()
        output_sheet_name = str(row[index["输出工作表名称"]] or category_name).strip()
        if not category_name or not field_name:
            raise RuleConfigError(f"分类规则配置第 {row_num} 行缺少分类名称或判断字段。")
        if field_name not in headers:
            raise RuleConfigError(f"分类规则配置第 {row_num} 行的判断字段不存在：{field_name}")

        try:
            priority = int(row[index["优先级"]] or 999)
        except ValueError as exc:
            raise RuleConfigError(f"分类规则配置第 {row_num} 行优先级不是整数。") from exc

        rules.append(
            ClassificationRule(
                enabled=enabled,
                category_name=category_name,
                field_name=field_name,
                include_keywords=split_keywords(row[index["包含关键词"]]),
                exclude_keywords=split_keywords(row[index["排除关键词"]]),
                output_sheet_name=output_sheet_name or category_name,
                keep_original_rows=truthy(row[index["是否保留原始行"]], default=True),
                deduplicate=truthy(row[index["是否去重"]]),
                deduplicate_fields=split_keywords(row[index["去重字段"]]),
                summarize_order_qty=truthy(row[index["是否按订单数汇总"]]),
                summary_fields=split_keywords(row[index["汇总维度字段"]]),
                priority=priority,
                note=str(row[index["备注"]] or ""),
            )
        )

    rules.sort(key=lambda rule: rule.priority)
    logger.info(f"已加载启用的分类规则：{len(rules)} 条。")
    return rules

