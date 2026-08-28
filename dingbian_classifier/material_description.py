"""Material-description preparation and backfill helpers."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
import csv

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .data_cleaner import deduplicate_headers, normalize_header
from .excel_io import remove_sheet_if_exists
from .exceptions import ClassifierError, MissingRequiredFieldsError
from .logger import ProcessingLogger
from .reporter import style_header

MATERIAL_DESCRIPTION_SUPPLEMENT_SHEET = "物料描述补充"
MATERIAL_DESCRIPTION_LOOKUP_SHEET = "物料描述查询表"
MATERIAL_DESCRIPTION_STILL_MISSING_SHEET = "物料描述仍缺失"


@dataclass
class MaterialDescriptionPrepareResult:
    missing_rows: int


@dataclass
class MaterialDescriptionFillResult:
    filled_rows: int
    remaining_rows: int
    lookup_rows: int


@dataclass
class ManualMaterialDescriptionApplyResult:
    applied_rows: int
    remaining_rows: int


@dataclass
class MissingMaterialDescriptionCleanupResult:
    deleted_rows: int


def prepare_material_descriptions(
    workbook: Workbook,
    values_workbook: Workbook,
    target_sheet_name: str,
    logger: ProcessingLogger,
) -> MaterialDescriptionPrepareResult:
    formula_sheet = workbook[target_sheet_name]
    values_sheet = values_workbook[target_sheet_name]
    row_numbers = find_missing_material_description_rows(values_sheet)
    refresh_material_description_supplement_sheet(workbook, formula_sheet, row_numbers, logger)
    logger.info(
        f"物料描述为空白或 #N/A 的行已摘取到“{MATERIAL_DESCRIPTION_SUPPLEMENT_SHEET}”：{len(row_numbers)} 行。"
    )
    return MaterialDescriptionPrepareResult(missing_rows=len(row_numbers))


def delete_missing_material_description_rows(
    workbook: Workbook,
    target_sheet_name: str,
    logger: ProcessingLogger,
) -> MissingMaterialDescriptionCleanupResult:
    formula_sheet = workbook[target_sheet_name]
    headers = _header_map(formula_sheet)
    _require_columns(headers, ["物料描述"], "主数据表")
    description_col = headers["物料描述"]
    rows_to_delete = [
        row_index
        for row_index in range(2, formula_sheet.max_row + 1)
        if _is_missing_description(formula_sheet.cell(row_index, description_col).value)
    ]
    for row_index in reversed(rows_to_delete):
        formula_sheet.delete_rows(row_index, 1)

    refresh_material_description_supplement_sheet(workbook, formula_sheet, [], logger)
    create_still_missing_sheet(workbook, formula_sheet, [], logger)
    logger.info(f"已删除物料描述缺失的未上单订单行：{len(rows_to_delete)} 行。")
    return MissingMaterialDescriptionCleanupResult(deleted_rows=len(rows_to_delete))


def fill_material_descriptions(
    workbook: Workbook,
    values_workbook: Workbook,
    target_sheet_name: str,
    lookup_path: Path,
    logger: ProcessingLogger,
) -> MaterialDescriptionFillResult:
    lookup_pairs = read_material_description_lookup(lookup_path, logger)
    lookup_map: dict[Any, Any] = {}
    conflicts = 0
    for code, description in lookup_pairs:
        if _is_blank(code) or _is_missing_description(description):
            continue
        code_key = _code_key(code)
        if code_key in lookup_map and lookup_map[code_key] != description:
            conflicts += 1
            continue
        lookup_map[code_key] = description

    if not lookup_map:
        raise ClassifierError("物料描述查询表没有可用的物料编码和物料描述。")

    import_material_description_lookup_sheet(workbook, lookup_pairs, logger)

    formula_sheet = workbook[target_sheet_name]
    values_sheet = values_workbook[target_sheet_name]
    headers = _header_map(values_sheet)
    _require_columns(headers, ["物料编码", "物料描述"], "主数据表")
    code_col = headers["物料编码"]
    description_col = headers["物料描述"]

    remaining_rows: list[int] = []
    filled_rows = 0
    for row_index in range(2, values_sheet.max_row + 1):
        current_description = values_sheet.cell(row_index, description_col).value
        if not _is_missing_description(current_description):
            continue
        code = values_sheet.cell(row_index, code_col).value
        matched_description = lookup_map.get(_code_key(code))
        if matched_description is None:
            remaining_rows.append(row_index)
            continue
        formula_sheet.cell(row_index, description_col).value = matched_description
        filled_rows += 1

    create_still_missing_sheet(workbook, formula_sheet, remaining_rows, logger)
    refresh_material_description_supplement_sheet(workbook, formula_sheet, remaining_rows, logger)
    if conflicts:
        logger.warning(f"物料描述查询表存在同一物料编码对应多个物料描述的情况，已保留首次匹配：{conflicts} 条。")
    logger.info(f"物料描述已按物料编码回填：{filled_rows} 行；仍待补充 {len(remaining_rows)} 行。")
    return MaterialDescriptionFillResult(
        filled_rows=filled_rows,
        remaining_rows=len(remaining_rows),
        lookup_rows=len(lookup_pairs),
    )


def fill_material_descriptions_for_invalid_sheet_metal(
    workbook: Workbook,
    values_workbook: Workbook,
    target_sheet_name: str,
    lookup_path: Path,
    logger: ProcessingLogger,
) -> MaterialDescriptionFillResult:
    """Backfill descriptions when sheet metal is #N/A/N/A/zero, even if a description exists."""
    lookup_pairs = read_material_description_lookup(lookup_path, logger)
    lookup_map: dict[str, Any] = {}
    conflicts = 0
    for code, description in lookup_pairs:
        if _is_blank(code) or _is_missing_description(description):
            continue
        code_key = _code_key(code)
        if code_key in lookup_map and lookup_map[code_key] != description:
            conflicts += 1
            continue
        lookup_map[code_key] = description

    if not lookup_map:
        raise ClassifierError("物料描述查询表没有可用的物料编码和物料描述。")

    import_material_description_lookup_sheet(workbook, lookup_pairs, logger)
    formula_sheet = workbook[target_sheet_name]
    values_sheet = values_workbook[target_sheet_name]
    headers = _header_map(values_sheet)
    _require_columns(headers, ["物料编码", "钣金型号", "物料描述"], "主数据表")
    code_col = headers["物料编码"]
    sheet_metal_col = headers["钣金型号"]
    description_col = headers["物料描述"]

    remaining_rows: list[int] = []
    filled_rows = 0
    for row_index in range(2, values_sheet.max_row + 1):
        sheet_metal = values_sheet.cell(row_index, sheet_metal_col).value
        if not _is_invalid_sheet_metal(sheet_metal):
            continue
        code = values_sheet.cell(row_index, code_col).value
        matched_description = lookup_map.get(_code_key(code))
        if matched_description is None:
            remaining_rows.append(row_index)
            continue
        formula_sheet.cell(row_index, description_col).value = matched_description
        values_sheet.cell(row_index, description_col).value = matched_description
        filled_rows += 1

    still_missing_rows = find_missing_material_description_rows(values_sheet)
    create_still_missing_sheet(workbook, formula_sheet, still_missing_rows, logger)
    refresh_material_description_supplement_sheet(workbook, formula_sheet, still_missing_rows, logger)
    if conflicts:
        logger.warning(f"物料描述查询表存在重复物料编码且描述不同，已保留首次匹配：{conflicts} 条。")
    logger.info(
        f"已按钣金型号异常条件回填物料描述：{filled_rows} 行；"
        f"钣金异常但未匹配 {len(remaining_rows)} 行。"
    )
    return MaterialDescriptionFillResult(
        filled_rows=filled_rows,
        remaining_rows=len(remaining_rows),
        lookup_rows=len(lookup_pairs),
    )


def apply_manual_material_descriptions(
    workbook: Workbook,
    target_sheet_name: str,
    logger: ProcessingLogger,
) -> ManualMaterialDescriptionApplyResult:
    source_sheet = find_manual_material_description_sheet(workbook)
    formula_sheet = workbook[target_sheet_name]
    source_headers = _header_map(source_sheet)
    main_headers = _header_map(formula_sheet)
    _require_columns(source_headers, ["原始行号", "物料描述"], source_sheet.title)
    _require_columns(main_headers, ["物料编码", "物料描述"], "主数据表")

    row_no_col = source_headers["原始行号"]
    source_description_col = source_headers["物料描述"]
    source_code_col = source_headers.get("物料编码")
    main_code_col = main_headers["物料编码"]
    main_description_col = main_headers["物料描述"]

    remaining_rows: list[int] = []
    applied_rows = 0
    if source_code_col:
        description_by_code: dict[str, Any] = {}
        remaining_codes: set[str] = set()
        invalid_original_rows: list[int] = []

        for row_index in range(2, source_sheet.max_row + 1):
            code = _code_key(source_sheet.cell(row_index, source_code_col).value)
            description = source_sheet.cell(row_index, source_description_col).value
            if _is_valid_manual_description(description):
                if code:
                    description_by_code[code] = description
            elif code:
                remaining_codes.add(code)
            else:
                original_row = source_sheet.cell(row_index, row_no_col).value
                if _is_blank(original_row):
                    continue
                try:
                    invalid_original_rows.append(int(original_row))
                except (TypeError, ValueError):
                    logger.warning(f"{source_sheet.title} 第 {row_index} 行原始行号无效，已跳过：{original_row}")

        if description_by_code:
            for row_index in range(2, formula_sheet.max_row + 1):
                code = _code_key(formula_sheet.cell(row_index, main_code_col).value)
                if code not in description_by_code:
                    continue
                if not _is_missing_description(formula_sheet.cell(row_index, main_description_col).value):
                    continue
                formula_sheet.cell(row_index, main_description_col).value = description_by_code[code]
                applied_rows += 1

        if remaining_codes:
            for row_index in range(2, formula_sheet.max_row + 1):
                code = _code_key(formula_sheet.cell(row_index, main_code_col).value)
                if code in remaining_codes and _is_missing_description(
                    formula_sheet.cell(row_index, main_description_col).value
                ):
                    remaining_rows.append(row_index)

        remaining_rows.extend(
            row_index
            for row_index in invalid_original_rows
            if 2 <= row_index <= formula_sheet.max_row
            and _is_missing_description(formula_sheet.cell(row_index, main_description_col).value)
        )
        remaining_rows = sorted(set(remaining_rows))
        create_still_missing_sheet(workbook, formula_sheet, remaining_rows, logger)
        refresh_material_description_supplement_sheet(workbook, formula_sheet, remaining_rows, logger)
        logger.info(
            f"已从“{source_sheet.title}”按物料编码回填手工物料描述：{applied_rows} 行；仍待补充 {len(remaining_rows)} 行。"
        )
        return ManualMaterialDescriptionApplyResult(applied_rows=applied_rows, remaining_rows=len(remaining_rows))

    for row_index in range(2, source_sheet.max_row + 1):
        original_row = source_sheet.cell(row_index, row_no_col).value
        description = source_sheet.cell(row_index, source_description_col).value
        if _is_blank(original_row):
            continue
        try:
            original_row_number = int(original_row)
        except (TypeError, ValueError):
            logger.warning(f"{source_sheet.title} 第 {row_index} 行原始行号无效，已跳过：{original_row}")
            continue

        if _is_valid_manual_description(description):
            formula_sheet.cell(original_row_number, main_description_col).value = description
            applied_rows += 1
        else:
            remaining_rows.append(original_row_number)

    create_still_missing_sheet(workbook, formula_sheet, remaining_rows, logger)
    refresh_material_description_supplement_sheet(workbook, formula_sheet, remaining_rows, logger)
    logger.info(
        f"已从“{source_sheet.title}”回填手工物料描述：{applied_rows} 行；仍待补充 {len(remaining_rows)} 行。"
    )
    return ManualMaterialDescriptionApplyResult(applied_rows=applied_rows, remaining_rows=len(remaining_rows))


def find_missing_material_description_rows(values_sheet: Worksheet) -> list[int]:
    headers = _header_map(values_sheet)
    _require_columns(headers, ["物料描述"], "主数据表")
    description_col = headers["物料描述"]
    return [
        row_index
        for row_index in range(2, values_sheet.max_row + 1)
        if _is_missing_description(values_sheet.cell(row_index, description_col).value)
    ]


def read_material_description_lookup(lookup_path: Path, logger: ProcessingLogger) -> list[tuple[Any, Any]]:
    if not lookup_path.exists():
        raise FileNotFoundError(f"物料描述查询表不存在：{lookup_path}")

    suffix = lookup_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        rows = read_lookup_rows_from_xlsx(lookup_path)
    else:
        rows = read_lookup_rows_from_text_export(lookup_path)
    logger.info(f"已读取物料描述查询表：{len(rows)} 条。")
    return rows


def read_lookup_rows_from_xlsx(lookup_path: Path) -> list[tuple[Any, Any]]:
    lookup_wb = openpyxl.load_workbook(lookup_path, data_only=True)
    source_sheet = lookup_wb[lookup_wb.sheetnames[0]]
    source_headers = _header_map(source_sheet)
    _require_columns(source_headers, ["物料编码", "物料描述"], "物料描述查询表")
    code_col = source_headers["物料编码"]
    description_col = source_headers["物料描述"]

    rows: list[tuple[Any, Any]] = []
    for row_index in range(2, source_sheet.max_row + 1):
        code = source_sheet.cell(row_index, code_col).value
        description = source_sheet.cell(row_index, description_col).value
        if _is_blank(code):
            continue
        rows.append((code, description))
    return rows


def read_lookup_rows_from_text_export(lookup_path: Path) -> list[tuple[Any, Any]]:
    data = lookup_path.read_bytes()
    for encoding in ("utf-16", "utf-8-sig", "gb18030"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ClassifierError(f"无法识别物料描述查询表编码：{lookup_path}")

    parsed_rows = list(csv.reader(text.splitlines(), delimiter="\t"))
    if not parsed_rows:
        raise ClassifierError("物料描述查询表为空。")
    header_row_index, source_headers = _find_text_export_header_row(parsed_rows)
    _require_text_columns(source_headers, ["物料编码", "物料描述"], "物料描述查询表")

    code_index = source_headers["物料编码"]
    description_index = source_headers["物料描述"]
    rows: list[tuple[Any, Any]] = []
    for row in parsed_rows[header_row_index + 1:]:
        code = row[code_index] if code_index < len(row) else None
        description = row[description_index] if description_index < len(row) else None
        if _is_blank(code):
            continue
        rows.append((code, description))
    return rows


def _find_text_export_header_row(parsed_rows: list[list[str]]) -> tuple[int, dict[str, int]]:
    for row_index, row in enumerate(parsed_rows[:50]):
        headers = _normalized_header_map(row)
        if "物料编码" in headers and "物料描述" in headers:
            return row_index, headers
    raise MissingRequiredFieldsError("物料描述查询表缺少关键字段：物料编码、物料描述")


def import_material_description_lookup_sheet(
    workbook: Workbook,
    lookup_pairs: list[tuple[Any, Any]],
    logger: ProcessingLogger,
) -> None:
    remove_sheet_if_exists(workbook, MATERIAL_DESCRIPTION_LOOKUP_SHEET)
    sheet = workbook.create_sheet(MATERIAL_DESCRIPTION_LOOKUP_SHEET)
    sheet.append(["物料编码", "物料描述"])
    for code, description in lookup_pairs:
        sheet.append([code, description])
    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = sheet.dimensions
    style_header(sheet)
    _fit_columns(sheet)
    logger.info(f"已导入“{MATERIAL_DESCRIPTION_LOOKUP_SHEET}”：{len(lookup_pairs)} 条。")


def create_still_missing_sheet(
    workbook: Workbook,
    formula_sheet: Worksheet,
    row_numbers: list[int],
    logger: ProcessingLogger,
) -> None:
    remove_sheet_if_exists(workbook, MATERIAL_DESCRIPTION_STILL_MISSING_SHEET)
    sheet = workbook.create_sheet(MATERIAL_DESCRIPTION_STILL_MISSING_SHEET)
    _append_rows_by_number(sheet, formula_sheet, row_numbers)
    logger.info(f"已生成“{MATERIAL_DESCRIPTION_STILL_MISSING_SHEET}”：{len(row_numbers)} 行。")


def refresh_material_description_supplement_sheet(
    workbook: Workbook,
    formula_sheet: Worksheet,
    row_numbers: list[int],
    logger: ProcessingLogger,
) -> None:
    remove_sheet_if_exists(workbook, MATERIAL_DESCRIPTION_SUPPLEMENT_SHEET)
    sheet = workbook.create_sheet(MATERIAL_DESCRIPTION_SUPPLEMENT_SHEET)
    _append_rows_by_number(sheet, formula_sheet, row_numbers)
    logger.info(f"已刷新“{MATERIAL_DESCRIPTION_SUPPLEMENT_SHEET}”：剩余待补充 {len(row_numbers)} 行。")


def find_manual_material_description_sheet(workbook: Workbook) -> Worksheet:
    for sheet_name in (
        "物料描述待补充",
        MATERIAL_DESCRIPTION_STILL_MISSING_SHEET,
        MATERIAL_DESCRIPTION_SUPPLEMENT_SHEET,
    ):
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
    raise ClassifierError("未找到可用于手工物料描述回填的工作表：物料描述待补充、物料描述仍缺失 或 物料描述补充")


def _append_rows_by_number(sheet: Worksheet, formula_sheet: Worksheet, row_numbers: list[int]) -> None:
    output_headers = ["原始行号"] + [
        normalize_header(cell.value) or f"空列{index}"
        for index, cell in enumerate(formula_sheet[1], start=1)
    ]
    sheet.append(output_headers)
    for row_index in row_numbers:
        row_values = [
            formula_sheet.cell(row_index, col_index).value
            for col_index in range(1, formula_sheet.max_column + 1)
        ]
        sheet.append([row_index] + row_values)

    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = sheet.dimensions
    style_header(sheet)
    _fit_columns(sheet)


def _header_map(sheet: Worksheet) -> dict[str, int]:
    raw_headers = [normalize_header(cell.value) for cell in sheet[1]]
    headers = deduplicate_headers(raw_headers)
    return {header: index + 1 for index, header in enumerate(headers)}


def _normalized_header_map(headers: list[Any]) -> dict[str, int]:
    return {normalize_header(header): index for index, header in enumerate(headers)}


def _require_columns(header_map: dict[str, int], columns: list[str], context: str) -> None:
    missing = [column for column in columns if column not in header_map]
    if missing:
        raise MissingRequiredFieldsError(f"{context}缺少关键字段：" + "、".join(missing))


def _require_text_columns(header_map: dict[str, int], columns: list[str], context: str) -> None:
    missing = [column for column in columns if column not in header_map]
    if missing:
        raise MissingRequiredFieldsError(f"{context}缺少关键字段：" + "、".join(missing))


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _code_key(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_na(value: Any) -> bool:
    return value is not None and str(value).strip().upper() in {"#N/A", "#NA", "N/A"}


def _is_invalid_sheet_metal(value: Any) -> bool:
    if _is_na(value):
        return True
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value == 0
    return str(value).strip() in {"0", "0.0", "0.00"}


def _is_missing_description(value: Any) -> bool:
    return _is_blank(value) or _is_na(value)


def _is_valid_manual_description(value: Any) -> bool:
    if _is_missing_description(value):
        return False
    if isinstance(value, str) and value.strip().startswith("="):
        return False
    return True


def _fit_columns(sheet: Worksheet) -> None:
    for column_cells in sheet.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 45)
