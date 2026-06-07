"""Excel workbook reading and output helpers."""

from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Any
import re

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .data_cleaner import deduplicate_headers, normalize_header, validate_required_fields
from .logger import ProcessingLogger


def build_output_path(input_path: Path, output_dir: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir.mkdir(parents=True, exist_ok=True)
    base_stem = re.sub(r"_分类结果_\d{8}_\d{6}$", "", input_path.stem)
    return output_dir / f"{base_stem}_分类结果_{timestamp}{input_path.suffix}"


def copy_workbook(input_path: Path, output_path: Path, logger: ProcessingLogger) -> None:
    shutil.copy2(input_path, output_path)
    logger.info(f"已生成处理副本：{output_path}")


def load_workbook_pair(output_path: Path) -> tuple[Workbook, Workbook]:
    formula_wb = openpyxl.load_workbook(output_path)
    values_wb = openpyxl.load_workbook(output_path, data_only=True)
    return formula_wb, values_wb


def read_main_table(values_sheet: Worksheet) -> tuple[list[str], list[dict[str, Any]], list[list[Any]]]:
    raw_headers = [normalize_header(cell.value) for cell in values_sheet[1]]
    headers = deduplicate_headers(raw_headers)
    validate_required_fields(headers)

    rows: list[dict[str, Any]] = []
    raw_rows: list[list[Any]] = []
    for row in values_sheet.iter_rows(min_row=2, max_row=values_sheet.max_row, values_only=True):
        row_values = list(row[: len(headers)])
        if not any(value is not None and value != "" for value in row_values):
            continue
        raw_rows.append(row_values)
        rows.append({headers[index]: row_values[index] if index < len(row_values) else None for index in range(len(headers))})

    return headers, rows, raw_rows


def remove_sheet_if_exists(workbook: Workbook, sheet_name: str) -> None:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]


AUXILIARY_SHEET_ORDER = [
    "排单分解表明细",
    "系数补充",
    "系数查询表",
    "系数仍缺失",
    "钣金型号补充",
    "钣金型号查询表",
]


def move_auxiliary_sheets_after(workbook: Workbook, anchor_sheet_name: str) -> None:
    """Move frequently reviewed helper sheets directly after the main detail sheet."""
    if anchor_sheet_name not in workbook.sheetnames:
        return

    sheets_by_name = {sheet.title: sheet for sheet in workbook._sheets}
    sheets_to_move = [
        sheets_by_name[name]
        for name in AUXILIARY_SHEET_ORDER
        if name in sheets_by_name and name != anchor_sheet_name
    ]
    if not sheets_to_move:
        return

    remaining = [sheet for sheet in workbook._sheets if sheet not in sheets_to_move]
    anchor_index = next(
        index for index, sheet in enumerate(remaining) if sheet.title == anchor_sheet_name
    )
    workbook._sheets = (
        remaining[: anchor_index + 1]
        + sheets_to_move
        + remaining[anchor_index + 1 :]
    )


INVALID_SHEET_CHARS = re.compile(r"[\[\]\*:/\\?]")


def sanitize_sheet_name(name: str) -> str:
    return INVALID_SHEET_CHARS.sub("_", name).strip()


def safe_result_sheet_name(workbook: Workbook, desired_name: str, protected_names: set[str], logger: ProcessingLogger) -> str:
    raw_base = desired_name.strip() or "分类结果"
    base = sanitize_sheet_name(raw_base) or "分类结果"
    if base != raw_base:
        logger.warning(f"输出表名“{raw_base}”包含 Excel 禁用字符，已改为“{base}”。")
    if base in protected_names:
        renamed = f"分类_{base}"
        logger.warning(f"输出表名“{base}”与原工作簿已有工作表冲突，改为“{renamed}”。")
        base = renamed

    name = base[:31]
    if name not in workbook.sheetnames:
        return name

    for index in range(2, 100):
        suffix = f"_{index}"
        candidate = f"{name[:31 - len(suffix)]}{suffix}"
        if candidate not in workbook.sheetnames:
            return candidate

    raise ValueError(f"无法生成唯一工作表名称：{desired_name}")


def write_table(sheet: Worksheet, headers: list[str], rows: list[dict[str, Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header) for header in headers])
    if sheet.max_row > 1:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
