"""Write classification result sheets, summary, and logs."""

from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .classifier import ClassificationResult
from .data_cleaner import to_number
from .excel_io import remove_sheet_if_exists, safe_result_sheet_name, write_table
from .logger import ProcessingLogger

RESULT_SHEETS = {"分类结果汇总表", "处理日志", "未分类数据"}


def style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font


def compact_unique(values: list[Any]) -> str:
    cleaned = sorted({str(value).strip() for value in values if value is not None and str(value).strip()})
    return "、".join(cleaned)


def build_summary_rows(result: ClassificationResult) -> list[list[Any]]:
    rows: list[list[Any]] = []
    has_unmatched = bool(result.unmatched)
    for category_name, category_rows in result.categories.items():
        rows.append(
            [
                category_name,
                result.output_names.get(category_name, category_name),
                len(category_rows),
                sum(to_number(row.get("订单数")) for row in category_rows),
                compact_unique([row.get("线体") for row in category_rows]),
                compact_unique([row.get("周次") for row in category_rows]),
                "是" if has_unmatched else "否",
            ]
        )
    rows.append(
        [
            "未分类数据",
            "未分类数据",
            len(result.unmatched),
            sum(to_number(row.get("订单数")) for row in result.unmatched),
            compact_unique([row.get("线体") for row in result.unmatched]),
            compact_unique([row.get("周次") for row in result.unmatched]),
            "是" if has_unmatched else "否",
        ]
    )
    return rows


def write_type_summary_results(
    workbook: Workbook,
    headers: list[str],
    rows: list[dict[str, Any]],
    logger: ProcessingLogger,
) -> None:
    """Write the final summary directly from the main-sheet 类型 values.

    This deliberately does not invoke keyword rules or create per-keyword
    category sheets. Rows with an empty 类型 stay auditable in 未分类数据.
    """
    for sheet_name in RESULT_SHEETS:
        remove_sheet_if_exists(workbook, sheet_name)

    totals: dict[str, list[float | int]] = {}
    unmatched: list[dict[str, Any]] = []
    for row in rows:
        type_name = str(row.get("类型") or "").strip()
        if not type_name:
            unmatched.append(row)
            continue
        bucket = totals.setdefault(type_name, [0.0, 0.0, 0])
        bucket[0] += to_number(row.get("订单数"))
        bucket[1] += to_number(row.get("标台数"))
        bucket[2] += 1

    summary_sheet = workbook.create_sheet("分类结果汇总表")
    summary_sheet.append(["类型", "订单数合计", "标台数合计", "行数"])
    for type_name in sorted(totals):
        order_qty, standard_units, row_count = totals[type_name]
        summary_sheet.append([type_name, order_qty, standard_units, row_count])
    summary_sheet.freeze_panes = "A2"
    summary_sheet.auto_filter.ref = summary_sheet.dimensions
    style_header(summary_sheet)
    for column in ("B", "C", "D"):
        for cell in summary_sheet[column][1:]:
            cell.number_format = "#,##0.##"

    unmatched_sheet = workbook.create_sheet("未分类数据")
    write_table(unmatched_sheet, headers, unmatched)
    style_header(unmatched_sheet)
    unmatched_sheet.freeze_panes = "A2"
    unmatched_sheet.auto_filter.ref = unmatched_sheet.dimensions

    log_sheet = workbook.create_sheet("处理日志")
    write_log_rows(log_sheet, logger)

    for sheet in (summary_sheet, unmatched_sheet, log_sheet):
        for column_cells in sheet.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 45)

    logger.info(
        f"已按主表‘类型’直接汇总：{len(totals)} 个类型；"
        f"类型为空的未分类数据 {len(unmatched)} 行。"
    )


def write_results(
    workbook: Workbook,
    headers: list[str],
    result: ClassificationResult,
    original_sheet_names: set[str],
    logger: ProcessingLogger,
) -> None:
    for sheet_name in RESULT_SHEETS:
        remove_sheet_if_exists(workbook, sheet_name)

    written_result_names: set[str] = set()
    for category_name, rows in result.categories.items():
        desired_name = result.output_names.get(category_name, category_name)
        sheet_name = safe_result_sheet_name(
            workbook,
            desired_name,
            original_sheet_names | RESULT_SHEETS,
            logger,
        )
        result.output_names[category_name] = sheet_name
        written_result_names.add(sheet_name)
        sheet = workbook.create_sheet(sheet_name)
        write_table(sheet, list(rows[0].keys()) if rows and set(rows[0].keys()) != set(headers) else headers, rows)
        style_header(sheet)
        logger.info(f"生成分类工作表：{sheet_name}，数据行数 {len(rows)}。")

    unmatched_sheet = workbook.create_sheet("未分类数据")
    write_table(unmatched_sheet, headers, result.unmatched)
    style_header(unmatched_sheet)
    logger.info(f"生成未分类数据工作表，数据行数 {len(result.unmatched)}。")

    summary_sheet = workbook.create_sheet("分类结果汇总表")
    summary_headers = ["分类名称", "输出工作表名称", "数据行数", "订单数合计", "涉及线体", "涉及周次", "是否存在未分类"]
    summary_sheet.append(summary_headers)
    for row in build_summary_rows(result):
        summary_sheet.append(row)
    summary_sheet.freeze_panes = "A2"
    summary_sheet.auto_filter.ref = summary_sheet.dimensions
    style_header(summary_sheet)

    log_sheet = workbook.create_sheet("处理日志")
    write_log_rows(log_sheet, logger)

    for sheet in workbook.worksheets:
        if sheet.title in written_result_names | RESULT_SHEETS:
            for column_cells in sheet.columns:
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 45)


def write_log_sheet(workbook: Workbook, logger: ProcessingLogger) -> None:
    remove_sheet_if_exists(workbook, "处理日志")
    log_sheet = workbook.create_sheet("处理日志")
    write_log_rows(log_sheet, logger)
    for column_cells in log_sheet.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        log_sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 80)


def write_log_rows(log_sheet, logger: ProcessingLogger) -> None:
    log_sheet.append(["时间", "级别", "消息"])
    for entry in logger.entries:
        log_sheet.append([entry.timestamp, entry.level, entry.message])
    log_sheet.freeze_panes = "A2"
    log_sheet.auto_filter.ref = log_sheet.dimensions
    style_header(log_sheet)
