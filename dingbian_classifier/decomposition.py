"""Production order decomposition rules and summary sheet."""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .data_cleaner import deduplicate_headers, normalize_header
from .excel_io import remove_sheet_if_exists
from .exceptions import MissingRequiredFieldsError
from .logger import ProcessingLogger

DETAIL_SHEET_NAME = "排单分解表明细"
LINE_CLASSIFICATION_SHEET_NAME = "各线体分类明细表"
WAVE_LINES = {"B线", "B线夜", "C线", "C线夜"}
ROLLING_LINE_ORDER = ["A线", "A线夜", "D线", "D线夜", "E线", "H线"]
WAVE_LINE_ORDER = ["B线", "B线夜", "C线", "C线夜"]
DECOMPOSE_COLUMNS = ["基本开始日期", "备注", "类型"]

SKD_MAIN_FILL = "00B050"
SKD_NORMAL_TYPE_FILL = "92D050"
SKD_DRY_TYPE_FILL = "F4CCCC"
CKD_MAIN_FILL = "FFC000"
CKD_NORMAL_TYPE_FILL = "FFD966"
CKD_DRUM_TYPE_FILL = "F4B183"
SAMSUNG_MAIN_FILL = "9DC3E6"
DOUBLE_DRUM_MAIN_FILL = "A9D18E"
T7P7_MAIN_FILL = "7030A0"
T7P7_WASH_TYPE_FILL = "D9EAD3"
T7P7_DRY_TYPE_FILL = "D9D2E9"
T7P7_WASH_TYPE = "T7/P7/T5/P5/追觅单洗"
T7P7_DRY_TYPE = "T7/P7/T5/P5/追觅烘干"
T9P9_MAIN_FILL = "00B0F0"
T9P9_WASH_TYPE_FILL = "C9DAF8"
T9P9_DRY_TYPE_FILL = "FCE4D6"
T9P9_WASH_TYPE = "T9/P9单洗"
T9P9_DRY_TYPE = "T9/P9烘干"
T9P9_DRYER_TYPE = "T9/P9干衣机"
T9P9_DRYER_TYPE_FILL = "DDEBF7"
T10P10_MAIN_FILL = "C00000"
T10P10_WASH_TYPE_FILL = "F4CCCC"
T10P10_DRYER_TYPE_FILL = "EADCF8"
T10P10_WASH_TYPE = "T10/P10洗衣机"
T10P10_DRYER_TYPE = "T10/P10干衣机"
C6_HEAT_PUMP_DRYER_TYPE = "C6热泵干衣机"
C6_HEAT_PUMP_DRYER_MAIN_FILL = "8064A2"
C6_HEAT_PUMP_DRYER_TYPE_FILL = "D9E2F3"
COMPOSITE_MAIN_FILL = "B4A7D6"
COMPOSITE_DRY_TYPE_FILL = "D9D2E9"
COMPOSITE_WASH_TYPE_FILL = "EADCF8"
COMPOSITE_DRY_TYPE = "复式烘干"
COMPOSITE_WASH_TYPE = "复式单洗"
PENGUIN_MAIN_FILL = "A6A6A6"
PENGUIN_DRYER_TYPE_FILL = "D9D9D9"
PENGUIN_WASH_TYPE_FILL = "E7E6E6"
PENGUIN_DRYER_TYPE = "企鹅干衣机"
PENGUIN_WASH_TYPE = "企鹅洗衣机"
C6_MAIN_FILL = "548235"
C6_WASH_TYPE_FILL = "D9EAD3"
C6_DRY_TYPE_FILL = "E2F0D9"
C6_WASH_TYPE = "C6单洗"
C6_DRY_TYPE = "C6烘干"
ORDINARY_DRY_MAIN_FILL = "F4B183"
ORDINARY_DRY_TYPE_FILL = "FCE4D6"
ORDINARY_DRY_TYPE = "普通烘干"
DOMESTIC_MAIN_FILL = "9EADCC"
DOMESTIC_TYPE_FILL = "D9E2F3"
DOMESTIC_TYPE = "普通内销"
EXPORT_MAIN_FILL = "B7DEE8"
EXPORT_TYPE_FILL = "DDEBF7"
EXPORT_TYPE = "外销"
WAVE_LG_TYPE = "LG"
WAVE_PLASTIC_DOMESTIC_TYPE = "塑料内销"
WAVE_P7P9_TYPE = "P7/P9"
WAVE_P7P9_LABEL = "P7P9"
WAVE_CKD_MAIN_FILL = "C6E0B4"
WAVE_CKD_TYPE_FILL = "E2F0D9"
WAVE_LG_MAIN_FILL = "F8CBAD"
WAVE_LG_TYPE_FILL = "FCE4D6"
WAVE_PLASTIC_MAIN_FILL = "BDD7EE"
WAVE_PLASTIC_TYPE_FILL = "DDEBF7"
WAVE_P7P9_MAIN_FILL = "D9EAD3"
WAVE_P7P9_TYPE_FILL = "E2F0D9"
WAVE_SKD_MAIN_FILL = "D9D2E9"
WAVE_SKD_TYPE_FILL = "EADCF8"
WAVE_BASIC_TYPES = {"CKD", WAVE_LG_TYPE, WAVE_PLASTIC_DOMESTIC_TYPE, WAVE_P7P9_TYPE, "SKD"}
WAVE_DOMESTIC_INVERTER_TYPE = "内销铁皮变频"
WAVE_EXPORT_INVERTER_TYPE = "外销普通变频"
WAVE_DOMESTIC_IRON_TYPE = "内销铁皮"
WAVE_EXPORT_IRON_TYPE = "外销铁皮"
WAVE_DOMESTIC_INVERTER_MAIN_FILL = "E2F0D9"
WAVE_DOMESTIC_INVERTER_TYPE_FILL = "C6E0B4"
WAVE_EXPORT_INVERTER_MAIN_FILL = "FFF2CC"
WAVE_EXPORT_INVERTER_TYPE_FILL = "FFE699"
WAVE_DOMESTIC_IRON_MAIN_FILL = "D9E2F3"
WAVE_DOMESTIC_IRON_TYPE_FILL = "BDD7EE"
WAVE_EXPORT_IRON_MAIN_FILL = "EADCF8"
WAVE_EXPORT_IRON_TYPE_FILL = "D9D2E9"
WAVE_FINAL_TYPES = {
    WAVE_DOMESTIC_INVERTER_TYPE,
    WAVE_EXPORT_INVERTER_TYPE,
    WAVE_DOMESTIC_IRON_TYPE,
    WAVE_EXPORT_IRON_TYPE,
}

ROLLING_LABELS = [
    "SKD烘干",
    "SKD总数(包含烘干)",
    "CKD",
    "CKD含筒部装",
    "三星",
    "双滚筒",
    T7P7_DRY_TYPE,
    T7P7_WASH_TYPE,
    T9P9_WASH_TYPE,
    T9P9_DRY_TYPE,
    T10P10_WASH_TYPE,
    T10P10_DRYER_TYPE,
    C6_HEAT_PUMP_DRYER_TYPE,
    T9P9_DRYER_TYPE,
    COMPOSITE_DRY_TYPE,
    COMPOSITE_WASH_TYPE,
    PENGUIN_DRYER_TYPE,
    PENGUIN_WASH_TYPE,
    C6_WASH_TYPE,
    C6_DRY_TYPE,
    ORDINARY_DRY_TYPE,
    "内销",
    EXPORT_TYPE,
]
WAVE_LABELS = [
    "CKD",
    WAVE_LG_TYPE,
    WAVE_PLASTIC_DOMESTIC_TYPE,
    WAVE_P7P9_LABEL,
    "SKD",
    WAVE_DOMESTIC_INVERTER_TYPE,
    WAVE_EXPORT_INVERTER_TYPE,
    WAVE_EXPORT_IRON_TYPE,
    WAVE_DOMESTIC_IRON_TYPE,
]


@dataclass
class DecomposeSkdResult:
    rolling_total: Decimal
    wave_total: Decimal
    skd_total: Decimal
    skd_dry: Decimal
    skd_rows: int
    skd_dry_rows: int


@dataclass
class DecomposeRollingRemarkResult:
    rolling_total: Decimal
    wave_total: Decimal
    category_totals: dict[str, Decimal] = field(default_factory=dict)
    category_rows: dict[str, int] = field(default_factory=dict)


@dataclass
class DecomposeT7P7T5P5DreameResult:
    rolling_total: Decimal
    wave_total: Decimal
    category_totals: dict[str, Decimal] = field(default_factory=dict)
    category_rows: dict[str, int] = field(default_factory=dict)
    skipped_dryer_rows: int = 0
    skipped_unknown_code_rows: int = 0


@dataclass
class DecomposeT9P9Result:
    rolling_total: Decimal
    wave_total: Decimal
    category_totals: dict[str, Decimal] = field(default_factory=dict)
    category_rows: dict[str, int] = field(default_factory=dict)
    skipped_dryer_rows: int = 0
    skipped_unknown_code_rows: int = 0
    skipped_remark_only_rows: int = 0


@dataclass
class DecomposeT9P9DryerResult:
    rolling_total: Decimal
    wave_total: Decimal
    category_totals: dict[str, Decimal] = field(default_factory=dict)
    category_rows: dict[str, int] = field(default_factory=dict)
    skipped_non_dryer_rows: int = 0
    skipped_unknown_code_rows: int = 0


@dataclass
class DecomposeT10P10Result:
    rolling_total: Decimal
    wave_total: Decimal
    category_totals: dict[str, Decimal] = field(default_factory=dict)
    category_rows: dict[str, int] = field(default_factory=dict)
    skipped_unknown_code_rows: int = 0


@dataclass
class DecomposeC6HeatPumpDryerResult:
    rolling_total: Decimal
    wave_total: Decimal
    category_totals: dict[str, Decimal] = field(default_factory=dict)
    category_rows: dict[str, int] = field(default_factory=dict)
    skipped_non_dryer_rows: int = 0
    skipped_unknown_code_rows: int = 0


@dataclass
class DecomposeCompositePenguinC6Result:
    rolling_total: Decimal
    wave_total: Decimal
    category_totals: dict[str, Decimal] = field(default_factory=dict)
    category_rows: dict[str, int] = field(default_factory=dict)
    skipped_unknown_code_rows: int = 0
    skipped_non_c6_dryer_rows: int = 0


@dataclass
class DecomposeRollingFinalResult:
    rolling_total: Decimal
    wave_total: Decimal
    category_totals: dict[str, Decimal] = field(default_factory=dict)
    category_rows: dict[str, int] = field(default_factory=dict)
    remaining_unclassified_rows: int = 0
    remaining_unclassified_total: Decimal = Decimal("0")
    rolling_decomposition_total: Decimal = Decimal("0")
    rolling_gap: Decimal = Decimal("0")


@dataclass
class DecomposeWaveBasicResult:
    rolling_total: Decimal
    wave_total: Decimal
    category_totals: dict[str, Decimal] = field(default_factory=dict)
    category_rows: dict[str, int] = field(default_factory=dict)
    remaining_unclassified_rows: int = 0
    remaining_unclassified_total: Decimal = Decimal("0")


@dataclass
class DecomposeWaveFinalResult:
    rolling_total: Decimal
    wave_total: Decimal
    category_totals: dict[str, Decimal] = field(default_factory=dict)
    category_rows: dict[str, int] = field(default_factory=dict)
    remaining_unclassified_rows: int = 0
    remaining_unclassified_total: Decimal = Decimal("0")
    wave_decomposition_total: Decimal = Decimal("0")
    wave_gap: Decimal = Decimal("0")


@dataclass(frozen=True)
class ExtraOrderSummaryRule:
    name: str
    scope: str
    fill_rgb: str
    criteria: str


@dataclass
class ExtraOrderSummaryResult:
    category_totals: dict[str, Decimal] = field(default_factory=dict)
    category_rows: dict[str, int] = field(default_factory=dict)


EXTRA_ORDER_SUMMARY_RULES = [
    ExtraOrderSummaryRule(
        "锥形筒",
        "rolling",
        "FFF2CC",
        "滚筒线；钣金型号含“锥形筒”",
    ),
    ExtraOrderSummaryRule(
        "波轮特殊内筒-10kg和9升10内筒",
        "wave",
        "D9EAD3",
        "波轮线；钣金型号含“10kg波轮”或“9升10”",
    ),
    ExtraOrderSummaryRule(
        "波轮特殊内筒-8升9内筒",
        "wave",
        "DDEBF7",
        "波轮线；钣金型号含“8升9”",
    ),
    ExtraOrderSummaryRule(
        "波轮箱体-10kg",
        "wave",
        "EADCF8",
        "波轮线；钣金型号含“10kg”或“10KG”",
    ),
    ExtraOrderSummaryRule(
        "波轮箱体-彩板",
        "wave",
        "FCE4D6",
        "波轮线；钣金型号含“PCM”；剔除波轮CKD",
    ),
]


def decompose_skd(workbook: Workbook, main_sheet_name: str, logger: ProcessingLogger) -> DecomposeSkdResult:
    sheet = workbook[main_sheet_name]
    headers = _header_map(sheet)
    _require_columns(headers, ["线体", "订单数", "备注", "类型", *DECOMPOSE_COLUMNS], "主数据表")

    rolling_total = Decimal("0")
    wave_total = Decimal("0")
    skd_total = Decimal("0")
    skd_dry = Decimal("0")
    skd_rows = 0
    skd_dry_rows = 0

    for row_index in range(2, sheet.max_row + 1):
        line = str(sheet.cell(row_index, headers["线体"]).value or "").strip()
        order_qty = _to_decimal(sheet.cell(row_index, headers["订单数"]).value)
        if _is_wave_line(line):
            wave_total += order_qty
            continue

        rolling_total += order_qty
        if not _is_uncolored_for_decomposition(sheet, row_index, headers):
            continue

        remark = str(sheet.cell(row_index, headers["备注"]).value or "")
        if "SKD" not in remark.upper():
            continue

        is_dry = "烘干" in remark
        type_name = "SKD烘干" if is_dry else "SKD"
        _apply_skd_style(sheet, row_index, headers, type_name, is_dry)

        skd_total += order_qty
        skd_rows += 1
        if is_dry:
            skd_dry += order_qty
            skd_dry_rows += 1

    summary = _collect_decomposition_summary(sheet, headers)
    write_decomposition_detail_sheet(workbook, main_sheet_name, summary, logger)
    logger.info(
        f"SKD排单分解完成：命中 {skd_rows} 行，SKD总数 {skd_total}，SKD烘干 {skd_dry}。"
    )
    return DecomposeSkdResult(
        rolling_total=rolling_total,
        wave_total=wave_total,
        skd_total=skd_total,
        skd_dry=skd_dry,
        skd_rows=skd_rows,
        skd_dry_rows=skd_dry_rows,
    )


def decompose_rolling_remark_rules(
    workbook: Workbook,
    values_workbook: Workbook,
    main_sheet_name: str,
    logger: ProcessingLogger,
) -> DecomposeRollingRemarkResult:
    """Run rolling-line remark rules after SKD: CKD, Samsung, and double-drum."""
    sheet = workbook[main_sheet_name]
    values_sheet = values_workbook[main_sheet_name]
    headers = _header_map(sheet)
    _require_columns(headers, ["线体", "订单数", "备注", "类型", "系数", *DECOMPOSE_COLUMNS], "主数据表")

    category_totals: dict[str, Decimal] = {"CKD": Decimal("0"), "CKD含筒部装": Decimal("0"), "三星": Decimal("0"), "双滚筒": Decimal("0")}
    category_rows: dict[str, int] = {"CKD": 0, "CKD含筒部装": 0, "三星": 0, "双滚筒": 0}

    for row_index in range(2, sheet.max_row + 1):
        line = str(sheet.cell(row_index, headers["线体"]).value or "").strip()
        if _is_wave_line(line) or not _is_uncolored_for_decomposition(sheet, row_index, headers):
            continue

        remark = str(sheet.cell(row_index, headers["备注"]).value or "")
        remark_upper = remark.upper()
        order_qty = _to_decimal(sheet.cell(row_index, headers["订单数"]).value)

        if "CKD" in remark_upper:
            coefficient = _to_decimal(values_sheet.cell(row_index, headers["系数"]).value)
            type_name = "CKD含筒部装" if coefficient > Decimal("1.6") else "CKD"
            type_fill = CKD_DRUM_TYPE_FILL if type_name == "CKD含筒部装" else CKD_NORMAL_TYPE_FILL
            _apply_decomposition_style(sheet, row_index, headers, type_name, CKD_MAIN_FILL, type_fill)
        elif "三星" in remark:
            type_name = "三星"
            _apply_decomposition_style(sheet, row_index, headers, type_name, SAMSUNG_MAIN_FILL, SAMSUNG_MAIN_FILL)
        elif "双滚筒" in remark:
            type_name = "双滚筒"
            _apply_decomposition_style(sheet, row_index, headers, type_name, DOUBLE_DRUM_MAIN_FILL, DOUBLE_DRUM_MAIN_FILL)
        else:
            continue

        category_totals[type_name] += order_qty
        category_rows[type_name] += 1

    summary = _collect_decomposition_summary(sheet, headers)
    write_decomposition_detail_sheet(workbook, main_sheet_name, summary, logger)
    logger.info(
        "滚筒排单分解规则完成："
        f"CKD {category_rows['CKD']} 行/{category_totals['CKD']}，"
        f"CKD含筒部装 {category_rows['CKD含筒部装']} 行/{category_totals['CKD含筒部装']}，"
        f"三星 {category_rows['三星']} 行/{category_totals['三星']}，"
        f"双滚筒 {category_rows['双滚筒']} 行/{category_totals['双滚筒']}。"
    )
    return DecomposeRollingRemarkResult(
        rolling_total=summary["rolling_total"],
        wave_total=summary["wave_total"],
        category_totals=category_totals,
        category_rows=category_rows,
    )


def decompose_t7p7t5p5_dreame(
    workbook: Workbook,
    main_sheet_name: str,
    logger: ProcessingLogger,
) -> DecomposeT7P7T5P5DreameResult:
    sheet = workbook[main_sheet_name]
    headers = _header_map(sheet)
    _require_columns(
        headers,
        ["线体", "订单数", "物料编码", "物料描述", "类型", *DECOMPOSE_COLUMNS],
        "主数据表",
    )

    category_totals: dict[str, Decimal] = {T7P7_DRY_TYPE: Decimal("0"), T7P7_WASH_TYPE: Decimal("0")}
    category_rows: dict[str, int] = {T7P7_DRY_TYPE: 0, T7P7_WASH_TYPE: 0}
    skipped_dryer_rows = 0
    skipped_unknown_code_rows = 0

    for row_index in range(2, sheet.max_row + 1):
        line = str(sheet.cell(row_index, headers["线体"]).value or "").strip()
        if _is_wave_line(line) or not _is_uncolored_for_decomposition(sheet, row_index, headers):
            continue

        description = str(sheet.cell(row_index, headers["物料描述"]).value or "")
        if not _is_t7p7t5p5_dreame_description(description):
            continue

        product_kind = _product_kind_from_material_code(sheet.cell(row_index, headers["物料编码"]).value)
        if product_kind == "dryer":
            skipped_dryer_rows += 1
            continue
        if product_kind == "unknown":
            skipped_unknown_code_rows += 1
            continue

        type_name = T7P7_DRY_TYPE if product_kind == "dry" else T7P7_WASH_TYPE
        type_fill = T7P7_DRY_TYPE_FILL if product_kind == "dry" else T7P7_WASH_TYPE_FILL
        _apply_decomposition_style(sheet, row_index, headers, type_name, T7P7_MAIN_FILL, type_fill)

        order_qty = _to_decimal(sheet.cell(row_index, headers["订单数"]).value)
        category_totals[type_name] += order_qty
        category_rows[type_name] += 1

    summary = _collect_decomposition_summary(sheet, headers)
    write_decomposition_detail_sheet(workbook, main_sheet_name, summary, logger)
    logger.info(
        "T7/P7/T5/P5/追觅排单分解完成："
        f"{T7P7_DRY_TYPE} {category_rows[T7P7_DRY_TYPE]} 行/{category_totals[T7P7_DRY_TYPE]}，"
        f"{T7P7_WASH_TYPE} {category_rows[T7P7_WASH_TYPE]} 行/{category_totals[T7P7_WASH_TYPE]}，"
        f"跳过605干衣机 {skipped_dryer_rows} 行，"
        f"跳过无法识别编码 {skipped_unknown_code_rows} 行。"
    )
    return DecomposeT7P7T5P5DreameResult(
        rolling_total=summary["rolling_total"],
        wave_total=summary["wave_total"],
        category_totals=category_totals,
        category_rows=category_rows,
        skipped_dryer_rows=skipped_dryer_rows,
        skipped_unknown_code_rows=skipped_unknown_code_rows,
    )


def decompose_t9p9(
    workbook: Workbook,
    main_sheet_name: str,
    logger: ProcessingLogger,
) -> DecomposeT9P9Result:
    sheet = workbook[main_sheet_name]
    headers = _header_map(sheet)
    _require_columns(
        headers,
        ["线体", "订单数", "备注", "物料编码", "物料描述", "类型", *DECOMPOSE_COLUMNS],
        "主数据表",
    )

    category_totals: dict[str, Decimal] = {T9P9_DRY_TYPE: Decimal("0"), T9P9_WASH_TYPE: Decimal("0")}
    category_rows: dict[str, int] = {T9P9_DRY_TYPE: 0, T9P9_WASH_TYPE: 0}
    skipped_dryer_rows = 0
    skipped_unknown_code_rows = 0
    skipped_remark_only_rows = 0

    for row_index in range(2, sheet.max_row + 1):
        line = str(sheet.cell(row_index, headers["线体"]).value or "").strip()
        if _is_wave_line(line) or not _is_uncolored_for_decomposition(sheet, row_index, headers):
            continue

        remark = str(sheet.cell(row_index, headers["备注"]).value or "")
        description = str(sheet.cell(row_index, headers["物料描述"]).value or "")
        has_remark_candidate = "T9" in remark.upper() or "P9" in remark.upper()
        has_description_candidate = _is_t9p9_description(description)
        if not has_remark_candidate and not has_description_candidate:
            continue
        if not has_description_candidate:
            skipped_remark_only_rows += 1
            continue

        product_kind = _product_kind_from_material_code(sheet.cell(row_index, headers["物料编码"]).value)
        if product_kind == "dryer":
            skipped_dryer_rows += 1
            continue
        if product_kind == "unknown":
            skipped_unknown_code_rows += 1
            continue

        type_name = T9P9_DRY_TYPE if product_kind == "dry" else T9P9_WASH_TYPE
        type_fill = T9P9_DRY_TYPE_FILL if product_kind == "dry" else T9P9_WASH_TYPE_FILL
        _apply_decomposition_style(sheet, row_index, headers, type_name, T9P9_MAIN_FILL, type_fill)

        order_qty = _to_decimal(sheet.cell(row_index, headers["订单数"]).value)
        category_totals[type_name] += order_qty
        category_rows[type_name] += 1

    summary = _collect_decomposition_summary(sheet, headers)
    write_decomposition_detail_sheet(workbook, main_sheet_name, summary, logger)
    logger.info(
        "T9/P9排单分解完成："
        f"{T9P9_DRY_TYPE} {category_rows[T9P9_DRY_TYPE]} 行/{category_totals[T9P9_DRY_TYPE]}，"
        f"{T9P9_WASH_TYPE} {category_rows[T9P9_WASH_TYPE]} 行/{category_totals[T9P9_WASH_TYPE]}，"
        f"跳过605干衣机 {skipped_dryer_rows} 行，"
        f"跳过无法识别编码 {skipped_unknown_code_rows} 行，"
        f"跳过仅备注命中但物料描述前段未命中 {skipped_remark_only_rows} 行。"
    )
    return DecomposeT9P9Result(
        rolling_total=summary["rolling_total"],
        wave_total=summary["wave_total"],
        category_totals=category_totals,
        category_rows=category_rows,
        skipped_dryer_rows=skipped_dryer_rows,
        skipped_unknown_code_rows=skipped_unknown_code_rows,
        skipped_remark_only_rows=skipped_remark_only_rows,
    )


def decompose_t9p9_dryer(
    workbook: Workbook,
    main_sheet_name: str,
    logger: ProcessingLogger,
) -> DecomposeT9P9DryerResult:
    sheet = workbook[main_sheet_name]
    headers = _header_map(sheet)
    _require_columns(
        headers,
        ["线体", "订单数", "备注", "物料编码", "物料描述", "类型", *DECOMPOSE_COLUMNS],
        "主数据表",
    )

    category_totals: dict[str, Decimal] = {T9P9_DRYER_TYPE: Decimal("0")}
    category_rows: dict[str, int] = {T9P9_DRYER_TYPE: 0}
    skipped_non_dryer_rows = 0
    skipped_unknown_code_rows = 0

    for row_index in range(2, sheet.max_row + 1):
        line = str(sheet.cell(row_index, headers["线体"]).value or "").strip()
        if _is_wave_line(line) or not _is_uncolored_for_decomposition(sheet, row_index, headers):
            continue

        remark = str(sheet.cell(row_index, headers["备注"]).value or "")
        description = str(sheet.cell(row_index, headers["物料描述"]).value or "")
        remark_upper = remark.upper()
        has_remark_candidate = "T9干衣机" in remark_upper or "P9干衣机" in remark_upper
        has_description_candidate = _is_t9p9_description(description)
        if not has_remark_candidate and not has_description_candidate:
            continue

        product_kind = _product_kind_from_material_code(sheet.cell(row_index, headers["物料编码"]).value)
        if product_kind == "unknown":
            skipped_unknown_code_rows += 1
            continue
        if product_kind != "dryer":
            skipped_non_dryer_rows += 1
            continue

        _apply_decomposition_style(
            sheet,
            row_index,
            headers,
            T9P9_DRYER_TYPE,
            T9P9_MAIN_FILL,
            T9P9_DRYER_TYPE_FILL,
        )

        order_qty = _to_decimal(sheet.cell(row_index, headers["订单数"]).value)
        category_totals[T9P9_DRYER_TYPE] += order_qty
        category_rows[T9P9_DRYER_TYPE] += 1

    summary = _collect_decomposition_summary(sheet, headers)
    write_decomposition_detail_sheet(workbook, main_sheet_name, summary, logger)
    logger.info(
        "T9/P9干衣机排单分解完成："
        f"{T9P9_DRYER_TYPE} {category_rows[T9P9_DRYER_TYPE]} 行/{category_totals[T9P9_DRYER_TYPE]}，"
        f"跳过非605机型 {skipped_non_dryer_rows} 行，"
        f"跳过无法识别编码 {skipped_unknown_code_rows} 行。"
    )
    return DecomposeT9P9DryerResult(
        rolling_total=summary["rolling_total"],
        wave_total=summary["wave_total"],
        category_totals=category_totals,
        category_rows=category_rows,
        skipped_non_dryer_rows=skipped_non_dryer_rows,
        skipped_unknown_code_rows=skipped_unknown_code_rows,
    )


def decompose_t10p10(
    workbook: Workbook,
    main_sheet_name: str,
    logger: ProcessingLogger,
) -> DecomposeT10P10Result:
    sheet = workbook[main_sheet_name]
    headers = _header_map(sheet)
    _require_columns(
        headers,
        ["线体", "订单数", "备注", "物料编码", "类型", *DECOMPOSE_COLUMNS],
        "主数据表",
    )

    category_totals: dict[str, Decimal] = {T10P10_WASH_TYPE: Decimal("0"), T10P10_DRYER_TYPE: Decimal("0")}
    category_rows: dict[str, int] = {T10P10_WASH_TYPE: 0, T10P10_DRYER_TYPE: 0}
    skipped_unknown_code_rows = 0

    for row_index in range(2, sheet.max_row + 1):
        line = str(sheet.cell(row_index, headers["线体"]).value or "").strip()
        if _is_wave_line(line) or not _is_uncolored_for_decomposition(sheet, row_index, headers):
            continue

        remark = str(sheet.cell(row_index, headers["备注"]).value or "")
        remark_upper = remark.upper()
        if "T10" not in remark_upper and "P10" not in remark_upper:
            continue
        if "干衣机" not in remark and "洗衣机" not in remark:
            continue

        product_kind = _product_kind_from_material_code(sheet.cell(row_index, headers["物料编码"]).value)
        if product_kind == "unknown":
            skipped_unknown_code_rows += 1
            continue

        type_name = T10P10_DRYER_TYPE if product_kind == "dryer" else T10P10_WASH_TYPE
        type_fill = T10P10_DRYER_TYPE_FILL if product_kind == "dryer" else T10P10_WASH_TYPE_FILL
        _apply_decomposition_style(sheet, row_index, headers, type_name, T10P10_MAIN_FILL, type_fill)

        order_qty = _to_decimal(sheet.cell(row_index, headers["订单数"]).value)
        category_totals[type_name] += order_qty
        category_rows[type_name] += 1

    summary = _collect_decomposition_summary(sheet, headers)
    write_decomposition_detail_sheet(workbook, main_sheet_name, summary, logger)
    logger.info(
        "T10/P10排单分解完成："
        f"{T10P10_WASH_TYPE} {category_rows[T10P10_WASH_TYPE]} 行/{category_totals[T10P10_WASH_TYPE]}，"
        f"{T10P10_DRYER_TYPE} {category_rows[T10P10_DRYER_TYPE]} 行/{category_totals[T10P10_DRYER_TYPE]}，"
        f"跳过无法识别编码 {skipped_unknown_code_rows} 行。"
    )
    return DecomposeT10P10Result(
        rolling_total=summary["rolling_total"],
        wave_total=summary["wave_total"],
        category_totals=category_totals,
        category_rows=category_rows,
        skipped_unknown_code_rows=skipped_unknown_code_rows,
    )


def decompose_c6_heat_pump_dryer(
    workbook: Workbook,
    main_sheet_name: str,
    logger: ProcessingLogger,
) -> DecomposeC6HeatPumpDryerResult:
    sheet = workbook[main_sheet_name]
    headers = _header_map(sheet)
    _require_columns(
        headers,
        ["线体", "订单数", "备注", "物料编码", "类型", *DECOMPOSE_COLUMNS],
        "主数据表",
    )

    category_totals: dict[str, Decimal] = {C6_HEAT_PUMP_DRYER_TYPE: Decimal("0")}
    category_rows: dict[str, int] = {C6_HEAT_PUMP_DRYER_TYPE: 0}
    skipped_non_dryer_rows = 0
    skipped_unknown_code_rows = 0

    for row_index in range(2, sheet.max_row + 1):
        line = str(sheet.cell(row_index, headers["线体"]).value or "").strip()
        if _is_wave_line(line) or not _is_uncolored_for_decomposition(sheet, row_index, headers):
            continue

        remark = str(sheet.cell(row_index, headers["备注"]).value or "")
        if "干衣机" not in remark:
            continue

        product_kind = _product_kind_from_material_code(sheet.cell(row_index, headers["物料编码"]).value)
        if product_kind == "unknown":
            skipped_unknown_code_rows += 1
            continue
        if product_kind != "dryer":
            skipped_non_dryer_rows += 1
            continue

        _apply_decomposition_style(
            sheet,
            row_index,
            headers,
            C6_HEAT_PUMP_DRYER_TYPE,
            C6_HEAT_PUMP_DRYER_MAIN_FILL,
            C6_HEAT_PUMP_DRYER_TYPE_FILL,
        )

        order_qty = _to_decimal(sheet.cell(row_index, headers["订单数"]).value)
        category_totals[C6_HEAT_PUMP_DRYER_TYPE] += order_qty
        category_rows[C6_HEAT_PUMP_DRYER_TYPE] += 1

    summary = _collect_decomposition_summary(sheet, headers)
    write_decomposition_detail_sheet(workbook, main_sheet_name, summary, logger)
    logger.info(
        "C6热泵干衣机排单分解完成："
        f"{C6_HEAT_PUMP_DRYER_TYPE} {category_rows[C6_HEAT_PUMP_DRYER_TYPE]} 行/{category_totals[C6_HEAT_PUMP_DRYER_TYPE]}，"
        f"跳过非605机型 {skipped_non_dryer_rows} 行，"
        f"跳过无法识别编码 {skipped_unknown_code_rows} 行。"
    )
    return DecomposeC6HeatPumpDryerResult(
        rolling_total=summary["rolling_total"],
        wave_total=summary["wave_total"],
        category_totals=category_totals,
        category_rows=category_rows,
        skipped_non_dryer_rows=skipped_non_dryer_rows,
        skipped_unknown_code_rows=skipped_unknown_code_rows,
    )


def decompose_composite_penguin_c6(
    workbook: Workbook,
    main_sheet_name: str,
    logger: ProcessingLogger,
) -> DecomposeCompositePenguinC6Result:
    sheet = workbook[main_sheet_name]
    headers = _header_map(sheet)
    _require_columns(
        headers,
        ["线体", "订单数", "备注", "物料编码", "物料描述", "类型", *DECOMPOSE_COLUMNS],
        "主数据表",
    )

    category_totals: dict[str, Decimal] = {
        COMPOSITE_DRY_TYPE: Decimal("0"),
        COMPOSITE_WASH_TYPE: Decimal("0"),
        PENGUIN_DRYER_TYPE: Decimal("0"),
        PENGUIN_WASH_TYPE: Decimal("0"),
        C6_WASH_TYPE: Decimal("0"),
        C6_DRY_TYPE: Decimal("0"),
    }
    category_rows: dict[str, int] = {category: 0 for category in category_totals}
    skipped_unknown_code_rows = 0
    skipped_non_c6_dryer_rows = 0

    for row_index in range(2, sheet.max_row + 1):
        line = str(sheet.cell(row_index, headers["线体"]).value or "").strip()
        if _is_wave_line(line) or not _is_uncolored_for_decomposition(sheet, row_index, headers):
            continue

        remark = str(sheet.cell(row_index, headers["备注"]).value or "")
        description = str(sheet.cell(row_index, headers["物料描述"]).value or "")
        product_kind = _product_kind_from_material_code(sheet.cell(row_index, headers["物料编码"]).value)
        if product_kind == "unknown":
            if "复式" in remark or "企鹅" in remark or _is_c6_description(description):
                skipped_unknown_code_rows += 1
            continue

        type_name = None
        main_fill = None
        type_fill = None

        if "复式" in remark:
            if product_kind == "dry":
                type_name = COMPOSITE_DRY_TYPE
                type_fill = COMPOSITE_DRY_TYPE_FILL
            elif product_kind == "wash":
                type_name = COMPOSITE_WASH_TYPE
                type_fill = COMPOSITE_WASH_TYPE_FILL
            main_fill = COMPOSITE_MAIN_FILL
        elif "企鹅" in remark:
            if product_kind == "dryer":
                type_name = PENGUIN_DRYER_TYPE
                type_fill = PENGUIN_DRYER_TYPE_FILL
            else:
                type_name = PENGUIN_WASH_TYPE
                type_fill = PENGUIN_WASH_TYPE_FILL
            main_fill = PENGUIN_MAIN_FILL
        elif _is_c6_description(description):
            if product_kind == "dryer":
                skipped_non_c6_dryer_rows += 1
                continue
            if product_kind == "dry":
                type_name = C6_DRY_TYPE
                type_fill = C6_DRY_TYPE_FILL
            elif product_kind == "wash":
                type_name = C6_WASH_TYPE
                type_fill = C6_WASH_TYPE_FILL
            main_fill = C6_MAIN_FILL

        if not type_name or not main_fill or not type_fill:
            continue

        _apply_decomposition_style(sheet, row_index, headers, type_name, main_fill, type_fill)
        order_qty = _to_decimal(sheet.cell(row_index, headers["订单数"]).value)
        category_totals[type_name] += order_qty
        category_rows[type_name] += 1

    summary = _collect_decomposition_summary(sheet, headers)
    write_decomposition_detail_sheet(workbook, main_sheet_name, summary, logger)
    logger.info(
        "复式/企鹅/C6排单分解完成："
        f"{COMPOSITE_DRY_TYPE} {category_rows[COMPOSITE_DRY_TYPE]} 行/{category_totals[COMPOSITE_DRY_TYPE]}，"
        f"{COMPOSITE_WASH_TYPE} {category_rows[COMPOSITE_WASH_TYPE]} 行/{category_totals[COMPOSITE_WASH_TYPE]}，"
        f"{PENGUIN_DRYER_TYPE} {category_rows[PENGUIN_DRYER_TYPE]} 行/{category_totals[PENGUIN_DRYER_TYPE]}，"
        f"{PENGUIN_WASH_TYPE} {category_rows[PENGUIN_WASH_TYPE]} 行/{category_totals[PENGUIN_WASH_TYPE]}，"
        f"{C6_WASH_TYPE} {category_rows[C6_WASH_TYPE]} 行/{category_totals[C6_WASH_TYPE]}，"
        f"{C6_DRY_TYPE} {category_rows[C6_DRY_TYPE]} 行/{category_totals[C6_DRY_TYPE]}，"
        f"跳过C6干衣机 {skipped_non_c6_dryer_rows} 行，"
        f"跳过无法识别编码 {skipped_unknown_code_rows} 行。"
    )
    return DecomposeCompositePenguinC6Result(
        rolling_total=summary["rolling_total"],
        wave_total=summary["wave_total"],
        category_totals=category_totals,
        category_rows=category_rows,
        skipped_unknown_code_rows=skipped_unknown_code_rows,
        skipped_non_c6_dryer_rows=skipped_non_c6_dryer_rows,
    )


def decompose_rolling_final(
    workbook: Workbook,
    main_sheet_name: str,
    logger: ProcessingLogger,
) -> DecomposeRollingFinalResult:
    sheet = workbook[main_sheet_name]
    headers = _header_map(sheet)
    _require_columns(
        headers,
        ["线体", "渠道", "订单数", "备注", "物料编码", "类型", *DECOMPOSE_COLUMNS],
        "主数据表",
    )

    category_totals: dict[str, Decimal] = {
        ORDINARY_DRY_TYPE: Decimal("0"),
        DOMESTIC_TYPE: Decimal("0"),
        EXPORT_TYPE: Decimal("0"),
    }
    category_rows: dict[str, int] = {category: 0 for category in category_totals}

    _clear_wave_basic_classifications(sheet, headers)

    for row_index in range(2, sheet.max_row + 1):
        line = str(sheet.cell(row_index, headers["线体"]).value or "").strip()
        if _is_wave_line(line) or not _is_uncolored_for_decomposition(sheet, row_index, headers):
            continue

        remark = str(sheet.cell(row_index, headers["备注"]).value or "")
        channel = str(sheet.cell(row_index, headers["渠道"]).value or "").strip()
        product_kind = _product_kind_from_material_code(sheet.cell(row_index, headers["物料编码"]).value)

        type_name = None
        main_fill = None
        type_fill = None
        if "烘干" in remark and product_kind == "dry":
            type_name = ORDINARY_DRY_TYPE
            main_fill = ORDINARY_DRY_MAIN_FILL
            type_fill = ORDINARY_DRY_TYPE_FILL
        elif channel in {"CBG", "国内ODM"}:
            type_name = DOMESTIC_TYPE
            main_fill = DOMESTIC_MAIN_FILL
            type_fill = DOMESTIC_TYPE_FILL
        elif channel in {"海外BG", "海外ODM"}:
            type_name = EXPORT_TYPE
            main_fill = EXPORT_MAIN_FILL
            type_fill = EXPORT_TYPE_FILL

        if not type_name or not main_fill or not type_fill:
            continue

        _apply_decomposition_style(sheet, row_index, headers, type_name, main_fill, type_fill)
        order_qty = _to_decimal(sheet.cell(row_index, headers["订单数"]).value)
        category_totals[type_name] += order_qty
        category_rows[type_name] += 1

    summary = _collect_decomposition_summary(sheet, headers)
    remaining_rows, remaining_total = _remaining_unclassified_rolling(sheet, headers)
    rolling_decomposition_total = _rolling_decomposition_total(summary["rolling_types"])
    rolling_gap = summary["rolling_total"] - rolling_decomposition_total
    write_decomposition_detail_sheet(workbook, main_sheet_name, summary, logger)
    logger.info(
        "滚筒排单分解收尾完成："
        f"{ORDINARY_DRY_TYPE} {category_rows[ORDINARY_DRY_TYPE]} 行/{category_totals[ORDINARY_DRY_TYPE]}，"
        f"{DOMESTIC_TYPE} {category_rows[DOMESTIC_TYPE]} 行/{category_totals[DOMESTIC_TYPE]}，"
        f"{EXPORT_TYPE} {category_rows[EXPORT_TYPE]} 行/{category_totals[EXPORT_TYPE]}，"
        f"滚筒剩余未分类 {remaining_rows} 行/{remaining_total}，"
        f"滚筒分解合计 {rolling_decomposition_total}，滚筒总数 {summary['rolling_total']}，差异 {rolling_gap}。"
    )
    return DecomposeRollingFinalResult(
        rolling_total=summary["rolling_total"],
        wave_total=summary["wave_total"],
        category_totals=category_totals,
        category_rows=category_rows,
        remaining_unclassified_rows=remaining_rows,
        remaining_unclassified_total=remaining_total,
        rolling_decomposition_total=rolling_decomposition_total,
        rolling_gap=rolling_gap,
    )


def decompose_wave_basic(
    workbook: Workbook,
    values_workbook: Workbook,
    main_sheet_name: str,
    logger: ProcessingLogger,
) -> DecomposeWaveBasicResult:
    sheet = workbook[main_sheet_name]
    values_sheet = values_workbook[main_sheet_name]
    headers = _header_map(sheet)
    _require_columns(
        headers,
        ["线体", "订单数", "备注", "系数", "钣金型号", "物料描述", "类型", *DECOMPOSE_COLUMNS],
        "主数据表",
    )

    category_totals: dict[str, Decimal] = {
        "CKD": Decimal("0"),
        WAVE_LG_TYPE: Decimal("0"),
        WAVE_PLASTIC_DOMESTIC_TYPE: Decimal("0"),
        WAVE_P7P9_TYPE: Decimal("0"),
        "SKD": Decimal("0"),
    }
    category_rows: dict[str, int] = {category: 0 for category in category_totals}

    _clear_wave_basic_classifications(sheet, headers)

    for row_index in range(2, sheet.max_row + 1):
        line = str(sheet.cell(row_index, headers["线体"]).value or "").strip()
        if not _is_wave_line(line) or not _is_uncolored_for_decomposition(sheet, row_index, headers):
            continue

        remark = str(sheet.cell(row_index, headers["备注"]).value or "")
        remark_upper = remark.upper()
        coefficient = _to_decimal_or_none(values_sheet.cell(row_index, headers["系数"]).value)
        if coefficient is None:
            coefficient = _to_decimal_or_none(sheet.cell(row_index, headers["系数"]).value)
        sheet_metal = str(
            values_sheet.cell(row_index, headers["钣金型号"]).value
            or sheet.cell(row_index, headers["钣金型号"]).value
            or ""
        )
        description = str(
            values_sheet.cell(row_index, headers["物料描述"]).value
            or sheet.cell(row_index, headers["物料描述"]).value
            or ""
        )

        type_name = None
        main_fill = None
        type_fill = None
        if "CKD" in remark_upper:
            type_name = "CKD"
            main_fill = WAVE_CKD_MAIN_FILL
            type_fill = WAVE_CKD_TYPE_FILL
        elif "LG" in remark_upper:
            type_name = WAVE_LG_TYPE
            main_fill = WAVE_LG_MAIN_FILL
            type_fill = WAVE_LG_TYPE_FILL
        elif coefficient is not None and coefficient < Decimal("1.217"):
            type_name = WAVE_PLASTIC_DOMESTIC_TYPE
            main_fill = WAVE_PLASTIC_MAIN_FILL
            type_fill = WAVE_PLASTIC_TYPE_FILL
        elif _is_wave_p7p9_candidate(sheet_metal, description):
            type_name = WAVE_P7P9_TYPE
            main_fill = WAVE_P7P9_MAIN_FILL
            type_fill = WAVE_P7P9_TYPE_FILL
        elif "SKD" in remark_upper:
            type_name = "SKD"
            main_fill = WAVE_SKD_MAIN_FILL
            type_fill = WAVE_SKD_TYPE_FILL

        if not type_name or not main_fill or not type_fill:
            continue

        _apply_decomposition_style(sheet, row_index, headers, type_name, main_fill, type_fill)
        order_qty = _to_decimal(sheet.cell(row_index, headers["订单数"]).value)
        category_totals[type_name] += order_qty
        category_rows[type_name] += 1

    summary = _collect_decomposition_summary(sheet, headers)
    remaining_rows, remaining_total = _remaining_unclassified_wave(sheet, headers)
    write_decomposition_detail_sheet(workbook, main_sheet_name, summary, logger)
    logger.info(
        "波轮排单分解第一组规则完成："
        f"CKD {category_rows['CKD']} 行/{category_totals['CKD']}，"
        f"{WAVE_LG_TYPE} {category_rows[WAVE_LG_TYPE]} 行/{category_totals[WAVE_LG_TYPE]}，"
        f"{WAVE_PLASTIC_DOMESTIC_TYPE} {category_rows[WAVE_PLASTIC_DOMESTIC_TYPE]} 行/{category_totals[WAVE_PLASTIC_DOMESTIC_TYPE]}，"
        f"{WAVE_P7P9_TYPE} {category_rows[WAVE_P7P9_TYPE]} 行/{category_totals[WAVE_P7P9_TYPE]}，"
        f"SKD {category_rows['SKD']} 行/{category_totals['SKD']}，"
        f"波轮剩余未分类 {remaining_rows} 行/{remaining_total}。"
    )
    return DecomposeWaveBasicResult(
        rolling_total=summary["rolling_total"],
        wave_total=summary["wave_total"],
        category_totals=category_totals,
        category_rows=category_rows,
        remaining_unclassified_rows=remaining_rows,
        remaining_unclassified_total=remaining_total,
    )


def decompose_wave_final(
    workbook: Workbook,
    values_workbook: Workbook,
    main_sheet_name: str,
    logger: ProcessingLogger,
) -> DecomposeWaveFinalResult:
    sheet = workbook[main_sheet_name]
    values_sheet = values_workbook[main_sheet_name]
    headers = _header_map(sheet)
    _require_columns(
        headers,
        ["线体", "渠道", "订单数", "系数", "物料描述", "类型", *DECOMPOSE_COLUMNS],
        "主数据表",
    )

    category_totals: dict[str, Decimal] = {
        WAVE_DOMESTIC_INVERTER_TYPE: Decimal("0"),
        WAVE_EXPORT_INVERTER_TYPE: Decimal("0"),
        WAVE_DOMESTIC_IRON_TYPE: Decimal("0"),
        WAVE_EXPORT_IRON_TYPE: Decimal("0"),
    }
    category_rows: dict[str, int] = {category: 0 for category in category_totals}

    _clear_wave_final_classifications(sheet, headers)

    for row_index in range(2, sheet.max_row + 1):
        line = str(sheet.cell(row_index, headers["线体"]).value or "").strip()
        if not _is_wave_line(line) or not _is_uncolored_for_decomposition(sheet, row_index, headers):
            continue
        if str(sheet.cell(row_index, headers["类型"]).value or "").strip():
            continue

        channel = str(sheet.cell(row_index, headers["渠道"]).value or "").strip()
        description = str(
            values_sheet.cell(row_index, headers["物料描述"]).value
            or sheet.cell(row_index, headers["物料描述"]).value
            or ""
        )
        coefficient = _to_decimal_or_none(values_sheet.cell(row_index, headers["系数"]).value)
        if coefficient is None:
            coefficient = _to_decimal_or_none(sheet.cell(row_index, headers["系数"]).value)

        type_name = None
        main_fill = None
        type_fill = None
        if channel in {"CBG", "国内ODM"} and _is_wave_domestic_inverter_description(description):
            type_name = WAVE_DOMESTIC_INVERTER_TYPE
            main_fill = WAVE_DOMESTIC_INVERTER_MAIN_FILL
            type_fill = WAVE_DOMESTIC_INVERTER_TYPE_FILL
        elif channel in {"海外BG", "海外ODM"} and coefficient is not None and coefficient > Decimal("2"):
            type_name = WAVE_EXPORT_INVERTER_TYPE
            main_fill = WAVE_EXPORT_INVERTER_MAIN_FILL
            type_fill = WAVE_EXPORT_INVERTER_TYPE_FILL
        elif channel in {"CBG", "国内ODM"}:
            type_name = WAVE_DOMESTIC_IRON_TYPE
            main_fill = WAVE_DOMESTIC_IRON_MAIN_FILL
            type_fill = WAVE_DOMESTIC_IRON_TYPE_FILL
        elif channel in {"海外BG", "海外ODM"}:
            type_name = WAVE_EXPORT_IRON_TYPE
            main_fill = WAVE_EXPORT_IRON_MAIN_FILL
            type_fill = WAVE_EXPORT_IRON_TYPE_FILL

        if not type_name or not main_fill or not type_fill:
            continue

        _apply_decomposition_style(sheet, row_index, headers, type_name, main_fill, type_fill)
        order_qty = _to_decimal(sheet.cell(row_index, headers["订单数"]).value)
        category_totals[type_name] += order_qty
        category_rows[type_name] += 1

    summary = _collect_decomposition_summary(sheet, headers)
    remaining_rows, remaining_total = _remaining_unclassified_wave(sheet, headers)
    wave_decomposition_total = _wave_decomposition_total(summary["wave_types"])
    wave_gap = summary["wave_total"] - wave_decomposition_total
    write_decomposition_detail_sheet(workbook, main_sheet_name, summary, logger)
    logger.info(
        "波轮排单分解收尾规则完成："
        f"{WAVE_DOMESTIC_INVERTER_TYPE} {category_rows[WAVE_DOMESTIC_INVERTER_TYPE]} 行/{category_totals[WAVE_DOMESTIC_INVERTER_TYPE]}，"
        f"{WAVE_EXPORT_INVERTER_TYPE} {category_rows[WAVE_EXPORT_INVERTER_TYPE]} 行/{category_totals[WAVE_EXPORT_INVERTER_TYPE]}，"
        f"{WAVE_DOMESTIC_IRON_TYPE} {category_rows[WAVE_DOMESTIC_IRON_TYPE]} 行/{category_totals[WAVE_DOMESTIC_IRON_TYPE]}，"
        f"{WAVE_EXPORT_IRON_TYPE} {category_rows[WAVE_EXPORT_IRON_TYPE]} 行/{category_totals[WAVE_EXPORT_IRON_TYPE]}，"
        f"波轮剩余未分类 {remaining_rows} 行/{remaining_total}，"
        f"波轮分解合计 {wave_decomposition_total}，波轮总数 {summary['wave_total']}，差异 {wave_gap}。"
    )
    return DecomposeWaveFinalResult(
        rolling_total=summary["rolling_total"],
        wave_total=summary["wave_total"],
        category_totals=category_totals,
        category_rows=category_rows,
        remaining_unclassified_rows=remaining_rows,
        remaining_unclassified_total=remaining_total,
        wave_decomposition_total=wave_decomposition_total,
        wave_gap=wave_gap,
    )


def write_extra_order_summary(
    workbook: Workbook,
    values_workbook: Workbook,
    main_sheet_name: str,
    logger: ProcessingLogger,
) -> ExtraOrderSummaryResult:
    sheet = workbook[main_sheet_name]
    values_sheet = values_workbook[main_sheet_name]
    headers = _header_map(sheet)
    _require_columns(headers, ["线体", "钣金型号", "订单数", "类型"], "主数据表")

    if DETAIL_SHEET_NAME not in workbook.sheetnames and "类型" in headers:
        summary = _collect_decomposition_summary(sheet, headers)
        write_decomposition_detail_sheet(workbook, main_sheet_name, summary, logger)
    elif DETAIL_SHEET_NAME not in workbook.sheetnames:
        workbook.create_sheet(DETAIL_SHEET_NAME)
        _move_sheet_after(workbook, DETAIL_SHEET_NAME, main_sheet_name)
        logger.warning(f"未找到“{DETAIL_SHEET_NAME}”和“类型”列，已仅创建右侧额外订单信息汇总表。")

    detail_sheet = workbook[DETAIL_SHEET_NAME]
    category_totals = {rule.name: Decimal("0") for rule in EXTRA_ORDER_SUMMARY_RULES}
    category_rows = {rule.name: 0 for rule in EXTRA_ORDER_SUMMARY_RULES}
    _clear_extra_summary_metal_fills(sheet, headers)

    for row_index in range(2, sheet.max_row + 1):
        line = str(sheet.cell(row_index, headers["线体"]).value or "").strip()
        metal_model = str(
            values_sheet.cell(row_index, headers["钣金型号"]).value
            or sheet.cell(row_index, headers["钣金型号"]).value
            or ""
        )
        type_name = str(sheet.cell(row_index, headers["类型"]).value or "").strip()
        order_qty = _to_decimal(sheet.cell(row_index, headers["订单数"]).value)
        matched_rules = [
            rule
            for rule in EXTRA_ORDER_SUMMARY_RULES
            if _matches_extra_order_rule(rule, line, metal_model, type_name)
        ]
        if not matched_rules:
            continue

        for rule in matched_rules:
            category_totals[rule.name] += order_qty
            category_rows[rule.name] += 1

        first_rule = matched_rules[0]
        metal_cell = sheet.cell(row_index, headers["钣金型号"])
        metal_cell.fill = PatternFill(fill_type="solid", fgColor=first_rule.fill_rgb)
        _set_font_for_fill(metal_cell, first_rule.fill_rgb)

    _write_extra_order_summary_table(detail_sheet, category_rows, category_totals)
    write_line_classification_detail_sheet(workbook, main_sheet_name, headers, logger)
    logger.info(
        "额外订单信息汇总完成："
        + "；".join(
            f"{rule.name} {category_rows[rule.name]} 行/{category_totals[rule.name]}"
            for rule in EXTRA_ORDER_SUMMARY_RULES
        )
    )
    return ExtraOrderSummaryResult(category_totals=category_totals, category_rows=category_rows)


def write_decomposition_detail_sheet(
    workbook: Workbook,
    main_sheet_name: str,
    summary: dict[str, Any],
    logger: ProcessingLogger,
) -> None:
    preserved_styles = _capture_detail_sheet_styles(workbook)
    remove_sheet_if_exists(workbook, DETAIL_SHEET_NAME)
    sheet = workbook.create_sheet(DETAIL_SHEET_NAME)

    rolling_total = summary["rolling_total"]
    wave_total = summary["wave_total"]
    rolling_types: dict[str, Decimal] = summary["rolling_types"]
    wave_types: dict[str, Decimal] = summary["wave_types"]

    rows = [
        ["排单分解", None, None, None],
        ["滚筒排单分解", None, "波轮排单分解", None],
        ["滚筒总数", _number_or_int(rolling_total), "波轮总数", _number_or_int(wave_total)],
    ]
    max_rows = max(len(ROLLING_LABELS), len(WAVE_LABELS))
    for index in range(max_rows):
        rolling_label = ROLLING_LABELS[index] if index < len(ROLLING_LABELS) else None
        wave_label = WAVE_LABELS[index] if index < len(WAVE_LABELS) else None
        rows.append(
            [
                rolling_label,
                _rolling_value_for_label(rolling_label, rolling_types),
                wave_label,
                _wave_value_for_label(wave_label, wave_types),
            ]
        )
    rows.append(["合计", f"=SUM(B5:B{3 + len(ROLLING_LABELS)})", "合计", f"=SUM(D4:D{3 + len(WAVE_LABELS)})"])

    for row in rows:
        sheet.append(row)

    sheet.merge_cells("A1:D1")
    sheet.merge_cells("A2:B2")
    sheet.merge_cells("C2:D2")
    _style_decomposition_sheet(sheet)
    _apply_preserved_detail_sheet_styles(sheet, preserved_styles)
    _move_sheet_after(workbook, DETAIL_SHEET_NAME, main_sheet_name)
    logger.info(f"已创建“{DETAIL_SHEET_NAME}”，并写入滚筒/波轮总数与当前排单分解统计。")


def write_line_classification_detail_sheet(
    workbook: Workbook,
    main_sheet_name: str,
    headers: dict[str, int],
    logger: ProcessingLogger,
) -> None:
    sheet = workbook[main_sheet_name]
    remove_sheet_if_exists(workbook, LINE_CLASSIFICATION_SHEET_NAME)
    detail_sheet = workbook.create_sheet(LINE_CLASSIFICATION_SHEET_NAME)

    rolling_line_types = _collect_line_type_totals(sheet, headers, ROLLING_LINE_ORDER)
    wave_line_types = _collect_line_type_totals(sheet, headers, WAVE_LINE_ORDER)
    rolling_line_totals = _collect_line_totals(sheet, headers, ROLLING_LINE_ORDER)
    wave_line_totals = _collect_line_totals(sheet, headers, WAVE_LINE_ORDER)

    rolling_rows = _build_line_detail_rows(ROLLING_LABELS, ROLLING_LINE_ORDER, rolling_line_types, "rolling")
    wave_rows = _build_line_detail_rows(WAVE_LABELS, WAVE_LINE_ORDER, wave_line_types, "wave")

    row_index = 1
    row_index = _write_line_detail_section(
        detail_sheet,
        row_index,
        "滚筒各线体分类明细",
        ROLLING_LINE_ORDER,
        rolling_rows,
        rolling_line_totals,
    )
    _write_line_detail_section(
        detail_sheet,
        row_index + 2,
        "波轮各线体分类明细",
        WAVE_LINE_ORDER,
        wave_rows,
        wave_line_totals,
    )
    _style_line_classification_detail_sheet(detail_sheet)
    _move_sheet_after(workbook, LINE_CLASSIFICATION_SHEET_NAME, DETAIL_SHEET_NAME)
    logger.info(f"已创建“{LINE_CLASSIFICATION_SHEET_NAME}”，并按线体汇总各分类订单数。")


def _capture_detail_sheet_styles(workbook: Workbook) -> dict[tuple[str, str], list[dict[str, Any]]]:
    if DETAIL_SHEET_NAME not in workbook.sheetnames:
        return {}

    sheet = workbook[DETAIL_SHEET_NAME]
    styles: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row_index in range(1, sheet.max_row + 1):
        for side, label_col in (("rolling", 1), ("wave", 3)):
            label = sheet.cell(row_index, label_col).value
            if label is None:
                continue
            styles[(side, str(label))] = [
                _capture_cell_style(sheet.cell(row_index, label_col)),
                _capture_cell_style(sheet.cell(row_index, label_col + 1)),
            ]
    return styles


def _capture_cell_style(cell) -> dict[str, Any]:
    return {
        "fill": copy(cell.fill),
        "font": copy(cell.font),
    }


def _apply_preserved_detail_sheet_styles(sheet: Worksheet, styles: dict[tuple[str, str], list[dict[str, Any]]]) -> None:
    if not styles:
        return

    for row_index in range(1, sheet.max_row + 1):
        for side, label_col in (("rolling", 1), ("wave", 3)):
            label = sheet.cell(row_index, label_col).value
            if label is None:
                continue
            saved_styles = styles.get((side, str(label)))
            if not saved_styles:
                continue
            for offset, saved_style in enumerate(saved_styles):
                cell = sheet.cell(row_index, label_col + offset)
                cell.fill = copy(saved_style["fill"])
                cell.font = copy(saved_style["font"])


def _apply_skd_style(sheet: Worksheet, row_index: int, headers: dict[str, int], type_name: str, is_dry: bool) -> None:
    _apply_decomposition_style(
        sheet,
        row_index,
        headers,
        type_name,
        SKD_MAIN_FILL,
        SKD_DRY_TYPE_FILL if is_dry else SKD_NORMAL_TYPE_FILL,
    )


def _apply_decomposition_style(
    sheet: Worksheet,
    row_index: int,
    headers: dict[str, int],
    type_name: str,
    main_fill_rgb: str,
    type_fill_rgb: str,
) -> None:
    main_fill = PatternFill(fill_type="solid", fgColor=main_fill_rgb)
    type_fill = PatternFill(fill_type="solid", fgColor=type_fill_rgb)
    for field in ["基本开始日期", "备注"]:
        cell = sheet.cell(row_index, headers[field])
        cell.fill = main_fill
        _set_font_for_fill(cell, main_fill_rgb)

    type_cell = sheet.cell(row_index, headers["类型"])
    type_cell.value = type_name
    type_cell.fill = type_fill
    _set_font_for_fill(type_cell, type_fill_rgb)


def _clear_wave_basic_classifications(sheet: Worksheet, headers: dict[str, int]) -> None:
    for row_index in range(2, sheet.max_row + 1):
        line = str(sheet.cell(row_index, headers["线体"]).value or "").strip()
        type_cell = sheet.cell(row_index, headers["类型"])
        if not _is_wave_line(line):
            continue
        for field in DECOMPOSE_COLUMNS:
            cell = sheet.cell(row_index, headers[field])
            cell.fill = PatternFill(fill_type=None)
        type_cell.value = None


def _clear_wave_final_classifications(sheet: Worksheet, headers: dict[str, int]) -> None:
    for row_index in range(2, sheet.max_row + 1):
        line = str(sheet.cell(row_index, headers["线体"]).value or "").strip()
        type_cell = sheet.cell(row_index, headers["类型"])
        type_name = str(type_cell.value or "").strip()
        if not _is_wave_line(line) or type_name not in WAVE_FINAL_TYPES:
            continue
        for field in DECOMPOSE_COLUMNS:
            cell = sheet.cell(row_index, headers[field])
            cell.fill = PatternFill(fill_type=None)
        type_cell.value = None


def _is_uncolored_for_decomposition(sheet: Worksheet, row_index: int, headers: dict[str, int]) -> bool:
    for field in DECOMPOSE_COLUMNS:
        cell = sheet.cell(row_index, headers[field])
        if cell.fill.fill_type is not None:
            return False
    return True


def _style_decomposition_sheet(sheet: Worksheet) -> None:
    blue_fill = PatternFill(fill_type="solid", fgColor="0070C0")
    section_fill = PatternFill(fill_type="solid", fgColor="D9E2F3")
    light_fill = PatternFill(fill_type="solid", fgColor="D9E2F3")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if cell.row <= 2 else "left", vertical="center")
            cell.font = Font(color="000000", bold=False)

    sheet["A1"].fill = blue_fill
    sheet["A1"].font = Font(color="FFFFFF", bold=True)
    for cell in sheet[2]:
        cell.fill = section_fill
        cell.font = Font(bold=True, color="000000")

    shaded_labels = {
        "CKD含筒部装",
        "内销铁皮变频",
        "内销铁皮",
        "C6热泵干衣机",
        "复式烘干",
        "C6单洗",
        "C6烘干",
        "普通烘干",
        "内销",
        "外销",
    }
    for row_index in range(3, sheet.max_row + 1):
        for label_col in (1, 3):
            value = sheet.cell(row_index, label_col).value
            if value in shaded_labels:
                sheet.cell(row_index, label_col).fill = light_fill
                sheet.cell(row_index, label_col + 1).fill = light_fill

    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True, color="000000")
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 22
    sheet.column_dimensions["D"].width = 14


def _write_extra_order_summary_table(
    sheet: Worksheet,
    category_rows: dict[str, int],
    category_totals: dict[str, Decimal],
) -> None:
    start_col = 6
    end_col = 9
    max_row = max(sheet.max_row, len(EXTRA_ORDER_SUMMARY_RULES) + 2)
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_col >= start_col and merged_range.max_col <= end_col:
            sheet.unmerge_cells(str(merged_range))

    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.font = Font(color="000000", bold=False)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    rows = [
        ["额外订单信息汇总", None, None, None],
        ["补充项目", "数据行数", "订单数合计", "筛选口径"],
    ]
    for rule in EXTRA_ORDER_SUMMARY_RULES:
        rows.append(
            [
                rule.name,
                category_rows[rule.name],
                _number_or_int(category_totals[rule.name]),
                rule.criteria,
            ]
        )

    for row_offset, row_values in enumerate(rows, start=1):
        for col_offset, value in enumerate(row_values, start=start_col):
            sheet.cell(row_offset, col_offset).value = value

    sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    blue_fill = PatternFill(fill_type="solid", fgColor="0070C0")
    header_fill = PatternFill(fill_type="solid", fgColor="D9E2F3")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in sheet.iter_rows(min_row=1, max_row=len(rows), min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if cell.row <= 2 else "left", vertical="center")
            cell.font = Font(color="000000", bold=False)

    for cell in sheet[1][start_col - 1 : end_col]:
        cell.fill = blue_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in sheet[2][start_col - 1 : end_col]:
        cell.fill = header_fill
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.column_dimensions["F"].width = 34
    sheet.column_dimensions["G"].width = 12
    sheet.column_dimensions["H"].width = 14
    sheet.column_dimensions["I"].width = 42


def _collect_line_type_totals(
    sheet: Worksheet,
    headers: dict[str, int],
    line_order: list[str],
) -> dict[str, dict[str, Decimal]]:
    totals = {line: {} for line in line_order}
    for row_index in range(2, sheet.max_row + 1):
        line = str(sheet.cell(row_index, headers["线体"]).value or "").strip()
        if line not in totals:
            continue
        type_name = str(sheet.cell(row_index, headers["类型"]).value or "").strip()
        if not type_name:
            continue
        order_qty = _to_decimal(sheet.cell(row_index, headers["订单数"]).value)
        totals[line][type_name] = totals[line].get(type_name, Decimal("0")) + order_qty
    return totals


def _collect_line_totals(
    sheet: Worksheet,
    headers: dict[str, int],
    line_order: list[str],
) -> dict[str, Decimal]:
    totals = {line: Decimal("0") for line in line_order}
    for row_index in range(2, sheet.max_row + 1):
        line = str(sheet.cell(row_index, headers["线体"]).value or "").strip()
        if line not in totals:
            continue
        totals[line] += _to_decimal(sheet.cell(row_index, headers["订单数"]).value)
    return totals


def _build_line_detail_rows(
    labels: list[str],
    line_order: list[str],
    line_type_totals: dict[str, dict[str, Decimal]],
    section: str,
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for label in labels:
        values = [
            _line_category_value(label, line_type_totals[line], section)
            for line in line_order
        ]
        rows.append([label, *values, sum(values, Decimal("0"))])
    return rows


def _line_category_value(label: str | None, type_totals: dict[str, Decimal], section: str) -> Decimal:
    if not label:
        return Decimal("0")
    if section == "rolling":
        if label == "SKD总数(包含烘干)":
            return type_totals.get("SKD", Decimal("0")) + type_totals.get("SKD烘干", Decimal("0"))
        if label == "内销":
            return type_totals.get("内销", Decimal("0")) + type_totals.get(DOMESTIC_TYPE, Decimal("0"))
        return type_totals.get(label, Decimal("0"))
    if label == WAVE_P7P9_LABEL:
        return type_totals.get(WAVE_P7P9_LABEL, Decimal("0")) + type_totals.get(WAVE_P7P9_TYPE, Decimal("0"))
    return type_totals.get(label, Decimal("0"))


def _write_line_detail_section(
    sheet: Worksheet,
    start_row: int,
    title: str,
    line_order: list[str],
    rows: list[list[Any]],
    line_totals: dict[str, Decimal],
) -> int:
    end_col = len(line_order) + 2
    sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=end_col)
    sheet.cell(start_row, 1).value = title
    header = ["分类名称", *line_order, "合计"]
    for col_index, value in enumerate(header, start=1):
        sheet.cell(start_row + 1, col_index).value = value

    for row_offset, row_values in enumerate(rows, start=start_row + 2):
        sheet.cell(row_offset, 1).value = row_values[0]
        for col_index, value in enumerate(row_values[1:], start=2):
            sheet.cell(row_offset, col_index).value = _number_or_int(value)

    total_row = start_row + len(rows) + 2
    sheet.cell(total_row, 1).value = "合计"
    total = Decimal("0")
    for col_index, line in enumerate(line_order, start=2):
        value = line_totals[line]
        sheet.cell(total_row, col_index).value = _number_or_int(value)
        total += value
    sheet.cell(total_row, end_col).value = _number_or_int(total)
    return total_row


def _style_line_classification_detail_sheet(sheet: Worksheet) -> None:
    blue_fill = PatternFill(fill_type="solid", fgColor="0070C0")
    header_fill = PatternFill(fill_type="solid", fgColor="D9E2F3")
    total_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(color="000000", bold=False)

    for row_index in range(1, sheet.max_row + 1):
        first_value = sheet.cell(row_index, 1).value
        if first_value in {"滚筒各线体分类明细", "波轮各线体分类明细"}:
            for cell in sheet[row_index]:
                cell.fill = blue_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        elif first_value == "分类名称":
            for cell in sheet[row_index]:
                cell.fill = header_fill
                cell.font = Font(color="000000", bold=True)
        elif first_value == "合计":
            for cell in sheet[row_index]:
                cell.fill = total_fill
                cell.font = Font(color="000000", bold=True)

    sheet.freeze_panes = "B3"
    sheet.column_dimensions["A"].width = 34
    for col_index in range(2, 9):
        sheet.column_dimensions[get_column_letter(col_index)].width = 12


def _clear_extra_summary_metal_fills(sheet: Worksheet, headers: dict[str, int]) -> None:
    extra_fill_rgbs = {rule.fill_rgb.upper() for rule in EXTRA_ORDER_SUMMARY_RULES}
    metal_col = headers["钣金型号"]
    for row_index in range(2, sheet.max_row + 1):
        cell = sheet.cell(row_index, metal_col)
        if _fill_rgb(cell) not in extra_fill_rgbs:
            continue
        cell.fill = PatternFill(fill_type=None)
        cell.font = copy(cell.font)
        cell.font = Font(
            name=cell.font.name,
            sz=cell.font.sz,
            bold=cell.font.bold,
            italic=cell.font.italic,
            vertAlign=cell.font.vertAlign,
            underline=cell.font.underline,
            strike=cell.font.strike,
            color="000000",
        )


def _fill_rgb(cell) -> str | None:
    color = cell.fill.fgColor
    if cell.fill.fill_type is None or color is None or color.type != "rgb" or color.rgb is None:
        return None
    return str(color.rgb)[-6:].upper()


def _matches_extra_order_rule(rule: ExtraOrderSummaryRule, line: str, metal_model: str, type_name: str) -> bool:
    is_wave = _is_wave_line(line)
    if rule.scope == "rolling" and is_wave:
        return False
    if rule.scope == "wave" and not is_wave:
        return False

    metal_text = str(metal_model or "")
    metal_upper = metal_text.upper()
    if rule.name == "锥形筒":
        return "锥形筒" in metal_text
    if rule.name == "波轮特殊内筒-10kg和9升10内筒":
        return "10KG波轮" in metal_upper or "9升10" in metal_text
    if rule.name == "波轮特殊内筒-8升9内筒":
        return "8升9" in metal_text
    if rule.name == "波轮箱体-10kg":
        return "10KG" in metal_upper
    if rule.name == "波轮箱体-彩板":
        return "PCM" in metal_upper and type_name != "CKD"
    return False


def _collect_decomposition_summary(sheet: Worksheet, headers: dict[str, int]) -> dict[str, Any]:
    rolling_total = Decimal("0")
    wave_total = Decimal("0")
    rolling_types = _empty_type_totals(ROLLING_LABELS)
    wave_types = _empty_type_totals(WAVE_LABELS)

    for row_index in range(2, sheet.max_row + 1):
        line = str(sheet.cell(row_index, headers["线体"]).value or "").strip()
        order_qty = _to_decimal(sheet.cell(row_index, headers["订单数"]).value)
        type_name = str(sheet.cell(row_index, headers["类型"]).value or "").strip()

        if _is_wave_line(line):
            wave_total += order_qty
            if type_name:
                wave_types[type_name] = wave_types.get(type_name, Decimal("0")) + order_qty
        else:
            rolling_total += order_qty
            if type_name:
                rolling_types[type_name] = rolling_types.get(type_name, Decimal("0")) + order_qty

    return {
        "rolling_total": rolling_total,
        "wave_total": wave_total,
        "rolling_types": rolling_types,
        "wave_types": wave_types,
    }


def _empty_type_totals(labels: list[str]) -> dict[str, Decimal]:
    return {label: Decimal("0") for label in labels if label}


def _rolling_value_for_label(label: str | None, rolling_types: dict[str, Decimal]):
    if not label:
        return None
    if label == "SKD总数(包含烘干)":
        value = rolling_types.get("SKD", Decimal("0")) + rolling_types.get("SKD烘干", Decimal("0"))
    elif label == "内销":
        value = rolling_types.get("内销", Decimal("0")) + rolling_types.get(DOMESTIC_TYPE, Decimal("0"))
    else:
        value = rolling_types.get(label, Decimal("0"))
    return _number_or_int(value)


def _wave_value_for_label(label: str | None, wave_types: dict[str, Decimal]):
    if not label:
        return None
    if label == WAVE_P7P9_LABEL:
        value = wave_types.get(WAVE_P7P9_LABEL, Decimal("0")) + wave_types.get(WAVE_P7P9_TYPE, Decimal("0"))
    else:
        value = wave_types.get(label, Decimal("0"))
    return _number_or_int(value)


def _wave_decomposition_total(wave_types: dict[str, Decimal]) -> Decimal:
    total = Decimal("0")
    for label in WAVE_LABELS:
        if label == WAVE_P7P9_LABEL:
            total += wave_types.get(WAVE_P7P9_LABEL, Decimal("0")) + wave_types.get(WAVE_P7P9_TYPE, Decimal("0"))
        else:
            total += wave_types.get(label, Decimal("0"))
    return total


def _rolling_decomposition_total(rolling_types: dict[str, Decimal]) -> Decimal:
    total = Decimal("0")
    for label in ROLLING_LABELS:
        if label == "SKD烘干":
            continue
        if label == "SKD总数(包含烘干)":
            total += rolling_types.get("SKD", Decimal("0")) + rolling_types.get("SKD烘干", Decimal("0"))
        elif label == "内销":
            total += rolling_types.get("内销", Decimal("0")) + rolling_types.get(DOMESTIC_TYPE, Decimal("0"))
        else:
            total += rolling_types.get(label, Decimal("0"))
    return total


def _remaining_unclassified_rolling(sheet: Worksheet, headers: dict[str, int]) -> tuple[int, Decimal]:
    rows = 0
    total = Decimal("0")
    for row_index in range(2, sheet.max_row + 1):
        line = str(sheet.cell(row_index, headers["线体"]).value or "").strip()
        if _is_wave_line(line):
            continue
        type_name = str(sheet.cell(row_index, headers["类型"]).value or "").strip()
        if type_name:
            continue
        rows += 1
        total += _to_decimal(sheet.cell(row_index, headers["订单数"]).value)
    return rows, total


def _remaining_unclassified_wave(sheet: Worksheet, headers: dict[str, int]) -> tuple[int, Decimal]:
    rows = 0
    total = Decimal("0")
    for row_index in range(2, sheet.max_row + 1):
        line = str(sheet.cell(row_index, headers["线体"]).value or "").strip()
        if not _is_wave_line(line):
            continue
        type_name = str(sheet.cell(row_index, headers["类型"]).value or "").strip()
        if type_name:
            continue
        rows += 1
        total += _to_decimal(sheet.cell(row_index, headers["订单数"]).value)
    return rows, total


def _product_kind_from_material_code(value: Any) -> str:
    material_code = str(value or "").strip().upper()
    if not material_code:
        return "unknown"
    marker_index = material_code.find("U")
    body = material_code[marker_index + 1 :] if marker_index >= 0 else material_code
    if body.startswith("605"):
        return "dryer"
    if body.startswith("60101"):
        return "wash"
    if body.startswith("60102"):
        return "dry"
    return "unknown"


def _is_t7p7t5p5_dreame_description(description: str) -> bool:
    text = str(description or "").strip()
    if not text:
        return False
    upper_text = text.upper()
    first_section = upper_text.split("/", 1)[0].strip()
    first_token = first_section.split()[0] if first_section.split() else first_section
    if any(series in first_token for series in ("T7", "P7", "T5", "P5")):
        return True
    return "追觅" in text or "DREAME" in upper_text or "DREMA" in upper_text


def _is_t9p9_description(description: str) -> bool:
    first_token = _description_first_token(description)
    return _contains_series_token(first_token, ("T9", "P9"))


def _is_c6_description(description: str) -> bool:
    return "C6" in _description_first_token(description)


def _is_wave_p7p9_candidate(sheet_metal: str, description: str) -> bool:
    sheet_metal_text = str(sheet_metal or "").upper()
    if "P7" in sheet_metal_text or "P9" in sheet_metal_text:
        return True
    return _contains_series_token(_description_first_token(description), ("P7", "P9"))


def _is_wave_domestic_inverter_description(description: str) -> bool:
    text = str(description or "").upper()
    return "DMP" in text or "BMP" in text or "DP" in text


def _description_first_token(description: str) -> str:
    text = str(description or "").strip().upper()
    if not text:
        return ""
    first_section = text.split("/", 1)[0].strip()
    return first_section.split()[0] if first_section.split() else first_section


def _contains_series_token(token: str, series_values: tuple[str, ...]) -> bool:
    for series in series_values:
        start = 0
        while True:
            index = token.find(series, start)
            if index < 0:
                break
            next_index = index + len(series)
            if next_index >= len(token) or not token[next_index].isdigit():
                return True
            start = index + 1
    return False


def _set_font_for_fill(cell, fill_rgb: str) -> None:
    font = copy(cell.font)
    font.color = "FFFFFF" if _is_dark_color(fill_rgb) else "000000"
    cell.font = font


def _is_dark_color(rgb: str) -> bool:
    value = rgb[-6:]
    r = int(value[0:2], 16)
    g = int(value[2:4], 16)
    b = int(value[4:6], 16)
    luminance = 0.2126 * r + 0.7152 * g + 0.0722 * b
    return luminance < 140


def _header_map(sheet: Worksheet) -> dict[str, int]:
    raw_headers = [normalize_header(cell.value) for cell in sheet[1]]
    headers = deduplicate_headers(raw_headers)
    return {header: index + 1 for index, header in enumerate(headers)}


def _require_columns(header_map: dict[str, int], columns: list[str], context: str) -> None:
    missing = [column for column in columns if column not in header_map]
    if missing:
        raise MissingRequiredFieldsError(f"{context}缺少关键字段：" + "、".join(missing))


def _is_wave_line(line: str) -> bool:
    return line in WAVE_LINES


def _to_decimal(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _to_decimal_or_none(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace(",", ""))
    except (InvalidOperation, ValueError):
        return None


def _number_or_int(value: Decimal):
    return int(value) if value == value.to_integral_value() else float(value)


def _move_sheet_after(workbook: Workbook, sheet_name: str, anchor_sheet_name: str) -> None:
    sheet = workbook[sheet_name]
    remaining = [item for item in workbook._sheets if item is not sheet]
    anchor_index = next(index for index, item in enumerate(remaining) if item.title == anchor_sheet_name)
    workbook._sheets = remaining[: anchor_index + 1] + [sheet] + remaining[anchor_index + 1 :]
