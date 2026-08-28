import unittest

from openpyxl import Workbook

from dingbian_classifier.decomposition import mark_missing_coefficient_rows_for_skip
from dingbian_classifier.logger import ProcessingLogger


class SkipMissingCoefficientRowsTest(unittest.TestCase):
    def test_marks_only_missing_coefficient_rows_to_exclude_them_from_decomposition(self):
        formula_workbook = Workbook()
        formula_sheet = formula_workbook.active
        formula_sheet.append(["基本开始日期", "备注", "类型", "系数"])
        formula_sheet.append([None, "订单A", None, 1.5])
        formula_sheet.append([None, "订单B", None, '=IFERROR(VLOOKUP(A3,系数查询表!A:B,2,0),"#N/A")'])

        values_workbook = Workbook()
        values_sheet = values_workbook.active
        values_sheet.append(["基本开始日期", "备注", "类型", "系数"])
        values_sheet.append([None, "订单A", None, 1.5])
        values_sheet.append([None, "订单B", None, "#N/A"])

        skipped = mark_missing_coefficient_rows_for_skip(
            formula_workbook,
            values_workbook,
            formula_sheet.title,
            ProcessingLogger(),
        )

        self.assertEqual(skipped, 1)
        self.assertIsNone(formula_sheet.cell(2, 1).fill.fill_type)
        self.assertEqual(formula_sheet.cell(3, 1).fill.fill_type, "solid")
        self.assertEqual(formula_sheet.cell(3, 3).fill.fill_type, "solid")

if __name__ == "__main__":
    unittest.main()
