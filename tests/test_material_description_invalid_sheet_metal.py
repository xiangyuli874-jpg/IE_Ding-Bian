import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from dingbian_classifier.logger import ProcessingLogger
from dingbian_classifier.material_description import (
    fill_material_descriptions_for_invalid_sheet_metal,
)


def _workbook_with_main_rows() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "主数据"
    sheet.append(["物料编码", "钣金型号", "物料描述"])
    sheet.append(["U60101000001", "#N/A", "旧描述一"])
    sheet.append(["U60101000002", 0, "旧描述二"])
    sheet.append(["U60101000003", "有效钣金", "不可覆盖"])
    sheet.append(["U60101000004", "N/A", "待匹配"])
    return workbook


def _write_lookup(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["物料编码", "物料描述"])
    sheet.append(["U60101000001", "回填描述一"])
    sheet.append(["U60101000002", "回填描述二"])
    sheet.append(["U60101000003", "不应写入"])
    workbook.save(path)


class InvalidSheetMetalDescriptionFillTest(unittest.TestCase):
    def test_overwrites_only_matched_invalid_rows(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            lookup_path = Path(temporary_directory) / "物料描述补充表.xlsx"
            _write_lookup(lookup_path)
            formula_workbook = _workbook_with_main_rows()
            values_workbook = _workbook_with_main_rows()

            result = fill_material_descriptions_for_invalid_sheet_metal(
                formula_workbook,
                values_workbook,
                "主数据",
                lookup_path,
                ProcessingLogger(),
            )

        formula_sheet = formula_workbook["主数据"]
        values_sheet = values_workbook["主数据"]
        self.assertEqual(formula_sheet.cell(2, 3).value, "回填描述一")
        self.assertEqual(formula_sheet.cell(3, 3).value, "回填描述二")
        self.assertEqual(formula_sheet.cell(4, 3).value, "不可覆盖")
        self.assertEqual(formula_sheet.cell(5, 3).value, "待匹配")
        self.assertEqual(values_sheet.cell(2, 3).value, "回填描述一")
        self.assertEqual(values_sheet.cell(3, 3).value, "回填描述二")
        self.assertEqual(result.filled_rows, 2)
        self.assertEqual(result.remaining_rows, 1)
