import unittest

from openpyxl import Workbook

from dingbian_classifier.decomposition import (
    MATERIAL_CODE_AUDIT_SHEET_NAME,
    audit_and_correct_material_code_types,
)
from dingbian_classifier.logger import ProcessingLogger


HEADERS = ["物料编码", "物料描述", "备注", "线体", "订单数", "钣金型号", "类型"]


def _workbook(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "主数据"
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)
    return workbook


class MaterialCodeTypeAuditTest(unittest.TestCase):
    def test_corrects_deterministic_conflicts_and_lists_unresolved_conflicts(self):
        rows = [
            ["Z3U60501080000", "DWD10L9ANACN 觅曜蓝", "追觅", "E线", 50, "T10热泵干衣机", "普通内销"],
            ["Z4U60102080139", "TWF120-E143021DA05e", "", "E线", 152, "#N/A", "外销"],
            ["Z4U60101080001", "洗衣机", "", "A线", 10, "有效钣金", "普通烘干"],
            ["Z4U60101080002", "热泵洗烘一体机", "", "E线", 20, "有效钣金", "热泵洗烘一体机"],
        ]
        formula_workbook = _workbook(rows)
        values_workbook = _workbook(rows)

        result = audit_and_correct_material_code_types(
            formula_workbook,
            values_workbook,
            "主数据",
            ProcessingLogger(),
        )

        sheet = formula_workbook["主数据"]
        self.assertEqual(sheet.cell(2, 7).value, "T10/P10干衣机")
        self.assertEqual(sheet.cell(3, 7).value, "普通烘干")
        self.assertEqual(sheet.cell(4, 7).value, "普通烘干")
        self.assertEqual(result.corrected_rows, 2)
        self.assertEqual(result.review_rows, 1)

        review_sheet = formula_workbook[MATERIAL_CODE_AUDIT_SHEET_NAME]
        self.assertEqual(review_sheet.max_row, 2)
        self.assertEqual(review_sheet.cell(2, 2).value, "Z4U60101080001")
        self.assertIn("单洗", review_sheet.cell(2, 10).value)

    def test_prioritizes_material_code_for_domestic_dry_and_c6_dryer(self):
        rows = [
            ["Z1U60102000142", "G100S3-HD", "", "E线", 100, "有效钣金", "普通内销"],
            ["Z4U60501080131", "THP90-G06H", "", "H线", 1, "C6热泵干衣机", "外销"],
        ]
        formula_workbook = _workbook(rows)
        values_workbook = _workbook(rows)

        result = audit_and_correct_material_code_types(
            formula_workbook,
            values_workbook,
            "主数据",
            ProcessingLogger(),
        )

        sheet = formula_workbook["主数据"]
        self.assertEqual(sheet.cell(2, 7).value, "普通烘干")
        self.assertEqual(sheet.cell(3, 7).value, "C6热泵干衣机")
        self.assertEqual(result.corrected_rows, 2)
        self.assertEqual(result.review_rows, 0)
