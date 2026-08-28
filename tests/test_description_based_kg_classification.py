# -*- coding: utf-8 -*-
import unittest

from openpyxl import Workbook

from dingbian_classifier.decomposition import write_extra_order_summary
from dingbian_classifier.logger import ProcessingLogger


HEADERS = ["线体", "钣金型号", "订单数", "类型", "物料描述", "备注", "标台数", "渠道"]


def build_workbooks(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "W1-W2周排产明细"
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)

    values_workbook = Workbook()
    values_sheet = values_workbook.active
    values_sheet.title = sheet.title
    values_sheet.append(HEADERS)
    for row in rows:
        values_sheet.append(row)
    return workbook, values_workbook, sheet.title


class DescriptionBasedKgClassificationTest(unittest.TestCase):
    def test_summary_writeback_uses_model_description_when_metal_is_missing(self):
        rows = [
            ["A线夜", "#N/A", 419, "普通内销", "TG-V8010BG 基准产品", "", 1000, "CBG"],
            ["A线", "#N/A", 200, "普通内销", "G100V3R-B/朗月白/国标插头/单洗", "", 500, "CBG"],
            ["D线", "#N/A", 100, "外销", "TWF80-P123021DA05e", "", 200, "海外ODI"],
            ["D线夜", "#N/A", 100, "外销", "TWF100-E143021DA05e", "", 200, "海外ODI"],
            ["D线", "#N/A", 100, "外销", "TWF120-L14709BA05e", "", 200, "海外ODI"],
            ["D线夜", "#N/A", 100, "外销", "WF75F1-14BU-AUA05eA", "", 200, "海外ODI"],
            ["E线", "#N/A", 153, "外销", "TWF140-P14707DB81e-CL", "", 300, "海外ODI"],
        ]
        workbook, values_workbook, sheet_name = build_workbooks(rows)

        write_extra_order_summary(workbook, values_workbook, sheet_name, ProcessingLogger())

        result_types = [workbook[sheet_name].cell(row, 4).value for row in range(2, 9)]
        self.assertEqual(
            result_types,
            [
                "普通内销6、7、8kg",
                "普通内销9、10kg",
                "外销6、7、8kg",
                "外销9、10kg",
                "外销12kg",
                "外销6、7、8kg",
                "外销12kg",
            ],
        )

    def test_summary_writeback_creates_heat_pump_washer_dryer_type(self):
        rows = [
            [
                "E线",
                "#N/A",
                150,
                "普通内销",
                "WH120V3W-14DIWB4e 星曜灰",
                "新品热泵洗烘一体机",
                300,
                "CBG",
            ],
        ]
        workbook, values_workbook, sheet_name = build_workbooks(rows)

        write_extra_order_summary(workbook, values_workbook, sheet_name, ProcessingLogger())

        self.assertEqual(workbook[sheet_name].cell(2, 4).value, "热泵洗烘一体机")

    def test_decomposition_detail_includes_kg_segment_types_in_rolling_total(self):
        rows = [
            ["A线", "#N/A", 100, "普通内销6、7、8kg", "TG-V80", "", 100, "CBG"],
            ["A线", "#N/A", 100, "普通内销9、10kg", "G100", "", 100, "CBG"],
            ["D线", "#N/A", 100, "外销6、7、8kg", "TWF80", "", 100, "海外ODI"],
            ["D线", "#N/A", 100, "外销9、10kg", "TWF100", "", 100, "海外ODI"],
            ["E线", "#N/A", 100, "外销12kg", "TWF140", "", 100, "海外ODI"],
        ]
        workbook, values_workbook, sheet_name = build_workbooks(rows)

        write_extra_order_summary(workbook, values_workbook, sheet_name, ProcessingLogger())

        detail_sheet = workbook["排单分解表明细"]
        rolling_values = {
            detail_sheet.cell(row, 1).value: detail_sheet.cell(row, 2).value
            for row in range(5, detail_sheet.max_row + 1)
            if detail_sheet.cell(row, 1).value
        }
        kg_labels = [
            "普通内销6、7、8kg",
            "普通内销9、10kg",
            "外销6、7、8kg",
            "外销9、10kg",
            "外销12kg",
        ]

        self.assertTrue(set(kg_labels).issubset(rolling_values))
        self.assertEqual([rolling_values[label] for label in kg_labels], [100, 100, 100, 100, 100])
        self.assertEqual(sum(rolling_values[label] for label in kg_labels), detail_sheet["B3"].value)


if __name__ == "__main__":
    unittest.main()
