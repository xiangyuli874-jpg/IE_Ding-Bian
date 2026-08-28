import unittest

from openpyxl import Workbook

from dingbian_classifier.coefficients import apply_manual_coefficients
from dingbian_classifier.logger import ProcessingLogger


class ManualCoefficientApplyTest(unittest.TestCase):
    def _make_workbooks(self, cached_value):
        formula_workbook = Workbook()
        main = formula_workbook.active
        main.title = "W2632-W2635周排产明细"
        main.append(["物料编码", "系数"])
        main.append(["Z4U60101080857", "#N/A"])
        supplement = formula_workbook.create_sheet("系数补充")
        supplement.append(["原始行号", "物料编码", "系数"])
        supplement.append([2, "Z4U60101080857", "=VLOOKUP(B2,外部表!A:B,2,0)"])

        values_workbook = Workbook()
        values_main = values_workbook.active
        values_main.title = main.title
        values_main.append(["物料编码", "系数"])
        values_main.append(["Z4U60101080857", "#N/A"])
        values_supplement = values_workbook.create_sheet("系数补充")
        values_supplement.append(["原始行号", "物料编码", "系数"])
        values_supplement.append([2, "Z4U60101080857", cached_value])
        return formula_workbook, values_workbook, main

    def test_applies_numeric_cached_value_from_lookup_formula(self):
        formula_workbook, values_workbook, main = self._make_workbooks(2.723)

        result = apply_manual_coefficients(
            formula_workbook,
            values_workbook,
            main.title,
            ProcessingLogger(),
        )

        self.assertEqual(result.applied_rows, 1)
        self.assertEqual(result.remaining_rows, 0)
        self.assertEqual(main.cell(2, 2).value, 2.723)

    def test_keeps_formula_with_error_cache_unresolved(self):
        formula_workbook, values_workbook, main = self._make_workbooks("#N/A")

        result = apply_manual_coefficients(
            formula_workbook,
            values_workbook,
            main.title,
            ProcessingLogger(),
        )

        self.assertEqual(result.applied_rows, 0)
        self.assertEqual(result.remaining_rows, 1)
        self.assertEqual(main.cell(2, 2).value, "#N/A")


if __name__ == "__main__":
    unittest.main()
