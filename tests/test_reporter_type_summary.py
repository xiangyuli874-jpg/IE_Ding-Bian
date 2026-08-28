import unittest

from openpyxl import Workbook

from dingbian_classifier.logger import ProcessingLogger
from dingbian_classifier.reporter import write_type_summary_results


class TypeSummaryReporterTest(unittest.TestCase):
    def test_summarizes_existing_main_sheet_types_without_keyword_rules(self):
        workbook = Workbook()
        headers = ["类型", "订单数", "标台数", "线体"]
        rows = [
            {"类型": "普通内销", "订单数": 10, "标台数": 15, "线体": "A线"},
            {"类型": "普通内销", "订单数": 20, "标台数": 30, "线体": "D线"},
            {"类型": "外销铁皮", "订单数": 5, "标台数": 8, "线体": "H线"},
            {"类型": "", "订单数": 7, "标台数": 9, "线体": "E线"},
        ]

        write_type_summary_results(workbook, headers, rows, ProcessingLogger())

        summary = workbook["分类结果汇总表"]
        self.assertEqual(
            list(summary.iter_rows(values_only=True)),
            [
                ("类型", "订单数合计", "标台数合计", "行数"),
                ("外销铁皮", 5.0, 8.0, 1),
                ("普通内销", 30.0, 45.0, 2),
            ],
        )
        unmatched = workbook["未分类数据"]
        self.assertEqual(list(unmatched.iter_rows(values_only=True)), [tuple(headers), ("", 7, 9, "E线")])
        self.assertNotIn("分类规则配置", workbook.sheetnames)


if __name__ == "__main__":
    unittest.main()
