import unittest

from openpyxl import Workbook

from dingbian_classifier.coefficients import (
    cleanup_order_rows,
    cleanup_order_rows_preserve_blank_orders,
    prepare_coefficients,
)
from dingbian_classifier.logger import ProcessingLogger


class CleanupOrderRowsTest(unittest.TestCase):
    def test_deletes_blank_order_blank_line_and_excluded_material_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["订单数", "物料编码", "线体"])
        sheet.append([100, "KEEP-1", "A线"])
        sheet.append([None, "BLANK-ORDER", "A线"])
        sheet.append([200, "BLANK-LINE", None])
        sheet.append([300, "Z4U6010100", "C线"])
        sheet.append([400, "KEEP-2", "H线"])

        result = cleanup_order_rows(sheet, ProcessingLogger())

        self.assertEqual(result.deleted_blank_order_rows, 1)
        self.assertEqual(result.deleted_blank_line_rows, 1)
        self.assertEqual(result.deleted_excluded_material_rows, 1)
        self.assertEqual(
            [sheet.cell(row_index, 2).value for row_index in range(2, sheet.max_row + 1)],
            ["KEEP-1", "KEEP-2"],
        )

    def test_prepare_coefficients_can_preserve_blank_order_rows(self):
        formula_workbook = Workbook()
        formula_sheet = formula_workbook.active
        formula_sheet.append(["订单数", "物料编码", "线体", "系数"])
        formula_sheet.append([None, "BLANK-ORDER", "A线", "#N/A"])
        formula_sheet.append([100, "KEEP-1", "A线", "#N/A"])

        values_workbook = Workbook()
        values_sheet = values_workbook.active
        values_sheet.append(["订单数", "物料编码", "线体", "系数"])
        values_sheet.append([None, "BLANK-ORDER", "A线", "#N/A"])
        values_sheet.append([100, "KEEP-1", "A线", "#N/A"])

        stale_missing = formula_workbook.create_sheet("系数仍缺失")
        stale_missing.append(["原始行号", "物料编码", "系数"])
        stale_missing.append([999, "STALE", "#N/A"])

        result = prepare_coefficients(
            formula_workbook,
            values_workbook,
            formula_sheet.title,
            ProcessingLogger(),
            cleanup_rows=False,
        )

        self.assertEqual(result.deleted_blank_order_rows, 0)
        self.assertEqual(formula_sheet.max_row, 3)
        supplement = formula_workbook["系数补充"]
        self.assertEqual(supplement.max_row, 3)
        self.assertNotIn("系数仍缺失", formula_workbook.sheetnames)

    def test_preserve_blank_order_cleanup_removes_only_other_foundation_exceptions(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["订单数", "物料编码", "线体"])
        sheet.append([None, "BLANK-ORDER", "A线"])
        sheet.append([100, "BLANK-LINE", None])
        sheet.append([100, "Z4U6010100", "A线"])
        sheet.append([100, "Z4U6010108", "A线"])
        sheet.append([100, "Z4U60501080", "A线"])
        sheet.append([100, "KEEP", "A线"])

        result = cleanup_order_rows_preserve_blank_orders(sheet, ProcessingLogger())

        self.assertEqual(result.deleted_blank_order_rows, 0)
        self.assertEqual(result.deleted_blank_line_rows, 1)
        self.assertEqual(result.deleted_excluded_material_rows, 3)
        self.assertEqual(
            [sheet.cell(row_index, 2).value for row_index in range(2, sheet.max_row + 1)],
            ["BLANK-ORDER", "KEEP"],
        )


if __name__ == "__main__":
    unittest.main()
