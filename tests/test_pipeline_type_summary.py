import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from dingbian_classifier.pipeline import run


class TypeSummaryPipelineTest(unittest.TestCase):
    def test_classify_writes_summary_from_existing_type_without_rule_configuration(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "input.xlsx"
            output_dir = Path(directory) / "output"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "W2632-W2635周排产明细"
            sheet.append(["物料编码", "物料描述", "钣金型号", "线体", "订单数", "标台数", "类型"])
            sheet.append(["Z4U60101000001", "型号A", "500滚筒", "A线", 10, 20, "普通内销"])
            sheet.append(["Z4U60101000002", "型号B", "500滚筒", "A线", 5, 10, "普通内销"])
            sheet.append(["Z4U60101000003", "型号C", "500滚筒", "H线", 7, 14, ""])
            workbook.save(source)

            with patch.dict(os.environ, {"DINGBIAN_SKIP_EXCEL_RESAVE": "1"}):
                result_path = run(source, output_dir, stage="classify")

            result = load_workbook(result_path, data_only=True)
            self.assertEqual(
                list(result["分类结果汇总表"].iter_rows(values_only=True)),
                [("类型", "订单数合计", "标台数合计", "行数"), ("普通内销", 15, 30, 2)],
            )
            self.assertNotIn("分类规则配置", result.sheetnames)
            self.assertEqual(result["未分类数据"].max_row, 2)


if __name__ == "__main__":
    unittest.main()
