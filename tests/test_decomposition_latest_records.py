import unittest

from openpyxl import Workbook

from dingbian_classifier.decomposition import (
    DETAIL_SHEET_NAME,
    LEGACY_LINE_CLASSIFICATION_SHEET_NAME,
    LINE_CLASSIFICATION_SHEET_NAME,
    WAVE_PLASTIC_DOMESTIC_TYPE,
    write_line_classification_detail_sheet,
)
from dingbian_classifier.logger import ProcessingLogger


class DecompositionLatestRecordsTest(unittest.TestCase):
    def test_line_classification_uses_current_sheet_name(self):
        workbook = Workbook()
        main_sheet = workbook.active
        main_sheet.title = "W1-W2周排产明细"
        main_sheet.append(["线体", "订单数", "类型", "标台数"])
        main_sheet.append(["B线", 10, WAVE_PLASTIC_DOMESTIC_TYPE, 12.5])
        workbook.create_sheet(DETAIL_SHEET_NAME)
        workbook.create_sheet(LEGACY_LINE_CLASSIFICATION_SHEET_NAME)

        values_workbook = Workbook()
        values_sheet = values_workbook.active
        values_sheet.title = main_sheet.title
        values_sheet.append(["线体", "订单数", "类型", "标台数"])
        values_sheet.append(["B线", 10, WAVE_PLASTIC_DOMESTIC_TYPE, 12.5])

        headers = {"线体": 1, "订单数": 2, "类型": 3, "标台数": 4}
        write_line_classification_detail_sheet(
            workbook,
            values_workbook,
            main_sheet.title,
            headers,
            ProcessingLogger(),
        )

        self.assertIn(LINE_CLASSIFICATION_SHEET_NAME, workbook.sheetnames)
        self.assertNotIn(LEGACY_LINE_CLASSIFICATION_SHEET_NAME, workbook.sheetnames)
        self.assertEqual(workbook[LINE_CLASSIFICATION_SHEET_NAME].freeze_panes, "B4")
        self.assertEqual(WAVE_PLASTIC_DOMESTIC_TYPE, "塑料机")


if __name__ == "__main__":
    unittest.main()
