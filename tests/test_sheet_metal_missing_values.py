import unittest

from openpyxl import Workbook

from dingbian_classifier.logger import ProcessingLogger
from dingbian_classifier.sheet_metal import create_sheet_metal_supplement_sheet


class SheetMetalMissingValuesTests(unittest.TestCase):
    def test_dash_and_blank_sheet_metal_values_are_added_to_supplement(self):
        formula_workbook = Workbook()
        formula_sheet = formula_workbook.active
        formula_sheet.title = "W2632-W2635周排产明细"
        formula_sheet.append(["物料编码", "钣金型号"])
        formula_sheet.append(["A", "-"])
        formula_sheet.append(["B", None])
        formula_sheet.append(["C", "有效钣金"])

        values_workbook = Workbook()
        values_sheet = values_workbook.active
        values_sheet.title = formula_sheet.title
        values_sheet.append(["物料编码", "钣金型号"])
        values_sheet.append(["A", "-"])
        values_sheet.append(["B", None])
        values_sheet.append(["C", "有效钣金"])

        missing_count = create_sheet_metal_supplement_sheet(
            formula_workbook,
            formula_sheet,
            values_sheet,
            ProcessingLogger(),
        )

        supplement = formula_workbook["钣金型号补充"]
        self.assertEqual(missing_count, 2)
        self.assertEqual([supplement.cell(row, 2).value for row in (2, 3)], ["A", "B"])


if __name__ == "__main__":
    unittest.main()
