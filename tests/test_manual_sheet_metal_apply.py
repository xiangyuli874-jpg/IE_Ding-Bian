import unittest

from openpyxl import Workbook

from dingbian_classifier.logger import ProcessingLogger
from dingbian_classifier.sheet_metal import apply_manual_sheet_metal_models


class ManualSheetMetalApplyTest(unittest.TestCase):
    def test_applies_numeric_cached_value_from_lookup_formula(self):
        formula_workbook = Workbook()
        main = formula_workbook.active
        main.title = "W2632-W2635周排产明细"
        main.append(["物料编码", "钣金型号"])
        main.append(["Z4U60501080114", "#N/A"])
        supplement = formula_workbook.create_sheet("钣金型号补充")
        supplement.append(["原始行号", "物料编码", "钣金型号"])
        supplement.append([2, "Z4U60501080114", "=VLOOKUP(B2,外部表!A:B,2,0)"])

        values_workbook = Workbook()
        values_main = values_workbook.active
        values_main.title = main.title
        values_main.append(["物料编码", "钣金型号"])
        values_main.append(["Z4U60501080114", "#N/A"])
        values_supplement = values_workbook.create_sheet("钣金型号补充")
        values_supplement.append(["原始行号", "物料编码", "钣金型号"])
        values_supplement.append([2, "Z4U60501080114", "T10热泵干衣机极地灰PCM板/123筒107盖/95门板PCM板"])

        result = apply_manual_sheet_metal_models(
            formula_workbook,
            values_workbook,
            main.title,
            ProcessingLogger(),
        )

        self.assertEqual(result.applied_rows, 1)
        self.assertEqual(main.cell(2, 2).value, "T10热泵干衣机极地灰PCM板/123筒107盖/95门板PCM板")

    def test_deduplicates_remaining_rows_for_repeated_missing_codes(self):
        formula_workbook = Workbook()
        main = formula_workbook.active
        main.title = "W2632-W2635周排产明细"
        main.append(["物料编码", "钣金型号"])
        main.append(["Z4U60501080115", "#N/A"])
        supplement = formula_workbook.create_sheet("钣金型号补充")
        supplement.append(["原始行号", "物料编码", "钣金型号"])
        supplement.append([2, "Z4U60501080115", "#N/A"])
        supplement.append([2, "Z4U60501080115", "#N/A"])

        values_workbook = Workbook()
        values_main = values_workbook.active
        values_main.title = main.title
        values_main.append(["物料编码", "钣金型号"])
        values_main.append(["Z4U60501080115", "#N/A"])
        values_supplement = values_workbook.create_sheet("钣金型号补充")
        values_supplement.append(["原始行号", "物料编码", "钣金型号"])
        values_supplement.append([2, "Z4U60501080115", "#N/A"])
        values_supplement.append([2, "Z4U60501080115", "#N/A"])

        result = apply_manual_sheet_metal_models(
            formula_workbook,
            values_workbook,
            main.title,
            ProcessingLogger(),
        )

        self.assertEqual(result.applied_rows, 0)
        self.assertEqual(result.remaining_rows, 1)
        self.assertEqual(formula_workbook["钣金型号补充"].max_row, 2)


if __name__ == "__main__":
    unittest.main()
